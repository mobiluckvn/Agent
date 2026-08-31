"""Xuất tài liệu ra .docx / .pptx / .xlsx / .md / .pdf — không thêm phụ thuộc.

EAA-AIS-05 §8.5; NFR-04. Xem `docs/SAI_LECH_THIET_KE.md` mục SL-105.

Vì sao tự dựng OOXML thay vì dùng ``python-docx``
--------------------------------------------------

NFR-04 chốt sản phẩm chỉ phụ thuộc Python, toolchain và Git. ``python-docx`` và
``python-pptx`` là hai phụ thuộc mới cho một việc mà thư viện chuẩn làm được:
một tệp ``.docx`` là một tệp ZIP chứa vài tệp XML, và ``zipfile`` nằm sẵn trong
Python.

Đánh đổi phải nói thẳng: bộ xuất này viết ra tập con **hẹp** của OOXML — tiêu
đề, đoạn, danh sách, bảng, khối mã. Không có ảnh, không có style tùy biến,
không có mục lục tự cập nhật. Đủ cho tài liệu thiết kế; không đủ nếu một ngày
cần bản trình bày có thương hiệu.

``.xlsx`` thì KHÁC: ``openpyxl`` đã là phụ thuộc dev sẵn có (dùng cho bảng năng
lực), và định dạng bảng tính có nhiều chỗ bẫy hơn hẳn — kiểu ô, chuỗi dùng
chung, tham chiếu. Tự dựng ở đó là chuốc lấy rủi ro để đổi lấy đúng số không.

``.pdf`` thì KHÔNG tự dựng
---------------------------

Sinh PDF có bố cục cần một bộ dàn trang: ngắt dòng, ngắt trang, đo chiều rộng
chữ theo font. Tự viết là viết một nửa bộ dàn trang, và một nửa bộ dàn trang
cho ra tài liệu **trông như đã hỏng** thay vì hỏng hẳn.

Nên đường PDF ở đây là: sinh ``.docx`` rồi nhờ LibreOffice chuyển. Nó là một
công cụ NGOÀI, và được đối xử đúng như mọi công cụ ngoài khác của sản phẩm này
— dò trước khi dùng, và nói rõ cách cài khi thiếu, thay vì hỏng lúc chạy.

Chuỗi trong XML
----------------

Mọi chỗ chèn chữ đều đi qua :func:`_thoat`. Một dấu ``&`` trong tên linh kiện
là đủ để tệp không mở được, và Word báo "tệp hỏng" chứ không báo dòng nào sai.
"""

from __future__ import annotations

import shutil
import subprocess
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

from eaa.docmodel import (
    Block,
    Bullets,
    Code,
    Doc,
    Heading,
    Note,
    PageBreak,
    Para,
    Table,
)

__all__ = [
    "OfficeError",
    "ThieuCongCu",
    "DINH_DANG",
    "write_docx",
    "write_pptx",
    "write_xlsx",
    "write_markdown",
    "write_pdf",
    "write",
]

#: Định dạng xuất được, kèm câu mô tả để in ra cho người chọn.
DINH_DANG: dict[str, str] = {
    "md": "Markdown — đọc được ngay trong terminal và trong Git diff",
    "docx": "Word — tự dựng OOXML, không thêm phụ thuộc",
    "xlsx": "Excel — qua openpyxl; hợp cho danh sách chức năng và ma trận truy vết",
    "pptx": "PowerPoint — tự dựng OOXML; mỗi tiêu đề mức 1 thành một slide",
    "pdf": "PDF — sinh .docx rồi nhờ LibreOffice chuyển (cần soffice trên máy)",
}

#: Tên tiến trình LibreOffice, theo thứ tự thử.
_SOFFICE = ("soffice", "libreoffice")
#: Chỗ macOS cài LibreOffice mà không đưa vào PATH.
_SOFFICE_MAC = "/Applications/LibreOffice.app/Contents/MacOS/soffice"


class OfficeError(Exception):
    """Không xuất được tài liệu."""


class ThieuCongCu(OfficeError):
    """Thiếu một công cụ ngoài. Khác lỗi — đây là chuyện cài đặt."""


def _thoat(s: str) -> str:
    """Thoát chuỗi cho XML.

    Không dùng ``xml.sax.saxutils.escape``: nó không bỏ ký tự điều khiển, mà
    một ký tự điều khiển lọt vào XML làm tệp không mở được — và trình mở chỉ
    báo "tệp hỏng", không báo ký tự nào.
    """
    s = "".join(c for c in str(s) if c in "\t\n\r" or ord(c) >= 0x20)
    return (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
             .replace('"', "&quot;"))


# ══════════════════════════════ .docx ══════════════════════════════

_DOCX_CONTENT_TYPES = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
<Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>
<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
</Types>"""

_DOCX_RELS = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
</Relationships>"""

_DOCX_DOC_RELS = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>"""


def _docx_styles() -> str:
    """Style tối thiểu: Normal, Heading1..4, Title, Code, Note.

    Định nghĩa tường minh chứ không dựa vào style mặc định của trình mở: cùng
    một tệp mở bằng Word, LibreOffice và Google Docs mà bố cục lệch nhau thì
    "tài liệu bàn giao" mất nghĩa.
    """
    phan = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
            '<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">',
            '<w:docDefaults><w:rPrDefault><w:rPr>'
            '<w:rFonts w:ascii="Calibri" w:hAnsi="Calibri" w:cs="Calibri"/>'
            '<w:sz w:val="22"/></w:rPr></w:rPrDefault></w:docDefaults>',
            '<w:style w:type="paragraph" w:default="1" w:styleId="Normal">'
            '<w:name w:val="Normal"/><w:pPr><w:spacing w:after="120"/></w:pPr></w:style>']
    co = {1: 32, 2: 26, 3: 24, 4: 22}
    for muc, sz in co.items():
        phan.append(
            f'<w:style w:type="paragraph" w:styleId="Heading{muc}">'
            f'<w:name w:val="heading {muc}"/><w:basedOn w:val="Normal"/>'
            f'<w:pPr><w:outlineLvl w:val="{muc - 1}"/>'
            f'<w:spacing w:before="{280 - muc * 40}" w:after="120"/></w:pPr>'
            f'<w:rPr><w:b/><w:sz w:val="{sz}"/><w:color w:val="1F3864"/></w:rPr></w:style>')
    phan.append(
        '<w:style w:type="paragraph" w:styleId="Title"><w:name w:val="Title"/>'
        '<w:basedOn w:val="Normal"/><w:pPr><w:spacing w:after="240"/></w:pPr>'
        '<w:rPr><w:b/><w:sz w:val="44"/><w:color w:val="1F3864"/></w:rPr></w:style>')
    phan.append(
        '<w:style w:type="paragraph" w:styleId="Code"><w:name w:val="Code"/>'
        '<w:basedOn w:val="Normal"/><w:pPr><w:spacing w:after="0"/>'
        '<w:ind w:left="360"/></w:pPr>'
        '<w:rPr><w:rFonts w:ascii="Consolas" w:hAnsi="Consolas"/><w:sz w:val="18"/></w:rPr></w:style>')
    phan.append(
        '<w:style w:type="paragraph" w:styleId="NoteStyle"><w:name w:val="Note"/>'
        '<w:basedOn w:val="Normal"/><w:pPr><w:ind w:left="360"/>'
        '<w:pBdr><w:left w:val="single" w:sz="18" w:space="8" w:color="C00000"/></w:pBdr>'
        '</w:pPr><w:rPr><w:i/><w:color w:val="7F1D1D"/></w:rPr></w:style>')
    phan.append('</w:styles>')
    return "".join(phan)


def _docx_core(doc: Doc) -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<cp:coreProperties '
        'xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        'xmlns:dc="http://purl.org/dc/elements/1.1/" '
        'xmlns:dcterms="http://purl.org/dc/terms/" '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        f'<dc:title>{_thoat(doc.title)}</dc:title>'
        f'<dc:subject>{_thoat(doc.subtitle)}</dc:subject>'
        f'<dc:creator>{_thoat(doc.author or "EAA")}</dc:creator>'
        f'<cp:lastModifiedBy>{_thoat(doc.author or "EAA")}</cp:lastModifiedBy>'
        + (f'<dcterms:created xsi:type="dcterms:W3CDTF">{_thoat(doc.created_at)}</dcterms:created>'
           if doc.created_at else "")
        + '</cp:coreProperties>'
    )


def _p(text: str, *, style: str = "", bold: bool = False) -> str:
    ppr = f'<w:pPr><w:pStyle w:val="{style}"/></w:pPr>' if style else ""
    rpr = "<w:rPr><w:b/></w:rPr>" if bold else ""
    return (f'<w:p>{ppr}<w:r>{rpr}'
            f'<w:t xml:space="preserve">{_thoat(text)}</w:t></w:r></w:p>')


def _p_bullet(text: str, *, numbered: bool) -> str:
    # Dùng thụt lề + ký tự dẫn thay vì numbering.xml: danh sách đánh số thật
    # của OOXML cần thêm một part và một bảng định nghĩa mức, và nó là chỗ dễ
    # sinh ra tệp không mở được nhất. Chữ dẫn cho kết quả nhìn giống hệt.
    return (f'<w:p><w:pPr><w:ind w:left="480" w:hanging="240"/>'
            f'<w:spacing w:after="40"/></w:pPr>'
            f'<w:r><w:t xml:space="preserve">{_thoat(text)}</w:t></w:r></w:p>')


def _o(b: Table) -> str:
    rong = 9360 // max(b.cols, 1)
    ra = ['<w:tbl><w:tblPr><w:tblStyle w:val="TableGrid"/>'
          '<w:tblW w:w="5000" w:type="pct"/><w:tblBorders>'
          + "".join(f'<w:{c} w:val="single" w:sz="4" w:color="AAAAAA"/>'
                    for c in ("top", "left", "bottom", "right", "insideH", "insideV"))
          + '</w:tblBorders></w:tblPr>']
    ra.append('<w:tblGrid>' + f'<w:gridCol w:w="{rong}"/>' * b.cols + '</w:tblGrid>')

    def hang(o: Sequence[str], dam: bool) -> str:
        d = list(o) + [""] * (b.cols - len(o))
        tc = "".join(
            '<w:tc><w:tcPr>'
            + (f'<w:tcW w:w="{rong}" w:type="dxa"/>')
            + ('<w:shd w:val="clear" w:fill="DEE7F5"/>' if dam else "")
            + '</w:tcPr>'
            + "".join(_p(dong, bold=dam) for dong in (str(x).split("\n") or [""]))
            + '</w:tc>' for x in d)
        return f'<w:tr>{tc}</w:tr>'

    if b.header:
        ra.append(hang(b.header, True))
    for r in b.rows:
        ra.append(hang(r, False))
    ra.append('</w:tbl>')
    if b.caption:
        ra.append(_p(b.caption))
    else:
        ra.append('<w:p/>')  # Word cần một đoạn sau bảng, nếu không hai bảng dính nhau.
    return "".join(ra)


def _khoi_docx(b: Block) -> str:
    if isinstance(b, Heading):
        return _p(b.text, style=f"Heading{min(max(b.level, 1), 4)}")
    if isinstance(b, Para):
        return _p(b.text)
    if isinstance(b, Bullets):
        return "".join(
            _p_bullet(f"{i}. {x}" if b.numbered else f"•  {x}", numbered=b.numbered)
            for i, x in enumerate(b.items, 1))
    if isinstance(b, Table):
        return _o(b)
    if isinstance(b, Code):
        return "".join(_p(d or " ", style="Code") for d in b.text.split("\n"))
    if isinstance(b, Note):
        nhan = f"[{b.level}] " if b.level else ""
        return _p(nhan + b.text, style="NoteStyle")
    if isinstance(b, PageBreak):
        return '<w:p><w:r><w:br w:type="page"/></w:r></w:p>'
    return ""


def write_docx(doc: Doc, path: str | Path) -> Path:
    """Ghi ``.docx``. Trả về đường dẫn đã ghi."""
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)

    than = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
            '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">',
            '<w:body>']
    if doc.title:
        than.append(_p(doc.title, style="Title"))
    if doc.subtitle:
        than.append(_p(doc.subtitle))
    than += [_khoi_docx(b) for b in doc.blocks]
    # sectPr phải là phần tử CUỐI trong body — sai chỗ là tệp không mở được.
    than.append('<w:sectPr><w:pgSz w:w="11906" w:h="16838"/>'
                '<w:pgMar w:top="1134" w:right="1134" w:bottom="1134" w:left="1418"/>'
                '</w:sectPr></w:body></w:document>')

    with zipfile.ZipFile(p, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", _DOCX_CONTENT_TYPES)
        z.writestr("_rels/.rels", _DOCX_RELS)
        z.writestr("docProps/core.xml", _docx_core(doc))
        z.writestr("word/_rels/document.xml.rels", _DOCX_DOC_RELS)
        z.writestr("word/styles.xml", _docx_styles())
        z.writestr("word/document.xml", "".join(than))
    return p


# ══════════════════════════════ .pptx ══════════════════════════════

_PPTX_CONTENT_TYPES_HEAD = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/ppt/presentation.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.presentation.main+xml"/>
<Override PartName="/ppt/slideMasters/slideMaster1.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slideMaster+xml"/>
<Override PartName="/ppt/slideLayouts/slideLayout1.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slideLayout+xml"/>
<Override PartName="/ppt/theme/theme1.xml" ContentType="application/vnd.openxmlformats-officedocument.theme+xml"/>"""

_PPTX_THEME = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<a:theme xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" name="EAA">
<a:themeElements>
<a:clrScheme name="EAA"><a:dk1><a:sysClr val="windowText" lastClr="000000"/></a:dk1>
<a:lt1><a:sysClr val="window" lastClr="FFFFFF"/></a:lt1>
<a:dk2><a:srgbClr val="1F3864"/></a:dk2><a:lt2><a:srgbClr val="EEECE1"/></a:lt2>
<a:accent1><a:srgbClr val="1F3864"/></a:accent1><a:accent2><a:srgbClr val="C0504D"/></a:accent2>
<a:accent3><a:srgbClr val="9BBB59"/></a:accent3><a:accent4><a:srgbClr val="8064A2"/></a:accent4>
<a:accent5><a:srgbClr val="4BACC6"/></a:accent5><a:accent6><a:srgbClr val="F79646"/></a:accent6>
<a:hlink><a:srgbClr val="0000FF"/></a:hlink><a:folHlink><a:srgbClr val="800080"/></a:folHlink></a:clrScheme>
<a:fontScheme name="EAA">
<a:majorFont><a:latin typeface="Calibri Light"/><a:ea typeface=""/><a:cs typeface=""/></a:majorFont>
<a:minorFont><a:latin typeface="Calibri"/><a:ea typeface=""/><a:cs typeface=""/></a:minorFont></a:fontScheme>
<a:fmtScheme name="EAA">
<a:fillStyleLst><a:solidFill><a:schemeClr val="phClr"/></a:solidFill>
<a:solidFill><a:schemeClr val="phClr"/></a:solidFill>
<a:solidFill><a:schemeClr val="phClr"/></a:solidFill></a:fillStyleLst>
<a:lnStyleLst><a:ln w="6350"><a:solidFill><a:schemeClr val="phClr"/></a:solidFill></a:ln>
<a:ln w="12700"><a:solidFill><a:schemeClr val="phClr"/></a:solidFill></a:ln>
<a:ln w="19050"><a:solidFill><a:schemeClr val="phClr"/></a:solidFill></a:ln></a:lnStyleLst>
<a:effectStyleLst><a:effectStyle><a:effectLst/></a:effectStyle>
<a:effectStyle><a:effectLst/></a:effectStyle>
<a:effectStyle><a:effectLst/></a:effectStyle></a:effectStyleLst>
<a:bgFillStyleLst><a:solidFill><a:schemeClr val="phClr"/></a:solidFill>
<a:solidFill><a:schemeClr val="phClr"/></a:solidFill>
<a:solidFill><a:schemeClr val="phClr"/></a:solidFill></a:bgFillStyleLst></a:fmtScheme>
</a:themeElements></a:theme>"""

_PPTX_MASTER = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<p:sldMaster xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
 xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main">
<p:cSld>
<p:bg><p:bgPr><a:solidFill><a:srgbClr val="FFFFFF"/></a:solidFill><a:effectLst/></p:bgPr></p:bg>
<p:spTree><p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr>
<p:grpSpPr/></p:spTree></p:cSld>
<p:clrMap bg1="lt1" tx1="dk1" bg2="lt2" tx2="dk2" accent1="accent1" accent2="accent2"
 accent3="accent3" accent4="accent4" accent5="accent5" accent6="accent6"
 hlink="hlink" folHlink="folHlink"/>
<p:sldLayoutIdLst><p:sldLayoutId id="2147483649" r:id="rId1"/></p:sldLayoutIdLst>
</p:sldMaster>"""

_PPTX_LAYOUT = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<p:sldLayout xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
 xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main" type="blank">
<p:cSld><p:spTree><p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr>
<p:grpSpPr/></p:spTree></p:cSld>
<p:clrMapOvr><a:overrideClrMapping bg1="lt1" tx1="dk1" bg2="lt2" tx2="dk2"
 accent1="accent1" accent2="accent2" accent3="accent3" accent4="accent4"
 accent5="accent5" accent6="accent6" hlink="hlink" folHlink="folHlink"/></p:clrMapOvr>
</p:sldLayout>"""


def _pptx_o_chu(id_: int, ten: str, x: int, y: int, cx: int, cy: int,
                dong: Sequence[tuple[str, int, bool]]) -> str:
    """Một ô chữ. ``dong`` là dãy (chữ, cỡ phần trăm điểm, đậm)."""
    p = "".join(
        f'<a:p><a:pPr><a:lnSpc><a:spcPct val="100000"/></a:lnSpc></a:pPr>'
        f'<a:r><a:rPr lang="vi-VN" sz="{sz}"{" b=\"1\"" if dam else ""} dirty="0"/>'
        f'<a:t>{_thoat(chu)}</a:t></a:r></a:p>'
        for chu, sz, dam in dong) or '<a:p><a:endParaRPr lang="vi-VN"/></a:p>'
    return (f'<p:sp><p:nvSpPr><p:cNvPr id="{id_}" name="{_thoat(ten)}"/>'
            f'<p:cNvSpPr txBox="1"/><p:nvPr/></p:nvSpPr>'
            f'<p:spPr><a:xfrm><a:off x="{x}" y="{y}"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm>'
            f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></p:spPr>'
            f'<p:txBody><a:bodyPr wrap="square"><a:normAutofit/></a:bodyPr>'
            f'<a:lstStyle/>{p}</p:txBody></p:sp>')


def _pptx_slide(tieu_de: str, dong: Sequence[tuple[str, int, bool]]) -> str:
    return ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<p:sld xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
            ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
            ' xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main">'
            '<p:cSld><p:spTree>'
            '<p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr>'
            '<p:grpSpPr/>'
            + _pptx_o_chu(2, "Tiêu đề", 685800, 457200, 10820400, 1000000,
                          [(tieu_de, 3200, True)])
            + _pptx_o_chu(3, "Nội dung", 685800, 1600200, 10820400, 4400000, dong)
            + '</p:spTree></p:cSld>'
            '<p:clrMapOvr><a:masterClrMapping/></p:clrMapOvr></p:sld>')


#: Trần dòng một slide. Quá số này thì cắt sang slide tiếp — chữ tràn ra ngoài
#: khung là lỗi chỉ thấy khi mở tệp, tức là lúc đã bàn giao.
DONG_MOI_SLIDE = 12


def _gom_slide(doc: Doc) -> list[tuple[str, list[tuple[str, int, bool]]]]:
    """Cắt tài liệu thành slide: mỗi tiêu đề mức 1–2 mở một slide.

    Slide đầu là slide bìa. Không có nó thì tiêu đề mức 1 đầu tiên đóng ngay
    slide đang mở và sinh ra một slide **trống chỉ có tiêu đề** — thứ ai xem
    cũng nghĩ là tệp hỏng.
    """
    slide: list[tuple[str, list[tuple[str, int, bool]]]] = []
    if doc.title:
        bia = [(doc.subtitle, 2000, False)] if doc.subtitle else []
        if doc.project:
            bia.append((f"Dự án: {doc.project}", 1600, False))
        if doc.created_at:
            bia.append((f"Dựng lúc {doc.created_at}", 1400, False))
        slide.append((doc.title, bia))

    hien: list[tuple[str, int, bool]] = []
    ten = doc.title or "Nội dung"

    def chot() -> None:
        # Slide rỗng bị bỏ: một tiêu đề không có nội dung dưới nó là một mục
        # trong tài liệu, không phải một slide.
        if hien:
            slide.append((ten, list(hien)))
        hien.clear()

    for b in doc.blocks:
        if isinstance(b, Heading) and b.level <= 2:
            chot()
            ten = b.text
        elif isinstance(b, Heading):
            hien.append((b.text, 1800, True))
        elif isinstance(b, Para):
            hien.append((b.text, 1600, False))
        elif isinstance(b, Bullets):
            hien += [(f"•  {x}", 1600, False) for x in b.items]
        elif isinstance(b, Note):
            nhan = f"[{b.level}] " if b.level else ""
            hien.append((nhan + b.text, 1400, False))
        elif isinstance(b, Table):
            if b.header:
                hien.append((" | ".join(b.header), 1400, True))
            hien += [(" | ".join(r), 1400, False) for r in b.rows]
        elif isinstance(b, Code):
            hien += [(d, 1200, False) for d in b.text.split("\n")]
        if len(hien) >= DONG_MOI_SLIDE:
            phu = hien[:DONG_MOI_SLIDE]
            con = hien[DONG_MOI_SLIDE:]
            slide.append((ten, phu))
            ten = f"{ten} (tiếp)"
            hien.clear()
            hien.extend(con)
    chot()
    # Một .pptx không có slide nào là một tệp PowerPoint từ chối mở.
    return slide or [(doc.title or "Trống", [("Tài liệu không có nội dung.", 1600, False)])]


def write_pptx(doc: Doc, path: str | Path) -> Path:
    """Ghi ``.pptx``. Mỗi tiêu đề mức 1–2 thành một slide."""
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    ds = _gom_slide(doc)

    ct = [_PPTX_CONTENT_TYPES_HEAD]
    for i in range(1, len(ds) + 1):
        ct.append(f'<Override PartName="/ppt/slides/slide{i}.xml" '
                  'ContentType="application/vnd.openxmlformats-officedocument.'
                  'presentationml.slide+xml"/>')
    ct.append('</Types>')

    id_slide = "".join(
        f'<p:sldId id="{255 + i}" r:id="rId{i + 1}"/>' for i in range(1, len(ds) + 1))
    presentation = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<p:presentation xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
        ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
        ' xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main">'
        '<p:sldMasterIdLst><p:sldMasterId id="2147483648" r:id="rId1"/></p:sldMasterIdLst>'
        f'<p:sldIdLst>{id_slide}</p:sldIdLst>'
        '<p:sldSz cx="12192000" cy="6858000"/><p:notesSz cx="6858000" cy="9144000"/>'
        '</p:presentation>')

    rels = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/'
            '2006/relationships/slideMaster" Target="slideMasters/slideMaster1.xml"/>']
    for i in range(1, len(ds) + 1):
        rels.append(f'<Relationship Id="rId{i + 1}" Type="http://schemas.openxmlformats.org/'
                    f'officeDocument/2006/relationships/slide" Target="slides/slide{i}.xml"/>')
    rels.append(f'<Relationship Id="rId{len(ds) + 2}" Type="http://schemas.openxmlformats.org/'
                'officeDocument/2006/relationships/theme" Target="theme/theme1.xml"/>')
    rels.append('</Relationships>')

    with zipfile.ZipFile(p, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", "".join(ct))
        z.writestr("_rels/.rels",
                   '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                   '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                   '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/'
                   '2006/relationships/officeDocument" Target="ppt/presentation.xml"/>'
                   '</Relationships>')
        z.writestr("ppt/presentation.xml", presentation)
        z.writestr("ppt/_rels/presentation.xml.rels", "".join(rels))
        z.writestr("ppt/theme/theme1.xml", _PPTX_THEME)
        z.writestr("ppt/slideMasters/slideMaster1.xml", _PPTX_MASTER)
        z.writestr("ppt/slideMasters/_rels/slideMaster1.xml.rels",
                   '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                   '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                   '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/'
                   '2006/relationships/slideLayout" Target="../slideLayouts/slideLayout1.xml"/>'
                   '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/'
                   '2006/relationships/theme" Target="../theme/theme1.xml"/>'
                   '</Relationships>')
        z.writestr("ppt/slideLayouts/slideLayout1.xml", _PPTX_LAYOUT)
        z.writestr("ppt/slideLayouts/_rels/slideLayout1.xml.rels",
                   '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                   '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                   '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/'
                   '2006/relationships/slideMaster" Target="../slideMasters/slideMaster1.xml"/>'
                   '</Relationships>')
        for i, (ten, dong) in enumerate(ds, 1):
            z.writestr(f"ppt/slides/slide{i}.xml", _pptx_slide(ten, dong))
            z.writestr(f"ppt/slides/_rels/slide{i}.xml.rels",
                       '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                       '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                       '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/'
                       '2006/relationships/slideLayout" Target="../slideLayouts/slideLayout1.xml"/>'
                       '</Relationships>')
    return p


# ══════════════════════════════ .xlsx ══════════════════════════════


def write_xlsx(doc: Doc, path: str | Path) -> Path:
    """Ghi ``.xlsx``: mỗi bảng một sheet, phần chữ vào sheet đầu.

    Qua ``openpyxl`` chứ không tự dựng — xem phần đầu module.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise ThieuCongCu(
            "Thiếu openpyxl. Nó là phụ thuộc dev của sản phẩm này: "
            'pip install -e ".[dev]"'
        ) from exc

    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tổng quan"
    dam = Font(bold=True)
    nen = PatternFill("solid", fgColor="DEE7F5")

    hang = 1
    if doc.title:
        ws.cell(hang, 1, doc.title).font = Font(bold=True, size=15)
        hang += 2
    for b in doc.blocks:
        if isinstance(b, Heading):
            hang += 1
            o = ws.cell(hang, 1, b.text)
            o.font = Font(bold=True, size=max(13 - b.level, 10))
            hang += 1
        elif isinstance(b, Para):
            ws.cell(hang, 1, b.text).alignment = Alignment(wrap_text=True, vertical="top")
            hang += 1
        elif isinstance(b, Bullets):
            for i, x in enumerate(b.items, 1):
                ws.cell(hang, 1, f"{i}. {x}" if b.numbered else f"•  {x}")
                hang += 1
        elif isinstance(b, Note):
            o = ws.cell(hang, 1, (f"[{b.level}] " if b.level else "") + b.text)
            o.font = Font(italic=True, color="7F1D1D")
            hang += 1
        elif isinstance(b, Code):
            for d in b.text.split("\n"):
                ws.cell(hang, 1, d).font = Font(name="Consolas", size=9)
                hang += 1

    # Mỗi bảng một sheet riêng: một bảng nhét vào sheet chữ thì mất hẳn cái lợi
    # duy nhất của .xlsx — lọc, sắp xếp, và đếm.
    dem: dict[str, int] = {}
    for b in doc.blocks:
        if not isinstance(b, Table) or not (b.header or b.rows):
            continue
        ten = (b.caption or "Bảng")[:28] or "Bảng"
        dem[ten] = dem.get(ten, 0) + 1
        if dem[ten] > 1:
            ten = f"{ten[:25]} {dem[ten]}"
        # openpyxl cấm các ký tự này trong tên sheet.
        for c in "[]:*?/\\":
            ten = ten.replace(c, "-")
        s = wb.create_sheet(ten[:31])
        r = 1
        if b.header:
            for c, x in enumerate(b.header, 1):
                o = s.cell(r, c, x)
                o.font, o.fill = dam, nen
                o.alignment = Alignment(wrap_text=True, vertical="top")
            s.freeze_panes = "A2"
            r += 1
        for row in b.rows:
            for c, x in enumerate(row, 1):
                s.cell(r, c, x).alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
        for c in range(1, b.cols + 1):
            do_dai = max(
                [len(str(b.header[c - 1])) if c <= len(b.header) else 0]
                + [len(str(row[c - 1])) if c <= len(row) else 0 for row in b.rows] or [10])
            s.column_dimensions[get_column_letter(c)].width = min(max(do_dai + 2, 12), 60)

    ws.column_dimensions["A"].width = 110
    wb.save(p)
    return p


# ══════════════════════════════ .md ══════════════════════════════


def write_markdown(doc: Doc, path: str | Path) -> Path:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(doc.render_markdown(), encoding="utf-8")
    return p


# ══════════════════════════════ .pdf ══════════════════════════════


def tim_soffice() -> str:
    """Đường dẫn LibreOffice, hoặc chuỗi rỗng."""
    for ten in _SOFFICE:
        d = shutil.which(ten)
        if d:
            return d
    return _SOFFICE_MAC if Path(_SOFFICE_MAC).is_file() else ""


def write_pdf(doc: Doc, path: str | Path, *, timeout_s: float = 180.0) -> Path:
    """Ghi ``.pdf`` bằng cách sinh ``.docx`` rồi nhờ LibreOffice chuyển.

    Không tự dựng PDF: xem phần đầu module. Thiếu LibreOffice thì ném
    :class:`ThieuCongCu` kèm cách cài — đây là chuyện cài đặt, không phải lỗi
    của tài liệu.
    """
    p = Path(path)
    soffice = tim_soffice()
    if not soffice:
        raise ThieuCongCu(
            "Chưa có LibreOffice trên máy này — nó là công cụ chuyển .docx sang "
            ".pdf.\n"
            "  macOS:  brew install --cask libreoffice\n"
            "  Linux:  sudo apt-get install -y libreoffice\n"
            "  Windows: winget install -e --id TheDocumentFoundation.LibreOffice\n"
            "Bản .docx vẫn xuất được ngay: đổi --format docx."
        )

    p.parent.mkdir(parents=True, exist_ok=True)
    tam = p.with_suffix(".docx")
    write_docx(doc, tam)
    try:
        kq = subprocess.run(
            [soffice, "--headless", "--convert-to", "pdf", "--outdir",
             str(p.parent), str(tam)],
            capture_output=True, text=True, timeout=timeout_s,
        )
    except subprocess.TimeoutExpired as exc:
        raise OfficeError(
            f"LibreOffice quá {timeout_s:.0f} giây chưa xong. Bản .docx đã ghi "
            f"tại {tam}."
        ) from exc

    sinh_ra = p.parent / (tam.stem + ".pdf")
    if not sinh_ra.is_file():
        raise OfficeError(
            f"LibreOffice không sinh ra PDF (mã {kq.returncode}).\n"
            f"{(kq.stderr or kq.stdout or '').strip()[:400]}\n"
            f"Bản .docx đã ghi tại {tam}."
        )
    if sinh_ra != p:
        sinh_ra.replace(p)
    return p


# ══════════════════════════════ điều phối ══════════════════════════════


def write(doc: Doc, path: str | Path, *, fmt: str = "") -> Path:
    """Ghi theo định dạng suy từ đuôi tệp, hoặc theo ``fmt`` nếu nêu."""
    p = Path(path)
    dinh_dang = (fmt or p.suffix.lstrip(".")).lower()
    bo = {"md": write_markdown, "markdown": write_markdown, "docx": write_docx,
          "xlsx": write_xlsx, "pptx": write_pptx, "pdf": write_pdf}
    if dinh_dang not in bo:
        raise OfficeError(
            f"Chưa xuất được định dạng {dinh_dang!r}. Đang có: "
            + ", ".join(sorted(DINH_DANG))
        )
    return bo[dinh_dang](doc, p)
