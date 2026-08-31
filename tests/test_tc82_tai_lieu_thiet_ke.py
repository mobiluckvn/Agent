"""TC-82 — sinh tài liệu thiết kế: URD, SRS, SDD, chức năng, luồng.

EAA-AIS-05 §8.5; NFR-04. Xem `docs/SAI_LECH_THIET_KE.md` mục SL-105.

Ba nhóm bài, theo ba cách hỏng khác nhau:

1. **Tệp không mở được.** Với một tài liệu bàn giao thì đây là hỏng hoàn toàn,
   và nó chỉ hiện ra sau khi đã gửi đi. Nên bài canh ở đây quét XML của mọi
   part trong gói, và canh riêng những ký tự làm vỡ XML.
2. **Tài liệu mở được nhưng nói sai.** Nguy hiểm hơn hẳn: một mục trống đọc
   như "không cần" trong khi thật ra là "chưa ai điền".
3. **Khuôn mẫu và mã trôi khỏi nhau.** Khuôn mẫu là dữ liệu, nên nó đổi được
   mà không ai chạy test — trừ khi có bài canh mối nối.
"""

from __future__ import annotations

import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import pytest

from eaa import office
from eaa.docmodel import Bullets, Code, Doc, Heading, Note, PageBreak, Para, Table
from eaa.designdoc import NGUON, SPEC_DIR, DesignDocError, build, list_specs, load_spec

GOC = Path(__file__).resolve().parents[1]


def _doc_day_du() -> Doc:
    d = Doc(title="Tiêu đề", subtitle="Phụ đề", project="du_an", author="EAA",
            created_at="2026-01-01T00:00:00Z")
    d.heading("Mục một", 1)
    d.para('Ký tự khó: & < > " và tiếng Việt có dấu: Cân Bằng, Đủ, Ưu.')
    d.bullets(["một", "hai & ba", "<bốn>"])
    d.bullets(["a", "b"], numbered=True)
    d.heading("Mục hai", 2)
    d.table(["Cột & 1", "Cột <2>"], [["ô a", "ô b"], ["ô c", "ô d"]], caption="Bảng thử")
    d.note("Ghi chú", "KHÔNG KIỂM ĐƯỢC")
    d.add(Code("if (a < b) {\n  x = 1 & 2;\n}", "c"), PageBreak())
    d.heading("Mục ba", 1)
    return d


# ═════════════════ 1. tệp phải mở được ═════════════════


@pytest.mark.parametrize("dinh_dang", ["md", "docx", "xlsx", "pptx"])
def test_xuat_duoc_moi_dinh_dang(tmp_path, dinh_dang):
    p = office.write(_doc_day_du(), tmp_path / f"t.{dinh_dang}")
    assert p.is_file() and p.stat().st_size > 0


@pytest.mark.parametrize("dinh_dang", ["docx", "pptx", "xlsx"])
def test_moi_part_XML_trong_goi_deu_phan_tich_duoc(tmp_path, dinh_dang):
    """Một part XML hỏng làm cả tệp không mở, và trình mở chỉ báo 'tệp hỏng'."""
    p = office.write(_doc_day_du(), tmp_path / f"t.{dinh_dang}")
    with zipfile.ZipFile(p) as z:
        assert z.testzip() is None
        for n in z.namelist():
            if n.endswith((".xml", ".rels")):
                ET.fromstring(z.read(n))  # ném ParseError nếu hỏng


def test_ky_tu_dieu_khien_bi_loai_khoi_XML(tmp_path):
    """Một ký tự điều khiển lọt vào XML là tệp không mở được.

    Ký tự này tới từ dữ liệu dự án — một tệp YAML dán từ PDF rất dễ mang theo
    chúng — nên đây không phải trường hợp giả tưởng.
    """
    d = Doc(title="T")
    d.para("trước\x01\x02sau")
    p = office.write_docx(d, tmp_path / "t.docx")
    with zipfile.ZipFile(p) as z:
        xml = z.read("word/document.xml").decode("utf-8")
    ET.fromstring(xml)
    assert "\x01" not in xml and "trướcsau" in xml


def test_dau_va_trong_du_lieu_khong_lam_vo_tep(tmp_path):
    d = Doc(title="A & B <C>")
    d.table(["x & y"], [["a < b"]])
    with zipfile.ZipFile(office.write_docx(d, tmp_path / "t.docx")) as z:
        for n in z.namelist():
            if n.endswith((".xml", ".rels")):
                ET.fromstring(z.read(n))


def test_pptx_luon_co_it_nhat_MOT_slide(tmp_path):
    """Một .pptx không slide nào là một tệp PowerPoint từ chối mở."""
    p = office.write_pptx(Doc(), tmp_path / "rong.pptx")
    with zipfile.ZipFile(p) as z:
        assert [n for n in z.namelist() if n.startswith("ppt/slides/slide")]


def test_pptx_khong_sinh_slide_TRONG(tmp_path):
    """Một tiêu đề không có gì dưới nó là một MỤC, không phải một slide.

    Slide đầu được miễn: slide bìa chỉ có tiêu đề là một slide bìa, đúng như
    mọi bản trình chiếu. Cái phải chặn là slide THỨ HAI trở đi chỉ có tiêu đề —
    người xem đọc nó thành tệp hỏng.
    """
    d = Doc(title="Bìa")
    d.heading("Mục không có gì dưới nó", 1)
    d.heading("Mục có nội dung", 1)
    d.para("nội dung")
    ds = office._gom_slide(d)
    assert all(dong for _, dong in ds[1:]), "có slide rỗng sau slide bìa"
    assert "Mục không có gì dưới nó" not in [t for t, _ in ds]


def test_xlsx_moi_bang_mot_sheet(tmp_path):
    from openpyxl import load_workbook

    d = Doc(title="T")
    d.table(["a"], [["1"]], caption="Bảng một")
    d.table(["b"], [["2"]], caption="Bảng hai")
    wb = load_workbook(office.write_xlsx(d, tmp_path / "t.xlsx"))
    assert "Bảng một" in wb.sheetnames and "Bảng hai" in wb.sheetnames
    # Cái lợi duy nhất của .xlsx là lọc/sắp/đếm — cần hàng tiêu đề đóng băng.
    assert wb["Bảng một"].freeze_panes == "A2"


def test_xlsx_ten_sheet_trung_khong_de_len_nhau(tmp_path):
    from openpyxl import load_workbook

    d = Doc(title="T")
    d.table(["a"], [["1"]], caption="Cùng tên")
    d.table(["b"], [["2"]], caption="Cùng tên")
    wb = load_workbook(office.write_xlsx(d, tmp_path / "t.xlsx"))
    assert len([n for n in wb.sheetnames if "Cùng tên" in n]) == 2


def test_dinh_dang_la_bao_loi_ro(tmp_path):
    with pytest.raises(office.OfficeError, match="Chưa xuất được"):
        office.write(Doc(title="T"), tmp_path / "t.rtf")


def test_sinh_pdf_KHONG_de_len_ban_docx_dang_co(tmp_path, monkeypatch):
    """`--format pdf` không được đè bản Word người dùng đã có.

    Bản .docx trung gian từng được dựng cạnh tệp đích, nghĩa là
    `design gen srs --format pdf` ghi ra `srs.docx` — đúng tên tệp lần chạy
    `--format docx` trước đó đã tạo. Mất một tệp bàn giao vì chạy một lệnh
    sinh tệp khác là loại hỏng người dùng không có cách nào đoán trước.
    """
    quy = tmp_path / "srs.docx"
    quy.write_text("bản Word của người dùng, không được mất", encoding="utf-8")

    def _gia(argv, **k):
        # Giả LibreOffice: sinh PDF vào đúng --outdir mà nó được bảo.
        ra = Path(argv[argv.index("--outdir") + 1])
        (ra / (Path(argv[-1]).stem + ".pdf")).write_bytes(b"%PDF-1.7\n%%EOF\n")
        return type("R", (), {"returncode": 0, "stdout": "", "stderr": ""})()

    monkeypatch.setattr(office, "tim_soffice", lambda: "/gia/soffice")
    monkeypatch.setattr(office.subprocess, "run", _gia)
    office.write_pdf(Doc(title="T"), tmp_path / "srs.pdf")

    assert (tmp_path / "srs.pdf").is_file()
    assert quy.read_text(encoding="utf-8") == "bản Word của người dùng, không được mất"
    assert sorted(p.name for p in tmp_path.iterdir()) == ["srs.docx", "srs.pdf"], \
        "để lại rác cạnh kết quả"


def test_sinh_pdf_dung_ho_so_LibreOffice_rieng(tmp_path, monkeypatch):
    """Người dùng đang mở LibreOffice thì tiến trình --headless phải không đụng nhau.

    Dùng chung hồ sơ người dùng, `--convert-to` lặng lẽ không chuyển gì và vẫn
    trả mã 0 — hỏng mà báo thành công, và tệp đích đơn giản là không có.
    """
    da_goi: list[list[str]] = []

    def _gia(argv, **k):
        da_goi.append(list(argv))
        ra = Path(argv[argv.index("--outdir") + 1])
        (ra / (Path(argv[-1]).stem + ".pdf")).write_bytes(b"%PDF-1.7\n")
        return type("R", (), {"returncode": 0, "stdout": "", "stderr": ""})()

    monkeypatch.setattr(office, "tim_soffice", lambda: "/gia/soffice")
    monkeypatch.setattr(office.subprocess, "run", _gia)
    office.write_pdf(Doc(title="T"), tmp_path / "t.pdf")

    assert any(a.startswith("-env:UserInstallation=") for a in da_goi[0]), \
        "thiếu hồ sơ riêng — sẽ đụng LibreOffice đang mở"


def test_LibreOffice_khong_sinh_ra_gi_thi_BAO_LOI(tmp_path, monkeypatch):
    """Trả mã 0 mà không có tệp vẫn là hỏng, và phải nói ra."""
    monkeypatch.setattr(office, "tim_soffice", lambda: "/gia/soffice")
    monkeypatch.setattr(
        office.subprocess, "run",
        lambda argv, **k: type("R", (), {"returncode": 0, "stdout": "", "stderr": ""})())
    with pytest.raises(office.OfficeError) as exc:
        office.write_pdf(Doc(title="T"), tmp_path / "t.pdf")
    assert "--format docx" in str(exc.value), "phải chỉ đường chạy được ngay"
    assert not (tmp_path / "t.pdf").exists()


#: Bốn bài dưới đây gọi LibreOffice THẬT và tốn ~58 giây trên tổng ~235 giây
#: của cả bộ. Chúng là thứ duy nhất kiểm được nhánh PDF, nên chúng chạy mặc
#: định; gắn nhãn `cham` để bỏ qua khi cần vòng lặp nhanh:
#:
#:     pytest -m "not cham"
#:
#: Tự bỏ qua trên máy chưa cài LibreOffice — CI không có nó vẫn xanh, và
#: "xanh vì bỏ qua" hiện ra trong bản tóm tắt của pytest chứ không im lặng.
co_soffice = pytest.mark.skipif(
    not office.tim_soffice(),
    reason="máy này chưa có LibreOffice — cài: brew install --cask libreoffice",
)
cham = pytest.mark.cham


@co_soffice
@cham
def test_pdf_that_doc_nguoc_lai_duoc_bang_chinh_bo_doc_cua_san_pham(tmp_path):
    """Vòng khép kín: dựng .pdf rồi đọc lại bằng eaa/pdftext.py.

    Đây là phép kiểm mạnh nhất có được mà không cần mắt người: nếu bộ xuất
    dựng ra một PDF hỏng, hoặc dấu tiếng Việt rụng trên đường docx → pdf, thì
    bộ đọc của chính sản phẩm sẽ không tìm thấy những chuỗi dưới đây.
    """
    from eaa.pdftext import extract_text

    d = Doc(title="Tiêu đề có dấu", subtitle="Phụ đề")
    d.heading("Mục một", 1)
    d.para("Chữ tiếng Việt đủ dấu: Cân Bằng, Đủ, Ưu, Nghiêng, Xung đột.")
    d.table(["Cột & 1", "Cột <2>"], [["giá trị a", "giá trị b"]], caption="Bảng thử")
    d.note("Ghi chú cảnh báo", "KHÔNG KIỂM ĐƯỢC")
    d.add(PageBreak())
    d.heading("Mục sau ngắt trang", 1)
    d.para("Nội dung trang sau.")

    p = office.write_pdf(d, tmp_path / "t.pdf", timeout_s=180)
    assert p.read_bytes()[:5] == b"%PDF-"

    t = extract_text(p)
    assert not t.empty and t.blank_pages == 0
    for s in ("Tiêu đề có dấu", "Cân Bằng, Đủ, Ưu", "giá trị a",
              "KHÔNG KIỂM ĐƯỢC", "Mục sau ngắt trang"):
        assert s in t.text, f"mất sau vòng docx → pdf → đọc lại: {s!r}"
    assert "Cột & 1" in t.text, "ký tự & không sống sót qua chuỗi xuất"


@co_soffice
@cham
def test_pdf_ngat_trang_ra_TRANG_MOI_that(tmp_path):
    """PageBreak phải thành trang mới, không thành một ký tự lạ giữa dòng."""
    from eaa.pdftext import extract_text

    d = Doc(title="T")
    d.para("trước ngắt")
    d.add(PageBreak())
    d.para("sau ngắt")
    t = extract_text(office.write_pdf(d, tmp_path / "t.pdf", timeout_s=180))
    assert len(t.pages) == 2, f"mong 2 trang, có {len(t.pages)}"
    assert "trước ngắt" in t.pages[0].text
    assert "sau ngắt" in t.pages[1].text


@co_soffice
@cham
def test_pdf_khong_de_len_ban_docx_ĐANG_CO_chay_that(tmp_path):
    """Bản mock ở trên canh hợp đồng; bài này canh LibreOffice thật."""
    quy = tmp_path / "t.docx"
    quy.write_bytes(b"ban Word cua nguoi dung")
    office.write_pdf(Doc(title="T"), tmp_path / "t.pdf", timeout_s=180)
    assert quy.read_bytes() == b"ban Word cua nguoi dung"
    assert sorted(x.name for x in tmp_path.iterdir()) == ["t.docx", "t.pdf"]


@co_soffice
@cham
def test_ba_luot_chuyen_SONG_SONG_deu_ra_tep(tmp_path):
    """Không có hồ sơ LibreOffice riêng thì một trong ba lượt không sinh gì.

    Đo được: chạy ba tiến trình `soffice --headless --convert-to` dùng chung
    hồ sơ mặc định thì 2/3 ra tệp. Với `-env:UserInstallation` riêng cho từng
    lượt thì 3/3. Bài này giữ con số ấy.
    """
    from concurrent.futures import ThreadPoolExecutor

    def _chay(i: int) -> Path:
        d = Doc(title=f"Tài liệu {i}")
        d.para(f"Nội dung {i}")
        return office.write_pdf(d, tmp_path / f"t{i}.pdf", timeout_s=180)

    with ThreadPoolExecutor(max_workers=3) as bo:
        ds = list(bo.map(_chay, range(3)))
    assert all(p.is_file() and p.stat().st_size > 0 for p in ds)


def test_thieu_libreoffice_thi_noi_CACH_CAI(tmp_path, monkeypatch):
    """Thiếu công cụ ngoài là chuyện cài đặt, không phải lỗi của tài liệu."""
    monkeypatch.setattr(office, "tim_soffice", lambda: "")
    with pytest.raises(office.ThieuCongCu) as exc:
        office.write_pdf(Doc(title="T"), tmp_path / "t.pdf")
    tin = str(exc.value)
    assert "brew" in tin and "apt-get" in tin
    assert "--format docx" in tin, "phải chỉ đường vòng chạy được ngay"


# ═════════════════ 2. tài liệu phải nói đúng ═════════════════


def _du_an(tmp_path: Path, **tep: str) -> Path:
    p = tmp_path / "du_an"
    p.mkdir(exist_ok=True)
    for ten, noi_dung in tep.items():
        (p / ten.replace("__", ".")).write_text(noi_dung, encoding="utf-8")
    return p


RANG_BUOC = """
version: 1
platform: pack_thu
mcu: chip_thu
limits: {vong_dieu_khien_ms: 10, so_dong_module_toi_da: 300}
forbidden: [ham_chan, cap_phat_dong]
acceptance:
  scenarios: [khởi động tĩnh]
  measurements: [{name: sai_so, key: sai_so, unit: đơn_vị, max: 1.0}]
"""

PHAN_CUNG = """
version: 1
project: du_an
description: Mô tả bo thử.
mcu: {part: chip_thu, clock_hz: 1000}
pin_functions: {P0: [gpio], P1: [gpio, chuc_nang_khac]}
components:
  - {id: cam_bien, part: linh_kien_thu, bus: bus_thu, provides: [so_do]}
  - {id: chua_lap, part: linh_kien_khac, pins: {tin_hieu: P1}, populated: false}
conflicts:
  - {pin: P1, claimed_by: [a.x, b.y], found_in: tep_nguon dòng 1, status: chưa phân xử}
"""

STATE = """{"schema_version": 1, "phase": "A",
"gates": {"G1": "pending", "G2": "approved", "G3": "pending", "G4": "pending", "G5": "pending"},
"backlog": [{"id": "mod_a", "status": "todo", "uses": ["bus_thu"], "depends_on": []},
            {"id": "mod_b", "status": "todo", "uses": [], "depends_on": ["mod_a"]}],
"constraints_version": "", "llm": {}, "created_at": "", "updated_at": ""}"""


@pytest.fixture
def du_an_day_du(tmp_path):
    return _du_an(tmp_path, constraints__yaml=RANG_BUOC,
                  hardware_profile__yaml=PHAN_CUNG, project_state__json=STATE)


@pytest.mark.parametrize("kind", ["urd", "srs", "sdd", "chuc_nang", "luong"])
def test_dung_duoc_moi_loai(du_an_day_du, kind):
    doc = build(load_spec(kind), du_an_day_du, created_at="2026-01-01")
    assert doc.blocks and doc.title
    assert len(doc.headings) >= len(load_spec(kind).sections)


def test_noi_dung_toi_tu_HO_SO_du_an(du_an_day_du):
    ra = build(load_spec("srs"), du_an_day_du, created_at="x").render_text()
    assert "Mô tả bo thử." in ra
    assert "mod_a" in ra and "mod_b" in ra
    assert "ham_chan" in ra, "ràng buộc cấm phải xuất hiện"
    assert "sai_so" in ra, "tiêu chí nghiệm thu phải xuất hiện"


def test_xung_dot_chua_phan_xu_ĐI_VAO_tai_lieu(du_an_day_du):
    """Xung đột chân không phân xử thành lỗi phần cứng, không phải lỗi biên dịch."""
    ra = build(load_spec("srs"), du_an_day_du, created_at="x").render_text()
    assert "XUNG ĐỘT CHƯA PHÂN XỬ" in ra
    assert "tep_nguon dòng 1" in ra, "phải nêu chỗ tìm thấy xung đột"


def test_xung_dot_da_phan_xu_thi_KHONG_bao_nua(tmp_path):
    pc = PHAN_CUNG.replace("status: chưa phân xử", "status: đã phân xử")
    p = _du_an(tmp_path, constraints__yaml=RANG_BUOC, hardware_profile__yaml=pc,
               project_state__json=STATE)
    assert "XUNG ĐỘT" not in build(load_spec("srs"), p, created_at="x").render_text()


def test_thieu_du_lieu_thi_NOI_RA_kem_lenh_phai_chay(tmp_path):
    """Mục trống đọc như 'không cần', trong khi thật ra là 'chưa ai điền'."""
    p = _du_an(tmp_path, project_state__json=STATE)
    doc = build(load_spec("srs"), p, created_at="x")
    ra = doc.render_text()
    assert "Chưa có dữ liệu" in ra
    assert "eaa " in ra, "mỗi chỗ thiếu phải kèm lệnh phải chạy"
    assert "KHÔNG KIỂM ĐƯỢC" in ra


def test_phu_luc_DEM_LAI_so_muc_thieu(tmp_path):
    p = _du_an(tmp_path, project_state__json=STATE)
    ra = build(load_spec("srs"), p, created_at="x").render_text()
    assert "mục chưa có dữ liệu:" in ra


def test_du_du_lieu_thi_phu_luc_noi_DA_KIEM(du_an_day_du):
    ra = build(load_spec("luong"), du_an_day_du, created_at="x").render_text()
    assert "Mọi mục đều có dữ liệu" in ra


def test_phu_luc_liet_ke_TEP_NGUON(du_an_day_du):
    ra = build(load_spec("srs"), du_an_day_du, created_at="x").render_text()
    for t in ("constraints.yaml", "hardware_profile.yaml", "project_state.json"):
        assert t in ra


def test_du_an_rong_van_dung_duoc_tai_lieu(tmp_path):
    """Không có tệp nào cũng phải ra một tài liệu, toàn phần thiếu được nêu tên."""
    p = tmp_path / "trong"
    p.mkdir()
    doc = build(load_spec("urd"), p, created_at="x")
    assert doc.blocks
    assert "không đọc được tệp nào" in doc.render_text()


def test_dung_hai_lan_ra_cung_noi_dung(du_an_day_du):
    """Không so được hai bản thì không dùng được để theo dõi thay đổi."""
    a = build(load_spec("sdd"), du_an_day_du, created_at="cùng-mốc").render_text()
    b = build(load_spec("sdd"), du_an_day_du, created_at="cùng-mốc").render_text()
    assert a == b


def test_thu_tu_dung_module_theo_phu_thuoc(du_an_day_du):
    ra = build(load_spec("sdd"), du_an_day_du, created_at="x").render_text()
    assert "Thứ tự dựng" in ra
    assert ra.index("mod_a") < ra.index("mod_b")


def test_vong_phu_thuoc_duoc_NOI_RA_chu_khong_treo(tmp_path):
    """Một vòng phụ thuộc phải thành một câu, không thành một vòng lặp vô hạn."""
    from eaa.designdoc import _thu_tu_dung

    class M:
        def __init__(self, i, d):
            self.id, self.depends_on = i, d

    lop = _thu_tu_dung([M("a", ["b"]), M("b", ["a"])])
    assert any("vòng phụ thuộc" in " ".join(l) for l in lop)


def test_gia_tri_nhieu_dong_KHONG_lam_vo_o_bang(tmp_path):
    """Khối YAML nhiều dòng giữ nguyên xuống dòng sẽ vỡ ô ở cả bốn định dạng."""
    pc = PHAN_CUNG.replace("description: Mô tả bo thử.",
                           "description: |\n  dòng một\n  dòng hai")
    p = _du_an(tmp_path, hardware_profile__yaml=pc, project_state__json=STATE)
    doc = build(load_spec("srs"), p, created_at="x")
    for b in doc.blocks:
        if isinstance(b, Table):
            for r in b.rows:
                assert all("\n" not in c for c in r), f"ô bảng có xuống dòng: {r}"


def test_khong_goi_mo_hinh_mot_lan_nao(du_an_day_du, monkeypatch):
    """Bài canh cấu trúc: tài liệu do mô hình viết không truy được về đâu cả."""
    import eaa.designdoc as dd

    nguon = (GOC / "eaa" / "designdoc.py").read_text(encoding="utf-8")
    for cam in ("_tao_llm", "GeminiClient", "llm.generate", "MockLLM"):
        assert cam not in nguon, f"designdoc.py chạm tới mô hình: {cam}"
    build(load_spec("srs"), du_an_day_du, created_at="x")


# ═════════════════ 3. khuôn mẫu là dữ liệu, phải khớp mã ═════════════════


def test_moi_khuon_mau_deu_doc_duoc():
    ds = list_specs()
    assert len(ds) >= 5
    for s in ds:
        assert s.kind and s.title and s.sections


def test_moi_nguon_trong_khuon_mau_deu_co_bo_cap_du_lieu():
    """Khuôn mẫu là dữ liệu nên đổi được mà không ai chạy test — trừ bài này."""
    thieu = []
    for s in list_specs():
        for m in s.sections:
            if m.nguon and m.nguon not in NGUON:
                thieu.append(f"{s.kind}:{m.title} → {m.nguon}")
    assert thieu == [], f"khuôn mẫu nêu bộ cấp dữ liệu chưa khai: {thieu}"


def test_bo_cap_du_lieu_thua_thi_KHONG_phai_loi():
    """Ngược lại thì không: một bộ chưa dùng là một mục sẵn cho khuôn mẫu mới."""
    dung = {m.nguon for s in list_specs() for m in s.sections if m.nguon}
    assert dung <= set(NGUON)


def test_khuon_mau_hong_bao_TEN_TEP(tmp_path):
    (tmp_path / "hong.yaml").write_text("a: [\n", encoding="utf-8")
    with pytest.raises(DesignDocError, match="hong.yaml"):
        load_spec("hong", tmp_path)


def test_khuon_mau_khong_co_muc_nao_bi_tu_choi(tmp_path):
    (tmp_path / "rong.yaml").write_text("kind: rong\ntitle: T\n", encoding="utf-8")
    with pytest.raises(DesignDocError, match="không có mục nào"):
        load_spec("rong", tmp_path)


def test_khuon_mau_khong_ton_tai_thi_LIET_KE_cai_dang_co(tmp_path):
    with pytest.raises(DesignDocError) as exc:
        load_spec("khong_co_that")
    assert "srs" in str(exc.value) and "sdd" in str(exc.value)


def test_nguon_la_khong_khai_thi_NOI_RA_chu_khong_bo_muc(tmp_path):
    (tmp_path / "la.yaml").write_text(
        "kind: la\ntitle: T\nsections:\n  - title: M\n    nguon: khong_co_bo_nay\n",
        encoding="utf-8")
    doc = build(load_spec("la", tmp_path), tmp_path, created_at="x")
    assert "khong_co_bo_nay" in doc.render_text()


def test_sdd_trinh_bay_theo_bon_muc_C4():
    s = load_spec("sdd")
    tieu_de = " ".join(m.title for m in s.sections)
    for muc in ("C1", "C2", "C3", "C4"):
        assert muc in tieu_de, f"SDD thiếu mức {muc} của mô hình C4"
    assert "C4" in s.standard


def test_khuon_mau_nam_trong_goi_khi_cai_dat():
    """Khuôn mẫu là dữ liệu — nếu không đóng gói thì bản cài thiếu hẳn nó."""
    cau_hinh = (GOC / "pyproject.toml").read_text(encoding="utf-8")
    assert "docspec" in cau_hinh, (
        "eaa/docspec/*.yaml chưa được khai trong package-data của pyproject.toml"
    )
    assert (SPEC_DIR / "srs.yaml").is_file()
