#!/usr/bin/env python3
"""Khảo sát Agent lập trình nhúng trên thị trường, và đối chiếu với EAA.

    python scripts/lam_bang_so_sanh.py

Sinh ``docs/EAA_So_sanh_Agent_nhung.xlsx``.

Bảng này trả lời câu gì
------------------------

Ba bảng đã có trong kho đều nhìn vào TRONG: `EAA_Bang_nang_luc` so mã với một
khung năng lực, `EAA_Nghiep_vu_Agent` so mã với nghề nhúng, `kiem_bang_nang_luc`
so lời khai với mã. Không bảng nào hỏi câu của người ngoài: **so với thứ người
khác đã làm được, ta đang ở đâu?**

Kỷ luật nguồn — chép đúng luật của chính đề án
----------------------------------------------

Đề án đã đặt hai hạng nguồn cho lớp truy cập mạng (SL-71..80): *chính chủ* mới
được thành tri thức, *mở* chỉ là manh mối. Bảng này chịu đúng luật ấy:

* **chính chủ** — trang của chính nhà cung cấp, hoặc bài báo khoa học có bản
  toàn văn đọc được. Ghi ĐÃ ĐỌC.
* **mở** — tin tức, trang tổng hợp, thông cáo báo chí đăng lại. Ghi NGHE NÓI,
  và không dòng nào của bảng đứng một mình trên hạng này.

Cột `nguồn` ghi ĐỊA CHỈ, không ghi kết luận — để người đọc đi kiểm lại được.
Ngày khảo sát ghi ở sheet đầu, vì một bảng khảo sát không có ngày là một bảng
sẽ cũ đi mà không ai biết.

Điều bảng này KHÔNG làm
-----------------------

Nó **không** chạy thử sản phẩm nào. Cột của đối thủ là *"họ khai có"*, không
phải *"đã đo được"* — và với sản phẩm thương mại thì khoảng cách giữa hai câu
ấy là khoảng cách ta không đo được từ bên ngoài. Cột của EAA thì ngược lại: mọi
ô đều truy về một mã TC chạy được.

Hai cột ấy KHÔNG cùng hạng bằng chứng, và gộp chúng thành một điểm số là việc
bảng này cố ý không làm.
"""

from __future__ import annotations

import sys
from pathlib import Path

GOC = Path(__file__).resolve().parent.parent
RA = GOC / "docs" / "EAA_So_sanh_Agent_nhung.xlsx"

NGAY_KHAO_SAT = "05/09/2026"

# Trạng thái của EAA
DU, PHAN, CHUA, KHONG = "ĐỦ", "MỘT PHẦN", "CHƯA", "KHÔNG THEO"
# Mức của đối thủ
CO, MOT_PHAN, KHONG_NEU = "có", "một phần", "không nêu"
# Ưu tiên bổ sung
CAO, VUA, THAP, DASH = "Cao", "Vừa", "Thấp", "—"

# --------------------------------------------------------------------------
# Nguồn khảo sát
# --------------------------------------------------------------------------

# (tên, dạng, hạng nguồn, mức đọc, địa chỉ, một câu nó là gì)
NGUON: list[tuple[str, str, str, str, str, str]] = [
    ("Embedder", "nền tảng thương mại", "chính chủ", "ĐÃ ĐỌC",
     "https://embedder.com/",
     "Nền tảng kỹ thuật firmware bằng AI. Đối thủ gần EAA nhất: nối đất bằng "
     "tài liệu phần cứng, trích dẫn nguồn cho từng giá trị, và kiểm chứng vòng "
     "kín trên silicon thật"),
    ("Embedder — tin nhận giải", "tin tức", "mở", "NGHE NÓI",
     "https://embeddedcomputing.com/technology/ai-machine-learning/"
     "ai-logic-devices-worload-acceleration/embedder-v031-ai-powered-firmware-"
     "engineering-platform-nominated-for-embedded-award-2026-in-the-startup-category",
     "v0.3.1 được đề cử Embedded Award 2026 hạng startup. Dùng để biết mức "
     "trưởng thành, không dùng để rút tính năng"),
    ("Skilled AI Agents for Embedded and IoT Systems Development",
     "bài báo khoa học", "chính chủ", "ĐÃ ĐỌC",
     "https://arxiv.org/html/2603.19583v1",
     "Kiến trúc KỸ NĂNG (skills) chưng cất mẫu lập trình + ràng buộc khởi tạo "
     "+ cách hỏng đã biết cho từng ngoại vi. Kèm IoT-SkillsBench: 42 nhiệm vụ, "
     "23 ngoại vi, 378 lượt chạy trên phần cứng THẬT"),
    ("Benchmarking LLMs for Embedded Systems Programming (MDPI Future Internet)",
     "bài báo khoa học", "mở", "CHƯA ĐỌC TOÀN VĂN",
     "https://www.mdpi.com/1999-5903/18/2/94",
     "27 LLM trên 8 kịch bản nhúng tăng dần độ khó. Trang trả 403 khi tải — "
     "chỉ dùng phần tóm tắt tìm được, và bảng này KHÔNG rút tính năng nào từ nó"),
    ("STM32 Sidekick (STMicroelectronics)", "trợ lý nhà sản xuất", "mở", "NGHE NÓI",
     "https://www.dataweek.co.za/27366r",
     "Trợ lý AI cho hệ sinh thái STM32, huấn luyện trên tài liệu chính thức "
     "của hãng. Là bộ TRA CỨU, không phải agent chạy vòng kín"),
    ("Espressif — Copilot4Eclipse & Documentation MCP Server",
     "tích hợp nhà sản xuất", "chính chủ", "ĐÃ ĐỌC",
     "https://developer.espressif.com/blog/2025/02/github-copilot-in-espressif-ide/",
     "Đưa Copilot vào IDE của hãng, và mở một máy chủ MCP nối agent thẳng tới "
     "tài liệu chính chủ ESP-IDF"),
    ("Trợ lý đa dụng trong miền nhúng (tổng hợp của Avnet)", "bài phân tích",
     "mở", "NGHE NÓI",
     "https://www.avnet.com/americas/resources/article/ai-coding-assistants-in-the-embedded-domain/",
     "Bức tranh chung: trợ lý đa dụng sinh mã tốt, nhưng dừng lại ở ranh giới "
     "phần cứng"),
    ("AI-Assisted Hardware-in-the-Loop for Embedded Linux", "bài kỹ thuật",
     "mở", "NGHE NÓI",
     "https://electronicsconsult.com/blog/ai-assisted-hardware-in-the-loop/",
     "Phát biểu gọn nhất về khoảng trống: agent sửa được tệp và chạy được "
     "build, nhưng KHÔNG nạp được bo, không xem nó khởi động, không đọc ngược "
     "được cái vừa xảy ra"),
    ("MISRA / ISO 26262 với mã do AI sinh", "chuẩn ngành", "mở", "NGHE NÓI",
     "https://www.parasoft.com/learning-center/iso-26262/misra/",
     "Mã AI sinh chịu ĐÚNG bộ luật như mã người viết; chuẩn đòi truy vết hai "
     "chiều yêu cầu ↔ mã ↔ kiểm chứng, và đòi công cụ phân tích phải TẤT ĐỊNH "
     "mới qualify được"),
]

# --------------------------------------------------------------------------
# Trục tính năng — rút từ các nguồn ĐÃ ĐỌC
# --------------------------------------------------------------------------

# (mã, nhóm, tính năng, Embedder, bài arXiv skills, trợ lý đa dụng,
#  trạng thái EAA, bằng chứng EAA, khoảng cách / nhận định, ưu tiên)
TINH_NANG: list[tuple[str, str, str, str, str, str, str, str, str, str]] = [
    # ------------------------------------------------------- A. nối đất ---
    ("A1", "A. Nối đất tri thức", "Chỉ mục datasheet / reference manual / errata",
     CO, CO, KHONG_NEU, DU,
     "eaa/kb.py DatasheetStore + eaa/ingest.py → G2; errata qua eaa/docplan.py "
     "ErrataAnalysis · TC-22, TC-55f,g,h",
     "Ngang. Khác ở chỗ EAA đòi NGƯỜI duyệt từng trích đoạn tại G2; Embedder "
     "khai chỉ mục tự động trên toàn bộ tài liệu", DASH),
    ("A2", "A. Nối đất tri thức", "Nạp SVD / file mô tả thanh ghi của hãng",
     CO, KHONG_NEU, KHONG_NEU, CHUA,
     "—",
     "THIẾU RÕ. SVD là bảng thanh ghi máy đọc được, chính hãng phát hành — nó "
     "làm được đúng việc mà trích đoạn thủ công đang làm, nhanh hơn nhiều lần "
     "và ít sai hơn. Đây là khoảng trống đáng lấp nhất của nhóm A", CAO),
    ("A3", "A. Nối đất tri thức", "Nạp SƠ ĐỒ NGUYÊN LÝ / netlist thành bối cảnh bo",
     CO, KHONG_NEU, KHONG_NEU, CHUA,
     "—",
     "THIẾU. Embedder khai đọc Altium/KiCad/Eagle/PADS/Xpedition rồi giải "
     "netlist ra chân, pull-up, địa chỉ bus. EAA nhận việc ấy bằng "
     "hardware_profile.yaml gõ tay — và SL-125 là lần chính chỗ gõ tay ấy sai", CAO),
    ("A4", "A. Nối đất tri thức", "Trích dẫn nguồn cho TỪNG giá trị thanh ghi",
     CO, KHONG_NEU, KHONG_NEU, DU,
     "Luật `// ref: <mã chunk>` cưỡng chế ở cổng phân tích tĩnh · TC-17",
     "NGANG, và EAA cưỡng chế mạnh hơn: thiếu trích dẫn là cổng ĐỎ, không phải "
     "một ghi chú. Embedder khai 'cites the source', không nêu có chặn hay không", DASH),
    ("A5", "A. Nối đất tri thức", "Nêu giá trị KHÔNG có trong tài liệu để người xem",
     CO, KHONG_NEU, KHONG_NEU, DU,
     "eaa/readiness.py chặn sinh mã khi thiếu tri thức; prompt cấm lấp chỗ "
     "trống · TC-24",
     "Ngang", DASH),
    ("A6", "A. Nối đất tri thức", "Truy vấn tài liệu chính chủ lúc chạy (MCP / API hãng)",
     CO, KHONG_NEU, MOT_PHAN, PHAN,
     "eaa/web.py hai hạng nguồn + eaa research / eaa read · TC-65, TC-66",
     "MỘT PHẦN. EAA tải được trang chính chủ và cưỡng chế hạng nguồn theo URL "
     "cuối — chỗ Espressif dùng MCP thì EAA dùng HTTP có kiểm hạng. Chưa nối "
     "máy chủ MCP nào của hãng", VUA),
    ("A7", "A. Nối đất tri thức", "Vòng đời tri thức: thay trích đoạn → truy ngược mã bị ảnh hưởng",
     KHONG_NEU, KHONG_NEU, KHONG_NEU, DU,
     "eaa/lifecycle.py ba đường truy ngược + eaa knowledge stale · TC-29, TC-132",
     "**EAA HƠN.** Không nguồn nào trong khảo sát nêu năng lực này. Sửa một "
     "datasheet mà không biết mã nào đứng trên bản cũ là để lỗi nằm im", DASH),

    # -------------------------------------------------------- B. sinh mã ---
    ("B1", "B. Sinh mã", "Sinh driver ngoại vi từ tài liệu",
     CO, CO, MOT_PHAN, DU,
     "eaa/composer.py 7 lớp + eaa/orchestrator.py · TC-04, TC-17", "Ngang", DASH),
    ("B2", "B. Sinh mã", "Dựng khung dự án / ráp firmware hoàn chỉnh",
     CO, CO, MOT_PHAN, DU,
     "eaa/firmware.py + eaa build · TC-41", "Ngang", DASH),
    ("B3", "B. Sinh mã", "KỸ NĂNG chưng cất theo ngoại vi/MCU (mẫu + ràng buộc + cách hỏng)",
     KHONG_NEU, CO, KHONG_NEU, PHAN,
     "eaa/skills.py rút kỹ năng từ chuỗi việc đã lặp · TC-71; prompts/ của dự án",
     "KHÁC HẠNG. `skills.py` của EAA rút kỹ năng QUY TRÌNH (chuỗi lệnh hay "
     "lặp); bài arXiv nói kỹ năng PHẦN CỨNG cho một ngoại vi — mẫu khởi tạo, "
     "thứ tự bắt buộc, cách hỏng đã biết. Và bài ấy ĐO được: kỹ năng do người "
     "soạn đạt 41-42/42, kỹ năng do LLM tự sinh thì lợi ích thất thường", CAO),
    ("B4", "B. Sinh mã", "Nhiều ngôn ngữ (C / C++ / Rust)",
     CO, KHONG_NEU, CO, KHONG,
     "Chỉ C. `packs/*/pack.yaml` quyết định toolchain",
     "KHÔNG THEO ở phạm vi đề án — nhưng ranh giới Platform Pack là chỗ nới "
     "được mà không sửa engine, nên nó là lựa chọn chứ không phải bức tường", THAP),
    ("B5", "B. Sinh mã", "Bao phủ nhiều nền tảng / RTOS / SDK",
     CO, MOT_PHAN, MOT_PHAN, PHAN,
     "packs/avr + packs/stm32 · TC-47 (pack thứ hai dùng chung engine)",
     "Khoảng cách LỚN về BỀ RỘNG: Embedder khai 500+ nền tảng, 13 hãng, 5.500+ "
     "ngoại vi; bài arXiv có 3 cặp nền tảng–framework. EAA có 2 pack. Nhưng "
     "TC-47 đã chứng minh thêm pack không phải sửa engine — đây là công việc "
     "tuyến tính, không phải rào kiến trúc", VUA),

    # ---------------------------------------------------- C. kiểm chứng ---
    ("C1", "C. Kiểm chứng", "Dịch + phân tích tĩnh theo ràng buộc dự án",
     CO, CO, MOT_PHAN, DU,
     "eaa/tools/compile.py + static.py · TC-07, TC-40", "Ngang", DASH),
    ("C2", "C. Kiểm chứng", "Bài kiểm trên máy chủ (host test / SIL)",
     CO, KHONG_NEU, KHONG_NEU, DU,
     "eaa/tools/unittests.py + hostmock của pack · TC-56f,g, TC-112", "Ngang", DASH),
    ("C3", "C. Kiểm chứng", "Nạp xuống bo và đọc ngược xác minh",
     CO, CO, KHONG_NEU, DU,
     "eaa/flash.py preflight + VerifyResult · TC-42, TC-52",
     "Ngang, và EAA đọc ngược đối chiếu băm ảnh — Embedder khai 'flash and "
     "serial smoke test', không nêu đọc ngược", DASH),
    ("C4", "C. Kiểm chứng", "Thu telemetry và kết luận từ số đo",
     CO, CO, KHONG_NEU, DU,
     "eaa/telemetry.py + eaa/diagnostics.py chẩn đoán HAI KÊNH · TC-27, TC-43",
     "Ngang về thu; EAA có phần đối thủ không nêu: kết luận bằng PHÉP GIAO của "
     "kênh máy và kênh người", DASH),
    ("C5", "C. Kiểm chứng", "Điều khiển máy đo: probe, logic analyzer, oscilloscope",
     CO, KHONG_NEU, KHONG_NEU, CHUA,
     "eaa/diagnostics.py ManualMeasurement — NGƯỜI đo, Agent ghi vết (T0)",
     "THIẾU. Embedder khai phối hợp 30+ probe và máy đo, phân tích dạng sóng, "
     "đo công suất tương quan với thực thi. Đây là khoảng cách phần cứng lớn "
     "nhất, và nó tốn tiền thiết bị chứ không chỉ tốn mã", VUA),
    ("C6", "C. Kiểm chứng", "Phiên gỡ lỗi qua debugger (đọc thanh ghi / bộ nhớ)",
     CO, KHONG_NEU, KHONG_NEU, PHAN,
     "eaa/debugsession.py dựng kế hoạch phiên, ghi vết — Agent KHÔNG tự chạy · TC-75",
     "MỘT PHẦN có chủ ý (T0). Nới lên T3 cần một mạch gỡ lỗi cắm vào bo thật, "
     "vốn ngoài phạm vi đề án — nhưng đây là chỗ đối thủ chạy tự động", VUA),
    ("C7", "C. Kiểm chứng", "Vòng kín tự động: build → nạp → kiểm → sửa",
     CO, CO, KHONG_NEU, PHAN,
     "Vòng chuẩn 13 bước tự chạy; riêng NẠP đòi người duyệt ảnh · TC-93",
     "KHÁC CÓ CHỦ Ý, không phải thiếu. Embedder khai 'autonomously' và cho "
     "phép cấu hình bước nào cần duyệt; EAA đặt duyệt-nạp thành BẤT BIẾN không "
     "cờ nào tắt được. Đây là chỗ hai triết lý tách nhau, và cần nói rõ trong "
     "luận văn thay vì ghi thành một ô thiếu", DASH),
    ("C8", "C. Kiểm chứng", "Sinh bài kiểm TỪ tài liệu phần cứng",
     CO, KHONG_NEU, KHONG_NEU, PHAN,
     "Bài kiểm sinh kèm mã mỗi module (SL-134); chưa sinh TỪ tài liệu",
     "MỘT PHẦN. Bài kiểm hiện sinh từ ý định module, không từ bảng thanh ghi. "
     "Sinh từ tài liệu là đường đi tới bài kiểm biết trước giá trị đúng", VUA),
    ("C9", "C. Kiểm chứng", "Nhiều agent kiểm chạy song song",
     CO, KHONG_NEU, KHONG_NEU, KHONG,
     "Một vòng lặp tuần tự, có chủ ý",
     "KHÔNG THEO. Song song làm hỏng tính tái lập, và tái lập là điều kiện "
     "qualify công cụ theo ISO 26262. Với một đề án lấy bằng chứng làm trung "
     "tâm thì đây là đánh đổi sai hướng", THAP),
    ("C10", "C. Kiểm chứng", "Đo độ NHẠY của bài kiểm mới sinh",
     KHONG_NEU, KHONG_NEU, KHONG_NEU, PHAN,
     "eaa/sensitivity.py chạy bộ kiểm mới trên mã vừa bị đánh đỏ · TC-128",
     "**EAA HƠN.** Không nguồn nào nêu. Một bài kiểm xanh chưa phải bằng chứng "
     "— và cả ba benchmark khảo sát được đều đo pass/fail chứ không hỏi bài "
     "kiểm ấy có phân biệt được gì không", DASH),
    ("C11", "C. Kiểm chứng", "Bắt mã TỰ CHỈNH cho vừa đồ đo của chính nó",
     KHONG_NEU, KHONG_NEU, KHONG_NEU, PHAN,
     "eaa/instrument.py ba dấu vết, dừng vòng vá và hỏi người · TC-131",
     "**EAA HƠN.** Đây là dạng hỏng qua sạch mọi cổng tự động — 3/12 lần từ "
     "chối G3 của chính đề án là nó. Không khảo sát nào nêu tới", DASH),
    ("C12", "C. Kiểm chứng", "Canh hợp đồng gọi và lời gọi bị đánh rơi giữa hai lượt sinh",
     KHONG_NEU, KHONG_NEU, KHONG_NEU, DU,
     "eaa/contract.py chữ ký + tập lời gọi liên module · TC-124, TC-127",
     "**EAA HƠN.** Bằng chứng: `app_init()` mất bốn lời gọi khởi tạo, firmware "
     "câm, 33 bài kiểm vẫn xanh", DASH),

    # ------------------------------------------------- D. quyền và an toàn ---
    ("D1", "D. Quyền và an toàn", "Người duyệt KẾ HOẠCH trước khi Agent làm",
     CO, KHONG_NEU, KHONG_NEU, DU,
     "G1 chốt ràng buộc/kiến trúc, G2 chốt tri thức · TC-01, TC-28", "Ngang", DASH),
    ("D2", "D. Quyền và an toàn", "Người duyệt từng diff trước khi vào nhánh chính",
     CO, KHONG_NEU, KHONG_NEU, DU,
     "G3 + eaa/vcs.py authorize_merge, không có nhánh thứ hai tới merge · TC-01, TC-02",
     "NGANG về chức năng, KHÁC về cưỡng chế: EAA neo quyết định vào BĂM NỘI "
     "DUNG — ráp lại là hồ sơ khác và phải duyệt lại", DASH),
    ("D3", "D. Quyền và an toàn", "Chính sách chặn theo tệp / lệnh / hành động ra ngoài",
     CO, KHONG_NEU, MOT_PHAN, DU,
     "eaa/policy.py + danh mục TOOLBOX tĩnh trong Git; khoá phạm vi ghi tệp · TC-08, SL-154",
     "NGANG, và EAA chặt hơn ở một điểm: Agent KHÔNG CÓ ĐƯỜNG GỌI tới lệnh "
     "duyệt, chứ không phải được dặn đừng gọi", DASH),
    ("D4", "D. Quyền và an toàn", "Nạp firmware đòi duyệt tường minh",
     CO, KHONG_NEU, KHONG_NEU, DU,
     "eaa/flash.py đòi bản duyệt neo vào băm ảnh, không cờ bỏ qua · TC-42c, TC-93",
     "NGANG, và EAA không có chế độ tắt", DASH),
    ("D5", "D. Quyền và an toàn", "Trọng tài phần cứng khi nhiều phiên dùng chung một bo",
     CO, KHONG_NEU, KHONG_NEU, CHUA,
     "—",
     "THIẾU, và nó sẽ thành lỗi thật ngay khi có người thứ hai chạy cùng bo. "
     "Rẻ: một khoá tệp trên cổng nối tiếp", VUA),
    ("D6", "D. Quyền và an toàn", "Chạy ngoại tuyến / trong mạng nội bộ / air-gapped",
     CO, KHONG_NEU, KHONG_NEU, PHAN,
     "EAA_NO_NET=1 cắt mọi lối ra mạng; MockLLM và bộ phát lại chạy không cần khoá",
     "MỘT PHẦN: chạy được không mạng, nhưng lượt gọi mô hình THẬT vẫn đi ra "
     "API ngoài. Air-gapped đòi mô hình chạy tại chỗ", THAP),
    ("D7", "D. Quyền và an toàn", "Chứng nhận tổ chức (SOC 2, ISO 27001, GDPR)",
     CO, KHONG_NEU, KHONG_NEU, KHONG,
     "—",
     "KHÔNG THEO: đây là chứng nhận của một DOANH NGHIỆP, không phải tính năng "
     "phần mềm. Ghi vào bảng để không ai nhầm nó là khoảng trống kỹ thuật", DASH),
    ("D8", "D. Quyền và an toàn", "Dữ liệu khách không dùng để huấn luyện",
     CO, KHONG_NEU, KHONG_NEU, DU,
     "Không gửi gì ngoài lượt gọi mô hình; khoá qua biến môi trường, không vào "
     "log hay commit · TC-14", "Ngang", DASH),

    # ------------------------------------------- E. đo lường và bằng chứng ---
    ("E1", "E. Đo lường & bằng chứng", "Benchmark có phần cứng THẬT, công bố được",
     KHONG_NEU, CO, KHONG_NEU, PHAN,
     "Một dự án mẫu chạy trên bo thật (robot cân bằng), 26 lượt nạp có đọc ngược",
     "KHOẢNG CÁCH RÕ. Bài arXiv có IoT-SkillsBench: 42 nhiệm vụ, 23 ngoại vi, "
     "3 cặp nền tảng, 378 lượt chạy trên phần cứng thật, chỉ số CF/BF/BC và "
     "pass@1/pass@5. EAA có chiều SÂU trên một bài, chưa có chiều RỘNG", CAO),
    ("E2", "E. Đo lường & bằng chứng", "Chỉ số chuẩn: trượt dịch / sai hành vi / đúng, pass@k",
     KHONG_NEU, CO, KHONG_NEU, CHUA,
     "kpi_log.csv đếm lượt gọi, vòng vá, quyết định gate — không có pass@k",
     "THIẾU. Đây là ngôn ngữ mà người đọc luận văn dùng để so EAA với văn "
     "liệu. Rẻ: bộ dữ liệu đã có trong kpi_log, chỉ thiếu cách tính", CAO),
    ("E3", "E. Đo lường & bằng chứng", "Đếm token và chi phí từng lượt gọi",
     KHONG_NEU, CO, KHONG_NEU, DU,
     "eaa/budget.py TokenBudget + llm_calls.jsonl; cộng cả token SUY NGHĨ (SL-170) · TC-53e,f, TC-130",
     "Ngang trở lên", DASH),
    ("E4", "E. Đo lường & bằng chứng", "Truy vết hai chiều yêu cầu ↔ mã ↔ kiểm chứng",
     MOT_PHAN, KHONG_NEU, KHONG_NEU, PHAN,
     "Trích đoạn ↔ mã có (`// ref:` + lifecycle ba đường); commit mang prompt "
     "hash, model, chunk ids · NFR-07",
     "MỘT PHẦN. Chuẩn ISO 26262 đòi yêu cầu ↔ mã ↔ bài kiểm; EAA mạnh ở nhánh "
     "TRI THỨC ↔ mã nhưng chưa nối tiêu chí nghiệm thu xuống từng bài kiểm", VUA),
    ("E5", "E. Đo lường & bằng chứng", "Tất định và tái lập được (điều kiện qualify công cụ)",
     KHONG_NEU, KHONG_NEU, KHONG_NEU, DU,
     "Ghim phiên bản model, prompt stateless, bộ phát lại từ nhật ký, env_lock "
     "băm toolchain · TC-11, TC-15, TC-36",
     "**EAA HƠN, và đây là điểm mạnh dễ bị bỏ qua nhất.** ISO 26262 đòi công "
     "cụ phân tích phải tất định mới qualify. Bộ phát lại cố ý KHÔNG bịa phản "
     "hồi khi trượt băm — một lượt phát lại tự sinh nội dung là bằng chứng giả", DASH),
    ("E6", "E. Đo lường & bằng chứng", "Sổ ghi mọi sai lệch giữa thiết kế và mã",
     KHONG_NEU, KHONG_NEU, KHONG_NEU, DU,
     "docs/SAI_LECH_THIET_KE.md 175 mục + eaa/deviation.py + TC-60 quét mỗi lần chạy",
     "**EAA HƠN.** Không nguồn nào nêu. Đây là dữ liệu gốc của phương pháp "
     "huấn luyện, và là thứ một sản phẩm thương mại không có lý do gì công bố", DASH),

    # ------------------------------------------------------ F. vận hành ---
    ("F1", "F. Vận hành", "Tích hợp IDE (VS Code / Eclipse)",
     CO, KHONG_NEU, CO, CHUA,
     "CLI 56 lệnh + tầng hội thoại `eaa chat`",
     "THIẾU. Đây là mặt tiếp xúc mà kỹ sư nhúng thật sự ngồi trong đó cả ngày. "
     "Không đụng tới lõi: một lớp mỏng gọi CLI", VUA),
    ("F2", "F. Vận hành", "Chạy trong CI, gọi được từ terminal và kho mã",
     CO, KHONG_NEU, MOT_PHAN, DU,
     "CLI thoát mã chuẩn (0/2/3/4); .github workflow chạy bộ test mỗi commit", "Ngang", DASH),
    ("F3", "F. Vận hành", "Dùng lại toolchain, build system, flasher sẵn có",
     CO, KHONG_NEU, KHONG_NEU, DU,
     "eaa/platform.py là interface DUY NHẤT; pack khai lệnh dưới dạng dữ liệu · TC-38, TC-47",
     "Ngang, và ranh giới của EAA rõ hơn: engine không có một hằng số phần "
     "cứng nào, có test quét mỗi commit", DASH),
    ("F4", "F. Vận hành", "Trả lời câu hỏi kỹ thuật từ kho tri thức đã duyệt",
     CO, KHONG_NEU, MOT_PHAN, DU,
     "eaa/rag.py search_chunks + eaa recall — chỉ trả trích đoạn đã qua G2 · TC-126",
     "Ngang, và EAA nói rõ chunk chưa duyệt KHÔNG tính vào kết quả", DASH),
]


# --------------------------------------------------------------------------
# Kế hoạch vượt lên — xem docs/KE_HOACH_VUOT_LEN.md
# --------------------------------------------------------------------------

# (giai đoạn, việc, lấp mã nào, nuôi chiều sâu thế nào, chỗ đặt trong ba tầng,
#  mã TC, nặng)
KE_HOACH: list[tuple[str, str, str, str, str, str, str]] = [
    ("GĐ1 · Bản đồ thanh ghi máy đọc được",
     "Bộ đọc CMSIS-SVD và ATDF về MỘT mô hình trung tính",
     "A2",
     "Cho `// ref:` một nguồn để đối chiếu — hôm nay cổng chỉ kiểm CÓ trích "
     "dẫn, không kiểm trích dẫn ĐÚNG",
     "ENGINE eaa/regmap.py + regmap_svd.py + regmap_atdf.py (đọc ĐỊNH DẠNG, "
     "không biết tên thanh ghi nào); PACK khai đường dẫn; PROJECT giữ tệp",
     "TC-136", "vừa"),
    ("GĐ1 · Bản đồ thanh ghi máy đọc được",
     "Cổng thứ năm `regcheck`: thanh ghi có thật, trường có thật, giá trị lọt "
     "vừa độ rộng, không ghi vào thanh ghi chỉ-đọc",
     "A2",
     "Chặn được hạng lỗi mà bốn cổng hiện tại không thấy: giá trị hợp cú pháp "
     "mà sai với silicon",
     "ENGINE eaa/tools/regcheck.py, đứng sau `static` trước `unittests`",
     "TC-136", "vừa"),
    ("GĐ1 · Bản đồ thanh ghi máy đọc được",
     "Nối bản đồ vào N-908 và N-911",
     "A2",
     "instrument.py biết giá trị mới có CÒN HỢP LỆ không, không chỉ biết nó đã "
     "đổi; dimension.py có nguồn thứ hai cho độ rộng và thang",
     "ENGINE — sửa eaa/instrument.py và eaa/dimension.py",
     "TC-137", "nhẹ"),

    ("GĐ2 · Thước đo mới",
     "Chỉ số của văn liệu: pass@1, pass@5, trượt dịch / sai hành vi / đúng",
     "E1, E2",
     "Không có nó thì không đối thoại được với văn liệu. Dữ liệu đã có trong "
     "kpi_log.csv và llm_calls.jsonl — thiếu cách tính, không thiếu số",
     "ENGINE eaa/bench.py; bộ nhiệm vụ nằm ở bench/<bo>/<task>/",
     "TC-138", "vừa"),
    ("GĐ2 · Thước đo mới",
     "BỐN TRỤC ĐO KHÔNG AI CÓ: độ nhạy bài kiểm · vá chỉnh đồ đo · mất việc im "
     "lặng · truy về được",
     "E1, E2",
     "Đây là đóng góp nghiên cứu: không đua trên thước của họ mà đề xuất thước "
     "mới. Bốn bộ đo đã có sẵn (sensitivity, instrument, contract, regcheck)",
     "ENGINE eaa/bench.py đọc kết quả của bốn module đã có",
     "TC-138", "vừa"),

    ("GĐ3 · Kỹ năng phần cứng",
     "Kỹ năng theo ngoại vi: mẫu khởi tạo, THỨ TỰ bắt buộc, cách hỏng đã biết",
     "B3, C8",
     "Chỗ DUY NHẤT bài arXiv đo được là nâng kết quả lên gần trần (41-42/42). "
     "Và nó chặn hạng lỗi đã làm robot ngã: mã đúng mọi dòng, sai ở THỨ TỰ",
     "PACK packs/<pack>/skills/<ngoại vi>.md — tri thức của một họ chip, không "
     "phải của engine; ENGINE chỉ chọn theo `uses`",
     "TC-139", "vừa"),
    ("GĐ3 · Kỹ năng phần cứng",
     "Lớp K9 trong bộ ghép prompt, chỉ nhận kỹ năng ĐÃ DUYỆT G2",
     "B3",
     "Bài arXiv đo được: kỹ năng LLM tự sinh cho lợi ích THẤT THƯỜNG. Cửa duyệt "
     "là chỗ chặn đúng điều ấy",
     "ENGINE eaa/composer.py; ngân sách lấy từ `repair` (SÀN, không phải trần)",
     "TC-139", "nhẹ"),

    ("GĐ4 · Bối cảnh bo từ sơ đồ",
     "Đọc netlist KiCad → linh kiện, chân, net",
     "A3",
     "hardware_profile.yaml gõ tay là chỗ SL-125 sai, và giá phải trả là robot "
     "lao thẳng một phía",
     "ENGINE eaa/netlist.py (đọc ĐỊNH DẠNG mở); PROJECT giữ tệp",
     "TC-140", "nặng"),
    ("GĐ4 · Bối cảnh bo từ sơ đồ",
     "ĐỐI CHIẾU netlist với hồ sơ ở G1 — không sinh đè",
     "A3",
     "Hồ sơ mang thứ netlist không có: mức tích cực, pull-up nội, lý do chọn bộ "
     "đếm. Sinh đè là mất phần đắt nhất",
     "ENGINE — lệch chân thì ĐỎ ở G1; linh kiện hồ sơ thiếu thì cảnh báo (ca "
     "SL-143: còi và nút có trên bo mà hồ sơ chưa bao giờ khai)",
     "TC-140", "vừa"),

    ("Xen kẽ · rẻ, đáng làm, không nuôi chiều sâu",
     "Trọng tài phần cứng: khoá cổng nối tiếp",
     "D5",
     "Không nuôi chiều sâu, nhưng thành lỗi THẬT ngay khi có người thứ hai chạy "
     "cùng bo — nên làm sớm nhất trong ba",
     "ENGINE eaa/serialport.py — một khoá tệp",
     "TC-141", "nhẹ"),
    ("Xen kẽ · rẻ, đáng làm, không nuôi chiều sâu",
     "Truy vết hai chiều: tiêu chí nghiệm thu ↔ bài kiểm",
     "E4",
     "ISO 26262 đòi yêu cầu ↔ mã ↔ kiểm chứng. EAA mạnh ở nhánh tri thức ↔ mã, "
     "chưa nối tiêu chí xuống từng bài kiểm",
     "ENGINE eaa/acceptance.py + eaa/tools/unittests.py",
     "TC-142", "vừa"),
    ("Xen kẽ · rẻ, đáng làm, không nuôi chiều sâu",
     "Lớp mỏng tích hợp IDE",
     "F1",
     "Mặt tiếp xúc kỹ sư nhúng ngồi trong đó cả ngày. Không đụng lõi: gọi CLI",
     "NGOÀI engine — một extension gọi `eaa` qua dòng lệnh",
     "TC-143", "vừa"),

    ("Sau cùng · tốn thiết bị",
     "Điều khiển máy đo, và phiên gỡ lỗi tự động",
     "C5, C6",
     "Chỉ có nghĩa khi đã có benchmark để chứng minh chúng cải thiện được gì — "
     "tức là sau GĐ2",
     "PACK khai tên máy đo và lệnh; ENGINE chỉ chuyển tiếp (đúng luật FR-PLT-01)",
     "TC-144", "nặng"),
]


def main() -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    DAM = Font(bold=True, color="FFFFFF")
    NEN_TIEU_DE = PatternFill("solid", fgColor="1F3864")
    NEN_NHOM = PatternFill("solid", fgColor="D9E2F3")
    MAU_TT = {
        DU: PatternFill("solid", fgColor="C6EFCE"),
        PHAN: PatternFill("solid", fgColor="FFEB9C"),
        CHUA: PatternFill("solid", fgColor="FFC7CE"),
        KHONG: PatternFill("solid", fgColor="E7E6E6"),
    }
    MAU_DOI_THU = {
        CO: PatternFill("solid", fgColor="DDEBF7"),
        MOT_PHAN: PatternFill("solid", fgColor="FFF2CC"),
    }
    MAU_UT = {
        CAO: PatternFill("solid", fgColor="FF9999"),
        VUA: PatternFill("solid", fgColor="FFE699"),
    }
    VIEN = Border(*[Side("thin", color="BFBFBF")] * 4)
    TREN = Alignment(vertical="top", wrap_text=True)
    GIUA = Alignment(horizontal="center", vertical="top", wrap_text=True)

    wb = Workbook()

    def tieu_de(ws, cot):
        for i, (ten, rong) in enumerate(cot, start=1):
            o = ws.cell(row=1, column=i, value=ten)
            o.font, o.fill = DAM, NEN_TIEU_DE
            o.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(i)].width = rong
        ws.row_dimensions[1].height = 34
        ws.freeze_panes = "A2"

    # ═══════════════════════════════════════════════════ 1. Đọc trước ═══
    ws = wb.active
    ws.title = "Đọc trước"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 104
    dong = [
        ("SO SÁNH AGENT LẬP TRÌNH NHÚNG — EAA vs THỊ TRƯỜNG", ""),
        ("", ""),
        ("Đề án", "Agent lập trình nhúng tổng quát (Embedded AIDD Agent)"),
        ("Học viên", "Vũ Trí Công"),
        ("GVHD", "TS. Nguyễn Trung Hiếu"),
        ("Ngày khảo sát", f"{NGAY_KHAO_SAT} — sinh lại: python scripts/lam_bang_so_sanh.py"),
        ("", ""),
        ("BẢNG NÀY TRẢ LỜI CÂU GÌ", ""),
        ("", "Ba bảng đã có trong kho đều nhìn vào TRONG. Bảng này hỏi câu của "
             "người ngoài: so với thứ người khác đã làm được, ta đang ở đâu, và "
             "phải làm gì để ngang bằng rồi hơn."),
        ("", ""),
        ("HAI CỘT KHÔNG CÙNG HẠNG BẰNG CHỨNG", ""),
        ("Cột đối thủ", "Là 'HỌ KHAI CÓ' — đọc từ trang chính chủ hoặc bài báo. "
                        "Bảng này KHÔNG chạy thử sản phẩm nào, và với sản phẩm "
                        "thương mại thì khoảng cách giữa 'khai có' và 'đo được' "
                        "là khoảng cách không đo được từ bên ngoài."),
        ("Cột EAA", "Là 'ĐÃ ĐO ĐƯỢC' — mọi ô truy về một mã TC chạy được trong "
                    "kho, hoặc ghi thẳng là CHƯA."),
        ("Hệ quả", "Gộp hai cột thành một điểm số là việc bảng này CỐ Ý KHÔNG "
                   "làm. Một bảng xếp hạng dựng trên hai hạng bằng chứng khác "
                   "nhau là một bảng xếp hạng sai."),
        ("", ""),
        ("KỶ LUẬT NGUỒN — theo đúng luật của chính đề án (SL-71..80)", ""),
        ("chính chủ", "Trang của chính nhà cung cấp, hoặc bài báo có toàn văn "
                      "đọc được. Ghi ĐÃ ĐỌC."),
        ("mở", "Tin tức, trang tổng hợp, thông cáo đăng lại. Ghi NGHE NÓI — và "
               "KHÔNG dòng tính năng nào đứng một mình trên hạng này."),
        ("", ""),
        ("CÁCH ĐỌC CỘT TRẠNG THÁI EAA", ""),
        (DU, "Có mã chạy được VÀ chỉ ra được mã TC."),
        (PHAN, "Có phần lõi, thiếu một nhánh. Cột nhận định nói rõ thiếu gì."),
        (CHUA, "Không có mã nào làm việc này."),
        (KHONG, "Quyết định thiết kế, không phải thiếu sót. Không nên lấp — và "
                "cột nhận định phải nói được VÌ SAO."),
        ("", ""),
        ("ĐIỀU BẢNG NÀY KHÔNG LÀM", ""),
        ("", "Nó không đo chất lượng mã sinh ra của bất kỳ bên nào. Muốn so "
             "chất lượng thì phải có benchmark chung chạy trên cùng phần cứng — "
             "và chính đó là dòng E1/E2 của bảng."),
    ]
    for r, (a, b) in enumerate(dong, start=1):
        ws.cell(row=r, column=1, value=a).alignment = TREN
        ws.cell(row=r, column=2, value=b).alignment = TREN
        if a and not b:
            ws.cell(row=r, column=1).font = Font(bold=True, size=12 if r > 2 else 14)
        if a in MAU_TT:
            ws.cell(row=r, column=1).fill = MAU_TT[a]

    # ═══════════════════════════════════════════════ 2. Nguồn khảo sát ═══
    ws = wb.create_sheet("Nguồn khảo sát")
    tieu_de(ws, [("Tên", 34), ("Dạng", 20), ("Hạng nguồn", 12), ("Mức đọc", 18),
                 ("Địa chỉ", 58), ("Nó là gì", 68)])
    for r, hang in enumerate(NGUON, start=2):
        for c, v in enumerate(hang, start=1):
            o = ws.cell(row=r, column=c, value=v)
            o.alignment = TREN
            o.border = VIEN
        ws.cell(row=r, column=3).fill = (
            MAU_TT[DU] if hang[2] == "chính chủ" else MAU_TT[KHONG]
        )
        ws.cell(row=r, column=3).alignment = GIUA
    ws.auto_filter.ref = f"A1:F{len(NGUON) + 1}"

    # ═════════════════════════════════════════════════ 3. So tính năng ═══
    ws = wb.create_sheet("So tính năng")
    tieu_de(ws, [("Mã", 6), ("Nhóm", 22), ("Tính năng", 46),
                 ("Embedder", 11), ("arXiv skills", 12), ("Trợ lý đa dụng", 13),
                 ("EAA", 12), ("Bằng chứng EAA", 54), ("Nhận định / khoảng cách", 62),
                 ("Ưu tiên", 9)])
    r = 2
    nhom_truoc = ""
    for hang in TINH_NANG:
        if hang[1] != nhom_truoc:
            nhom_truoc = hang[1]
            for c in range(1, 11):
                ws.cell(row=r, column=c).fill = NEN_NHOM
            ws.cell(row=r, column=1, value=nhom_truoc).font = Font(bold=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
            r += 1
        for c, v in enumerate(hang, start=1):
            o = ws.cell(row=r, column=c, value=v)
            o.alignment = TREN
            o.border = VIEN
        for c in (4, 5, 6):
            o = ws.cell(row=r, column=c)
            o.alignment = GIUA
            if o.value in MAU_DOI_THU:
                o.fill = MAU_DOI_THU[o.value]
        ws.cell(row=r, column=7).fill = MAU_TT.get(hang[6], PatternFill())
        ws.cell(row=r, column=7).alignment = GIUA
        if hang[9] in MAU_UT:
            ws.cell(row=r, column=10).fill = MAU_UT[hang[9]]
        ws.cell(row=r, column=10).alignment = GIUA
        r += 1
    ws.auto_filter.ref = f"A1:J{r - 1}"

    # ════════════════════════════════════════════ 4. Việc phải làm ═══
    ws = wb.create_sheet("Việc phải làm")
    tieu_de(ws, [("Ưu tiên", 9), ("Mã", 6), ("Tính năng", 46), ("EAA", 12),
                 ("Vì sao đáng làm", 86)])
    thu_tu = {CAO: 0, VUA: 1, THAP: 2, DASH: 3}
    # Lọc theo ƯU TIÊN, không theo trạng thái: một dòng MỘT PHẦN mà cột ưu
    # tiên để '—' là dòng đã cân nhắc và quyết định KHÔNG làm tiếp — nửa còn
    # thiếu của nó thuộc về người, hoặc nó là chỗ EAA đang đi trước. Đưa nó vào
    # danh sách việc là biến một kết luận thành một món nợ giả.
    thieu = sorted(
        (h for h in TINH_NANG if h[6] in (PHAN, CHUA) and h[9] in (CAO, VUA, THAP)),
        key=lambda h: (thu_tu.get(h[9], 9), h[0]),
    )
    for r, h in enumerate(thieu, start=2):
        for c, v in enumerate((h[9], h[0], h[2], h[6], h[8]), start=1):
            o = ws.cell(row=r, column=c, value=v)
            o.alignment = TREN
            o.border = VIEN
        if h[9] in MAU_UT:
            ws.cell(row=r, column=1).fill = MAU_UT[h[9]]
        ws.cell(row=r, column=4).fill = MAU_TT.get(h[6], PatternFill())
        for c in (1, 4):
            ws.cell(row=r, column=c).alignment = GIUA
    ws.auto_filter.ref = f"A1:E{len(thieu) + 1}"

    # ═══════════════════════════════════════════════ 5. EAA hơn ở đâu ═══
    ws = wb.create_sheet("EAA hơn ở đâu")
    tieu_de(ws, [("Mã", 6), ("Tính năng", 46), ("Bằng chứng trong kho", 58),
                 ("Vì sao nó đáng kể", 76)])
    hon = [h for h in TINH_NANG if "EAA HƠN" in h[8]]
    for r, h in enumerate(hon, start=2):
        for c, v in enumerate((h[0], h[2], h[7], h[8]), start=1):
            o = ws.cell(row=r, column=c, value=v)
            o.alignment = TREN
            o.border = VIEN
    ws.auto_filter.ref = f"A1:D{len(hon) + 1}"

    # ══════════════════════════════════════════ 6. Kế hoạch vượt lên ═══
    ws = wb.create_sheet("Kế hoạch vượt lên")
    tieu_de(ws, [("Giai đoạn", 30), ("Việc", 46), ("Lấp mã", 9),
                 ("Nuôi chiều sâu thế nào", 60), ("Chỗ đặt trong ba tầng", 56),
                 ("Mã TC", 9), ("Nặng", 8)])
    r = 2
    gd_truoc = ""
    for hang in KE_HOACH:
        if hang[0] != gd_truoc:
            gd_truoc = hang[0]
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = NEN_NHOM
            ws.cell(row=r, column=1, value=gd_truoc).font = Font(bold=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            r += 1
        for c, v in enumerate(hang, start=1):
            o = ws.cell(row=r, column=c, value=v)
            o.alignment = TREN
            o.border = VIEN
        for c in (3, 6, 7):
            ws.cell(row=r, column=c).alignment = GIUA
        r += 1
    ws.auto_filter.ref = f"A1:G{r - 1}"

    RA.parent.mkdir(parents=True, exist_ok=True)
    wb.save(RA)

    dem = {k: sum(1 for h in TINH_NANG if h[6] == k) for k in (DU, PHAN, CHUA, KHONG)}
    print(f"Đã ghi {RA.relative_to(GOC)}")
    print(f"  Nguồn khảo sát : {len(NGUON)} "
          f"({sum(1 for n in NGUON if n[2] == 'chính chủ')} chính chủ)")
    print(f"  Tính năng      : {len(TINH_NANG)} — EAA ĐỦ {dem[DU]} · MỘT PHẦN "
          f"{dem[PHAN]} · CHƯA {dem[CHUA]} · KHÔNG THEO {dem[KHONG]}")
    print(f"  Việc phải làm  : {len(thieu)} dòng "
          f"({sum(1 for h in thieu if h[9] == CAO)} ưu tiên Cao)")
    print(f"  EAA hơn        : {len(hon)} dòng")
    print(f"  Kế hoạch       : {len(KE_HOACH)} việc / "
          f"{len({k[0] for k in KE_HOACH})} giai đoạn — xem docs/KE_HOACH_VUOT_LEN.md")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
