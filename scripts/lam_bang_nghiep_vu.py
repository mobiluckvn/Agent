"""Sinh bảng Excel phân tích NGHIỆP VỤ của Agent lập trình nhúng.

Chạy: python scripts/lam_bang_nghiep_vu.py [đường-dẫn-xlsx]

Bảng này KHÔNG nói về mã nguồn hiện có. Nó trả lời câu hỏi đứng trước đó: một
Agent lập trình nhúng, muốn tự chủ cao nhất, thì phải làm những việc gì, ở giai
đoạn nào, và làm BẰNG CÁCH NÀO.

Vì sao tách khỏi bảng tính năng: bảng tính năng đo "đã xây được gì", nên nó chỉ
nhìn thấy những việc đã nghĩ ra. Bảng này đi từ nghiệp vụ của nghề — từ lúc
nhận yêu cầu tới lúc bảo trì ngoài hiện trường — nên nó nhìn thấy cả những việc
chưa ai nghĩ tới. Đối chiếu hai bảng mới ra được khoảng trống thật.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Thang tự chủ — cột quan trọng nhất của cả bảng
# --------------------------------------------------------------------------

THANG_TU_CHU: list[tuple[str, str, str]] = [
    ("T0", "Người làm, Agent ghi vết",
     "Agent không làm được việc này. Nó chỉ ghi lại ai làm, làm lúc nào, kết "
     "quả ra sao. Dùng cho việc đòi tay người hoặc phán đoán vật lý."),
    ("T1", "Agent đề xuất, người quyết",
     "Agent dựng phương án/dữ kiện ở trạng thái ĐỀ XUẤT. Không có hiệu lực cho "
     "tới khi người chọn. Dùng khi có nhiều cách đều đúng, hoặc khi sai thì tốn kém."),
    ("T2", "Agent làm, người duyệt trước khi có hiệu lực",
     "Agent làm trọn việc và trình kết quả; kết quả chỉ vào hệ thống sau khi "
     "người duyệt tại một gate. Dùng cho thứ khó hoàn tác."),
    ("T3", "Agent tự làm, báo lại",
     "Agent làm và có hiệu lực ngay, nhưng để lại bằng chứng đầy đủ và báo cáo. "
     "Dùng khi việc hoàn tác được và sai thì phát hiện được."),
    ("T4", "Agent tự làm, không cần báo",
     "Chỉ dành cho việc đọc, việc tính, việc không đổi trạng thái gì. Mọi việc "
     "GHI ra ngoài đều không thuộc mức này."),
]

# --------------------------------------------------------------------------
# Nghiệp vụ: (mã, giai đoạn, nghiệp vụ, kết quả cần đạt, đầu vào,
#             cách Agent làm, mức tự chủ, người quyết gì, rủi ro nếu sai)
# --------------------------------------------------------------------------

G0 = "GĐ0 · Tiếp nhận & phạm vi"
G1 = "GĐ1 · Phân tích & đặc tả"
G2 = "GĐ2 · Môi trường & công cụ"
G3 = "GĐ3 · Tri thức phần cứng"
G4 = "GĐ4 · Thiết kế & phân rã"
G5 = "GĐ5 · Sinh mã & kiểm chứng tĩnh"
G6 = "GĐ6 · Mô phỏng"
G7 = "GĐ7 · Ráp & nạp"
G8 = "GĐ8 · Đưa lên mạch & đo"
G9 = "GĐ9 · Nghiệm thu & phát hành"
G10 = "GĐ10 · Vận hành & bảo trì"
XS = "XS · Xuyên suốt"

NGHIEP_VU: list[tuple[str, str, str, str, str, str, str, str, str]] = [
    # ---------------- GĐ0 ----------------
    ("N-001", G0, "Tiếp nhận yêu cầu bằng lời của người dùng",
     "Một phát biểu bài toán đủ rõ để bắt đầu: sản phẩm làm gì, chạy trên cái gì, ràng buộc nào",
     "Câu mô tả tự do của người dùng",
     "Hỏi có cấu trúc theo danh mục cố định (mục tiêu · đối tượng điều khiển · bo · nguồn · thời gian thực · môi trường). Mỗi câu hỏi kèm ví dụ trả lời. Không hỏi lại thứ đã suy được từ bước dò phần cứng.",
     "T1", "Chốt phát biểu bài toán",
     "Hiểu sai bài toán thì mọi việc sau đều đúng quy trình mà sai mục đích"),

    ("N-002", G0, "Dò phần cứng đang cắm vào máy",
     "Danh sách thiết bị: VID/PID, cổng nối tiếp, ổ nạp, số serial",
     "Máy có cổng USB",
     "Liệt kê cổng nối tiếp + đọc VID/PID qua thư viện hệ thống; đọc ổ mass-storage của mạch nạp; đối chiếu với cơ sở dữ liệu bo đã biết. Hỏi người CHỈ những gì không dò được.",
     "T4", "—",
     "Đoán nhầm bo dẫn tới nạp nhầm thiết bị ở GĐ7"),

    ("N-003", G0, "Nhận dạng bo/MCU từ dấu hiệu dò được",
     "Tên bo và mã MCU, kèm mức tin cậy",
     "Kết quả N-002",
     "Tra VID/PID và nhãn ổ nạp trong danh mục bo; nếu không chắc thì đưa 2–3 ứng viên kèm cách phân biệt (đọc mã in trên chip, đo điện trở, xem nhãn hộp) rồi hỏi người.",
     "T1", "Xác nhận đúng bo nào",
     "Sai MCU thì sai toàn bộ thanh ghi, chân, và bản đồ bộ nhớ"),

    ("N-004", G0, "Thu thập tài liệu gốc",
     "Bộ tài liệu: datasheet, reference manual, schematic bo, errata, application note",
     "Mã MCU và tên bo",
     "Liệt kê ĐÍCH DANH những tài liệu cần, kèm liên kết tới trang chính thức của hãng. Kiểm tài liệu người đưa có đúng phiên bản/rev không. Errata phải hỏi riêng — nó là tài liệu hay bị quên nhất.",
     "T1", "Cung cấp tệp hoặc xác nhận nguồn",
     "Thiếu errata là loại lỗi tốn nhiều giờ nhất, vì mã đúng theo datasheet mà vẫn chạy sai"),

    ("N-005", G0, "Lập danh mục giả định ban đầu",
     "Danh sách điều CHƯA BIẾT, mỗi mục kèm cách kiểm",
     "Kết quả N-001..N-004",
     "Mọi thứ chưa có nguồn đều vào danh mục giả định ở trạng thái đề xuất, KHÔNG viết lẫn vào hồ sơ như sự thật. Mỗi giả định ghi rõ: nó chặn việc gì, và kiểm bằng cách nào.",
     "T3", "—",
     "Giả định trộn lẫn với sự thật là cách kho tri thức mục ruỗng từ bên trong"),

    ("N-006", G0, "Xác định phạm vi và cái KHÔNG làm",
     "Danh sách tính năng trong phạm vi và ngoài phạm vi, có lý do",
     "Phát biểu bài toán",
     "Đề xuất ranh giới dựa trên năng lực bo và thời gian; nêu rõ thứ bị loại và vì sao (ví dụ: điều khiển màn hình DSI cần thư viện hãng, là một quyết định kiến trúc chứ không phải chi tiết kỹ thuật).",
     "T1", "Chốt phạm vi",
     "Phạm vi trôi làm hỏng mọi ước lượng và mọi tiêu chí nghiệm thu"),

    # ---------------- GĐ1 ----------------
    ("N-010", G1, "Chốt ràng buộc cứng",
     "Tệp ràng buộc có phiên bản: giới hạn thời gian, bộ nhớ, điều cấm, quy ước mã",
     "Phát biểu bài toán, hồ sơ bo",
     "Đề xuất bộ ràng buộc từ đặc tính vật lý của đối tượng (chu kỳ điều khiển suy từ động lực học) và từ năng lực chip. Mỗi ràng buộc kèm HỆ QUẢ nếu vi phạm, để người duyệt có căn cứ.",
     "T1", "Duyệt tại gate ràng buộc",
     "Ràng buộc sai thì mọi cổng kiểm chứng phía sau đo nhầm thước"),

    ("N-011", G1, "Định nghĩa tiêu chí nghiệm thu ĐO ĐƯỢC",
     "Danh sách số đo kèm ngưỡng, và cách đo từng số",
     "Ràng buộc cứng",
     "Ép mọi tiêu chí thành một con số có đơn vị và một cách đo. Từ chối tiêu chí kiểu 'chạy mượt'. Với mỗi số, nêu luôn nó lấy từ đâu (telemetry? đồng hồ đo? quan sát người?).",
     "T1", "Duyệt tiêu chí",
     "Tiêu chí không đo được thì nghiệm thu thành cảm tính, và không ai bác bỏ được"),

    ("N-012", G1, "Chọn kiến trúc phần mềm",
     "Quyết định có ghi lý do: bare-metal / vòng lặp hợp tác / RTOS / ngắt-ưu-tiên",
     "Ràng buộc thời gian thực, số việc đồng thời",
     "Trình 2–4 phương án kèm đánh đổi thật (mỗi phương án BẮT BUỘC nêu mặt trái), gợi ý một cách kèm lý do, nhưng KHÔNG tự chọn. Lưu cả phương án bị loại.",
     "T1", "Chọn phương án",
     "Chọn RTOS cho việc không cần nó là gánh thêm phức tạp suốt vòng đời"),

    ("N-013", G1, "Phân bổ tài nguyên phần cứng",
     "Bảng: module nào dùng timer/bus/DMA/chân nào",
     "Danh sách module dự kiến, hồ sơ bo",
     "Dựng đồ thị chiếm dụng; mặc định coi tài nguyên là ĐỘC CHIẾM, chia sẻ phải khai tường minh. Phát hiện xung đột ngay lúc khai báo, không đợi tới lúc chạy.",
     "T3", "Phân xử khi có xung đột",
     "Hai module cùng dùng một timer là lỗi chỉ hiện ra khi chạy đồng thời, rất khó lần"),

    ("N-014", G1, "Chốt sơ đồ chân",
     "Bảng chân: tên chân, hướng, ngoại vi, chức năng thay thế",
     "Schematic bo, phân bổ tài nguyên",
     "Trích từ schematic nếu có; nếu không có thì ĐỀ XUẤT dựa trên bảng chức năng thay thế của MCU và hỏi người xác nhận từng chân quan trọng. Kiểm chân có hỗ trợ chức năng cần không.",
     "T1", "Xác nhận từng chân",
     "Một chân sai làm mọi module chạm vào nó phải sửa lại"),

    ("N-015", G1, "Lập ngân sách tài nguyên",
     "Ngân sách flash / RAM / CPU / năng lượng, chia theo module",
     "Năng lực chip, danh sách module",
     "Chia ngân sách trước khi viết mã, không phải đo sau. Mỗi module có trần riêng; tổng phải nhỏ hơn năng lực chip trừ dự phòng. Cảnh báo khi một module ăn quá phần của nó.",
     "T1", "Duyệt cách chia",
     "Không có ngân sách thì tràn bộ nhớ chỉ lộ ra ở module cuối cùng"),

    ("N-016", G1, "Phân tích hỏng hóc và hệ quả",
     "Bảng: cái gì hỏng → biểu hiện → hậu quả → biện pháp",
     "Kiến trúc, sơ đồ chân, đối tượng điều khiển",
     "Duyệt từng ngoại vi và từng cảm biến: nếu nó im lặng thì sao, nếu nó trả rác thì sao, nếu nguồn sụt thì sao. Đề xuất cơ chế phát hiện cho từng trường hợp (quá hạn, kiểm tổng, giới hạn dải).",
     "T1", "Chốt mức xử lý cho từng hỏng hóc",
     "Hệ nhúng không có ai nhìn: hỏng mà không phát hiện được là hỏng âm thầm"),

    ("N-017", G1, "Xác định chế độ an toàn",
     "Định nghĩa trạng thái an toàn và điều kiện vào/ra khỏi nó",
     "Phân tích hỏng hóc",
     "Với mọi cơ cấu chấp hành, hỏi: mất điều khiển thì đưa nó về đâu. Đề xuất chế độ an toàn (tắt động cơ, giữ nguyên, về vị trí gốc) và điều kiện kích hoạt.",
     "T1", "Chốt chế độ an toàn",
     "Không có chế độ an toàn thì lỗi phần mềm thành hỏng cơ khí hoặc tai nạn"),

    # ---------------- GĐ2 ----------------
    ("N-020", G2, "Kiểm công cụ đã có trên máy",
     "Danh sách công cụ, phiên bản, và cổng kiểm chứng nào bị chặn nếu thiếu",
     "Nền tảng đã chọn",
     "Suy nhu cầu công cụ TỪ khai báo của nền tảng, không từ danh sách chép tay. Chạy lệnh kiểm phiên bản của từng công cụ; thiếu là KHÔNG ĐẠT, không phải bỏ qua.",
     "T4", "—",
     "Cổng im lặng cho qua vì không tìm thấy chương trình còn tệ hơn không có cổng"),

    ("N-021", G2, "Tìm công cụ chưa biết cách cài",
     "Đề xuất: cài bằng lệnh gì trên từng hệ điều hành, kiểm phiên bản ra sao",
     "Danh sách công cụ thiếu",
     "Tra cứu bằng mô hình nền; kiểm nguồn cài trước khi trình lên người (chỉ trình quản lý gói chính thống, hoặc tải trực tiếp từ miền cho phép kèm checksum). Đề xuất không đạt bị chặn ngay.",
     "T1", "Duyệt công cụ vào danh mục",
     "Một lệnh cài trông hợp lý nằm cạnh chín đề xuất hợp lệ là thứ dễ được bấm duyệt nhất"),

    ("N-022", G2, "Cài công cụ",
     "Công cụ có mặt, đúng phiên bản tối thiểu, chạy được",
     "Danh mục công cụ đã duyệt",
     "Hỏi người TRƯỚC TỪNG LỆNH cài, kể cả trong phiên tự động. Chạy smoke test sau khi cài. Phiên không có người thì KHÔNG cài.",
     "T2", "Xác nhận từng lệnh cài",
     "Cài phần mềm là thay đổi máy của kỹ sư — không phải việc máy tự quyết"),

    ("N-023", G2, "Khóa phiên bản môi trường",
     "Bản khóa: mỗi công cụ một phiên bản, kèm băm tổng thể",
     "Môi trường đã đủ",
     "Ghi phiên bản mọi công cụ vào bản khóa; mỗi lần chạy sau đối chiếu và cảnh báo khi trôi. Băm môi trường đi kèm mọi dòng chỉ số đo được.",
     "T3", "Chấp nhận khi trôi phiên bản",
     "Toolchain đổi phiên bản giữa kỳ làm hỏng so sánh A/B mà không ai biết"),

    ("N-024", G2, "Dựng kho mã và quy ước commit",
     "Kho Git, quy ước nhánh, mẫu commit có đủ dấu vết",
     "Thư mục dự án",
     "Khởi tạo kho; mỗi module một nhánh; commit mã sinh ra mang theo băm prompt, mã model, phiên bản ràng buộc, danh sách nguồn tri thức đã dùng.",
     "T3", "—",
     "Không truy vết được thì không bảo vệ được kết quả trước hội đồng"),

    # ---------------- GĐ3 ----------------
    ("N-030", G3, "Chọn trang tài liệu cần trích",
     "Danh sách trang/mục cần chưng cất thành tri thức",
     "Tài liệu gốc, danh sách ngoại vi dùng tới",
     "Từ đồ thị tài nguyên suy ra cần thanh ghi nào, rồi nêu ĐÍCH DANH trang cần trích. Không nạp tự động cả tệp — tài liệu hàng trăm trang chỉ cho ra vài chục trích đoạn có ích.",
     "T1", "Chọn/duyệt trang",
     "Nạp bừa cả tệp làm loãng truy xuất và tăng chi phí mọi lần gọi mô hình"),

    ("N-031", G3, "Chưng cất trích đoạn thành bảng thanh ghi–bit",
     "Trích đoạn có cấu trúc: thanh ghi, bit, ý nghĩa, giá trị, nguồn",
     "Trang tài liệu đã chọn",
     "Trích thành bảng chứ không giữ văn xuôi; mọi trích đoạn ở trạng thái ĐỀ XUẤT kèm băm nguồn và số trang. Ghi cả cảnh báo trong tài liệu (ví dụ: phải che bit trạng thái trước khi so sánh).",
     "T2", "Duyệt vào kho tri thức",
     "Một bit hiểu sai tạo ra mã chạy được nhưng sai hành vi — cổng nào cũng không bắt được"),

    ("N-032", G3, "Dựng đồ thị tri thức",
     "Đồ thị: module → ngoại vi → thanh ghi → trích đoạn → chân",
     "Hồ sơ bo, trích đoạn đã duyệt, khai báo module",
     "Dựng tự động từ dữ liệu đã có, không nhập tay. Đồ thị là thứ trả lời 'module này cần biết gì' và 'sửa trích đoạn này thì ảnh hưởng mã nào'.",
     "T4", "—",
     "Không có đồ thị thì truy xuất tri thức thành tìm kiếm mù"),

    ("N-033", G3, "Phát hiện mâu thuẫn nguồn",
     "Danh sách mục có hai nguồn nói khác nhau",
     "Kho trích đoạn",
     "So giá trị cùng một thanh ghi giữa các nguồn, sau khi chuẩn hóa cách viết số. Mâu thuẫn thì DỪNG chờ người — không chọn bản mới hơn hay nguồn có vẻ chính thống hơn.",
     "T1", "Phân xử mâu thuẫn",
     "Độ mới không phải bằng chứng đúng; máy tự chọn là máy tự bịa ra thẩm quyền"),

    ("N-034", G3, "Tự đánh giá đủ thông tin trước khi sinh mã",
     "Bảng kiểm: mỗi mục CÓ (kèm nguồn) / THIẾU / MÂU THUẪN",
     "Đồ thị tri thức, đặc tả module",
     "Sinh bảng kiểm tự động; chặn vòng sinh mã khi còn mục bắt buộc thiếu. CẤM đoán giá trị để lấp chỗ trống.",
     "T3", "—",
     "Giá trị đoán trông y hệt giá trị tra được, và nó qua được mọi cổng phía sau"),

    ("N-035", G3, "Đi tìm thứ còn thiếu (leo thang 3 bậc)",
     "Trích đoạn bổ sung, hoặc câu hỏi đích danh gửi người, hoặc kết luận hết cách",
     "Bảng kiểm còn mục thiếu",
     "Bậc 1 lục lại tài liệu người đã đưa mà chưa trích hết; bậc 2 hỏi người ĐÍCH DANH ('cần trang mô tả thanh ghi X ở chế độ Y'); bậc 3 tra miền nhà sản xuất cho phép. Tối đa 2 vòng mỗi mục rồi chuyển người.",
     "T2", "Cung cấp tài liệu; duyệt thứ tìm được",
     "Agent đứng im khi thiếu tri thức thì người dùng phải tự đoán ra mình cần làm gì"),

    ("N-036", G3, "Quản lý vòng đời tri thức",
     "Trích đoạn cũ được thay thế chứ không xóa; mã dùng trích đoạn cũ bị đánh dấu",
     "Kho tri thức, mã đã merge",
     "Thay thế theo lối bổ sung: bản mới trỏ về bản cũ. Khi một trích đoạn bị thay, truy ngược ra mọi module đã dùng nó (qua đồ thị, qua trích dẫn trong mã, qua dấu vết commit) và đánh dấu cần xem lại.",
     "T3", "Duyệt bản thay thế",
     "Sửa datasheet mà không biết mã nào bị ảnh hưởng là để lỗi nằm im chờ ngày lộ"),

    ("N-037", G3, "Theo dõi errata của nhà sản xuất",
     "Danh sách lỗi chip đã biết, và module nào chịu ảnh hưởng",
     "Mã MCU, rev silicon",
     "Hỏi rev silicon (in trên chip); tra errata của đúng rev đó; đối chiếu với ngoại vi dự án dùng; cảnh báo module nào chạm vào lỗi đã biết.",
     "T1", "Xác nhận rev; quyết cách né lỗi",
     "Mã đúng theo datasheet vẫn chạy sai nếu chip có lỗi đã công bố"),

    # ---------------- GĐ4 ----------------
    ("N-040", G4, "Phân rã bài toán thành module",
     "Danh sách module, mỗi module một trách nhiệm rõ",
     "Phát biểu bài toán, kiến trúc đã chọn",
     "Đề xuất phân rã theo ngoại vi và theo tầng (driver / logic / điều phối); mỗi module nêu rõ trách nhiệm, tài nguyên chiếm, và tiêu chí xong.",
     "T1", "Duyệt danh sách module",
     "Phân rã sai làm module phụ thuộc chằng chịt, không kiểm chứng riêng được"),

    ("N-041", G4, "Định nghĩa giao diện giữa module",
     "Tệp tiêu đề: hàm, kiểu dữ liệu, hợp đồng gọi",
     "Danh sách module",
     "Sinh giao diện TRƯỚC khi sinh phần thân, để module phụ thuộc có thể làm song song. Nêu rõ hàm nào gọi trong ngắt được, hàm nào chặn, hàm nào tái nhập được.",
     "T2", "Duyệt giao diện",
     "Đổi giao diện sau khi đã viết thân là sửa lan ra mọi nơi gọi tới"),

    ("N-042", G4, "Xếp thứ tự làm theo phụ thuộc",
     "Thứ tự module, và cái nào làm song song được",
     "Đồ thị phụ thuộc",
     "Sắp topo theo phụ thuộc; làm driver trước, logic sau; nêu rõ module nào chặn module nào để người biết đường ưu tiên.",
     "T3", "—",
     "Làm sai thứ tự thì phải giả lập phụ thuộc, tốn công gấp đôi"),

    ("N-043", G4, "Thiết kế lịch chạy",
     "Bảng việc: module nào chạy mỗi bao nhiêu, ưu tiên ra sao",
     "Ràng buộc thời gian, danh sách module",
     "Đề xuất chu kỳ từ yêu cầu vật lý (chu kỳ điều khiển, tốc độ lấy mẫu); kiểm tổng tải CPU ước lượng; cảnh báo khi tổng vượt 100% hoặc khi một việc dài hơn chu kỳ của việc khác.",
     "T1", "Chốt chu kỳ từng việc",
     "Chu kỳ là quyết định vật lý — máy đọc được tên hàm, không đọc được rằng con lắc cần 10 ms"),

    # ---------------- GĐ5 ----------------
    ("N-050", G5, "Sinh mã một module",
     "Mã nguồn + tệp tiêu đề của module",
     "Đặc tả module, tri thức đã duyệt, ràng buộc",
     "Nén ngữ cảnh có ngân sách: ràng buộc luôn có mặt, trích đoạn liên quan được chọn theo đồ thị, giao diện module phụ thuộc kèm theo. Mã cấu hình thanh ghi BẮT BUỘC trích dẫn nguồn.",
     "T2", "Review diff trước khi hợp nhất",
     "Mã không trích dẫn nguồn thì không phân biệt được tra được với bịa ra"),

    ("N-051", G5, "Biên dịch",
     "Tệp đối tượng, hoặc danh sách lỗi có vị trí",
     "Mã nguồn, toolchain",
     "Dịch từng tệp riêng (không liên kết), vì một module driver không có hàm main và không cần có. Dịch hết mọi tệp rồi mới kết luận, để lượt sửa thấy đủ lỗi.",
     "T4", "—",
     "Gộp dịch với liên kết làm mọi module trượt vì lý do chẳng liên quan tới chất lượng mã"),

    ("N-052", G5, "Phân tích tĩnh theo ràng buộc dự án",
     "Danh sách vi phạm, có vị trí và mã quy tắc",
     "Mã nguồn, bộ quy tắc nền tảng + ràng buộc dự án",
     "Hợp nhất quy tắc của nền tảng với điều cấm của dự án. Dự án khai Ý ĐỊNH ('cấm hàm chặn'), nền tảng lo cách phát hiện ý định ấy. Thiếu quy tắc cho một điều cấm là LỖI CẤU HÌNH, không phải đạt.",
     "T4", "—",
     "Cổng chỉ chặn được thứ nó ĐƯỢC BẢO là cấm; im lặng ở đây là cho qua"),

    ("N-053", G5, "Kiểm thử đơn vị trên máy chủ",
     "Kết quả test, độ phủ",
     "Mã nguồn, lớp trừu tượng phần cứng giả",
     "Chạy logic thuần trên máy chủ qua lớp phần cứng giả; tập trung vào tính toán và máy trạng thái. Nêu rõ phần nào KHÔNG kiểm được trên máy chủ.",
     "T4", "—",
     "Nhầm 'qua test trên máy chủ' với 'chạy đúng trên chip' là ngộ nhận tốn kém nhất"),

    ("N-054", G5, "Đo chiếm dụng bộ nhớ",
     "Số byte flash/RAM, đối chiếu ngân sách",
     "Tệp đối tượng hoặc ảnh liên kết",
     "Đo và so với ngân sách đã chia. Nói rõ đang đo tầm MODULE hay tầm FIRMWARE — trần của cả firmware áp lên một module lẻ là phép kiểm dễ dãi hơn nó trông.",
     "T4", "—",
     "Đo nhầm tầm làm ngân sách mất tác dụng đúng lúc cần nhất"),

    ("N-055", G5, "Vòng tự sửa khi cổng không đạt",
     "Mã đã sửa, hoặc bàn giao người sau N lần",
     "Báo cáo cổng hỏng",
     "Gửi bản vá kèm ĐÚNG lỗi và hàm liên quan, không gửi lại cả tệp. Tối đa N lần rồi dừng. Lỗi môi trường và lỗi cấu hình KHÔNG vào vòng này — mô hình không sửa được chúng.",
     "T3", "Nhận bàn giao khi quá N lần",
     "Đốt lượt sửa vào lỗi môi trường làm mất luôn cơ hội sửa lỗi thật"),

    ("N-056", G5, "Review và hợp nhất",
     "Mã vào nhánh chính, kèm dấu vết đầy đủ",
     "Toàn bộ báo cáo cổng đều đạt",
     "Trình diff kèm bảng kiểm sinh từ đồ thị tri thức (thanh ghi nào phải đối chiếu). Hợp nhất CHỈ khi mọi cổng đạt VÀ người duyệt; băm nội dung nối liền 'thứ đã duyệt' với 'thứ được hợp nhất'.",
     "T2", "Duyệt diff",
     "Một nhánh thứ hai dẫn tới hợp nhất là chỗ mọi bất biến bị vô hiệu"),

    ("N-057", G5, "Ghi nhận lỗi mô hình bịa ra",
     "Nhật ký: bịa cái gì, phát hiện ở cổng nào, sửa ra sao",
     "Kết quả các cổng",
     "Mỗi lần mã dùng thanh ghi không có thật, hoặc giá trị không khớp nguồn, thì ghi lại. Nhật ký này là dữ liệu cho chương đánh giá, và là đầu vào cải tiến prompt.",
     "T3", "—",
     "Không đo được tỉ lệ bịa thì không chứng minh được quy trình có tác dụng"),

    # ---------------- GĐ6 ----------------
    ("N-060", G6, "Dựng mô hình đối tượng điều khiển",
     "Mô hình toán của đối tượng, có tham số vật lý",
     "Thông số cơ khí, điện",
     "Đề xuất mô hình từ loại đối tượng (con lắc ngược, động cơ, nhiệt); tham số chưa đo được thì vào danh mục giả định. Nêu rõ mô hình bỏ qua hiện tượng nào.",
     "T1", "Duyệt mô hình và tham số",
     "Mô hình sai làm mọi kết quả mô phỏng thành vô nghĩa mà vẫn trông thuyết phục"),

    ("N-061", G6, "Chạy mô phỏng vòng kín",
     "Quỹ đạo, các chỉ số chất lượng điều khiển",
     "Mô hình, mã điều khiển, kịch bản",
     "Chạy các kịch bản đã khai (khởi động, nhiễu, tải nặng); tính chỉ số theo định nghĩa viết ra được (thời gian ổn định đo từ đâu tới đâu).",
     "T4", "—",
     "Chỉ số định nghĩa mơ hồ thì hai lần đo không so được với nhau"),

    ("N-062", G6, "Quét tham số và đề xuất bộ tham số",
     "Bảng kết quả theo tham số, và bộ được gợi ý",
     "Mô hình, dải tham số",
     "Quét lưới, xếp hạng theo chỉ số, nêu rõ tham số nào thật sự quyết định kết quả. Gợi ý một bộ kèm lý do; không tự chốt.",
     "T1", "Chọn bộ tham số",
     "Tự chốt tham số là quyết định thay người ở chỗ ảnh hưởng trực tiếp tới an toàn"),

    ("N-063", G6, "Tiêm lỗi trong mô phỏng",
     "Hành vi hệ thống khi cảm biến hỏng/trôi/mất",
     "Mô hình, phân tích hỏng hóc",
     "Tiêm từng lỗi đã liệt kê ở N-016: cảm biến trả rác, mất mẫu, trôi điểm không, nguồn sụt. Kiểm hệ có vào chế độ an toàn không.",
     "T3", "—",
     "Chỉ thử đường đi thuận là chỉ chứng minh hệ chạy khi không có gì hỏng"),

    ("N-064", G6, "Chạy chính mã C sinh ra trong mô phỏng",
     "Kết quả mô phỏng dùng artifact thật, không phải bản chép tay",
     "Mã đã dịch, mô hình",
     "Nối mã nguồn thật vào mô hình qua lớp phần cứng giả. Nếu chưa làm được thì phải nói rõ mô phỏng đang chạy MÔ HÌNH của thuật toán chứ không phải mã sẽ nạp.",
     "T3", "—",
     "Mô phỏng một bản chép tay rồi kết luận cho bản sẽ nạp là ngộ nhận âm thầm"),

    # ---------------- GĐ7 ----------------
    ("N-070", G7, "Ráp firmware hoàn chỉnh",
     "Ảnh liên kết từ mọi module đã hợp nhất + vòng lặp chính",
     "Module đã merge, bản thiết kế ráp",
     "Sinh vòng lặp chính từ khuôn của nền tảng và bảng việc của dự án. MỌI module đã merge phải có mặt trong bản thiết kế ráp — bỏ quên là LỖI, không phải cảnh báo.",
     "T2", "Duyệt bản thiết kế ráp",
     "Firmware thiếu một module đã qua kiểm chứng là thứ mà mọi bằng chứng đều nói là có"),

    ("N-071", G7, "Kiểm bộ nhớ ở tầm firmware",
     "Số byte thật của ảnh sẽ nạp, đối chiếu ngân sách",
     "Ảnh liên kết",
     "Đây là lần đầu ngưỡng bộ nhớ áp lên thứ sẽ thật sự chạy. Kiểm cả khoảng trống ngăn xếp còn lại, không chỉ tổng flash.",
     "T4", "—",
     "Tràn ngăn xếp là lỗi hiện ra ngẫu nhiên, rất khó lần ra nguyên nhân"),

    ("N-072", G7, "Kiểm trước khi nạp",
     "Kết luận: được nạp hay không, kèm lý do",
     "Ảnh, kho mã, nhật ký nạp",
     "Bốn phép kiểm đều là 'không' chứ không phải cảnh báo: có ảnh · kho mã sạch · ảnh mới hơn nguồn · người xác nhận.",
     "T2", "Xác nhận nạp",
     "Nạp ảnh cũ là cách hỏng âm thầm nhất: mạch chạy mã cũ, người đọc mã mới"),

    ("N-073", G7, "Chọn đúng thiết bị để nạp",
     "Cổng/mạch nạp đã xác định chắc chắn",
     "Danh sách thiết bị dò được, hồ sơ bo",
     "Tự chọn CHỈ khi danh tính xác nhận bằng VID/PID. Khớp theo tên cổng là phỏng đoán — cắm hai bo cùng lúc thì gợi ý tên rất dễ trúng nhầm.",
     "T3", "Chỉ định cổng khi mơ hồ",
     "Nạp nhầm thiết bị là hỏng thật, không phải một lượt chạy lại"),

    ("N-074", G7, "Nạp và ghi nhật ký",
     "Firmware trên chip; bản ghi: commit nào, ảnh nào, cổng nào, ai, lúc nào",
     "Ảnh đã kiểm, xác nhận của người",
     "Nạp qua công cụ của nền tảng; ghi cả lần TRƯỢT, vì 'đã thử và trượt' là dữ kiện chẩn đoán y như 'đã nạp xong'.",
     "T2", "Xác nhận",
     "Không biết bản nào đang trên chip thì mọi số đo về sau đều mất neo"),

    ("N-075", G7, "Kiểm sau khi nạp",
     "Xác nhận thứ trên chip đúng là thứ vừa gửi đi",
     "Ảnh, thiết bị",
     "Đọc ngược bộ nhớ hoặc so kiểm tổng nếu mạch nạp hỗ trợ; nếu không hỗ trợ thì nói rõ là không kiểm được, đừng ngầm coi 'nạp không báo lỗi' là 'nạp đúng'.",
     "T3", "—",
     "Nạp lỗi một phần tạo ra hành vi kỳ quái không thể lần theo mã nguồn"),

    # ---------------- GĐ8 ----------------
    ("N-080", G8, "Thiết lập kênh đo",
     "Kênh telemetry thông, khung tin kiểm được",
     "Bo đã nạp, cổng nối tiếp",
     "Khung tin có checksum và nhãn thời gian; luôn có hạn thời gian chờ; khung hỏng được ĐẾM chứ không nuốt; giữ bản ghi nguyên văn.",
     "T3", "—",
     "Phiên đo nhiều khung hỏng vẫn cho ra vài con số trông hợp lý"),

    ("N-081", G8, "Sinh firmware đo cho từng kịch bản chẩn đoán",
     "Ảnh firmware đo riêng cho mỗi kịch bản",
     "Kịch bản chẩn đoán, khuôn của nền tảng",
     "Ghép khung của nền tảng với phần đo của dự án bằng LIÊN KẾT, không dán chuỗi. Kịch bản chưa khai phần đo thì DỪNG, không dựng firmware rỗng.",
     "T2", "Duyệt trước khi nạp",
     "Firmware đo im lặng không phân biệt được với mạch hỏng"),

    ("N-082", G8, "Chẩn đoán hai kênh",
     "Kết luận nguyên nhân, có bằng chứng cả máy lẫn người",
     "Telemetry, quan sát của người",
     "Lấy GIAO của thứ máy đo được và thứ người quan sát được. Thiếu một kênh thì TỪ CHỐI kết luận. Kịch bản gây chuyển động đòi xác nhận đủ checklist an toàn.",
     "T1", "Trả lời quan sát; xác nhận an toàn",
     "Kết luận chỉ từ một kênh là đoán mò có vẻ khoa học"),

    ("N-083", G8, "Đo đặc tính thời gian thực",
     "Chu kỳ thật, độ trễ, dao động chu kỳ, tải CPU",
     "Firmware đo, kênh telemetry",
     "Đo trên thiết bị thật, không suy từ mô phỏng. So với ràng buộc thời gian đã chốt; nêu rõ trường hợp xấu nhất chứ không chỉ trung bình.",
     "T3", "—",
     "Trung bình đạt mà đỉnh trượt là lỗi chỉ xuất hiện lúc tải nặng"),

    ("N-084", G8, "Đo điện và nhiệt",
     "Dòng tiêu thụ, sụt áp khi tải, nhiệt độ linh kiện",
     "Đồng hồ đo, thiết bị chạy",
     "Hướng dẫn người đo đích danh: đo ở điểm nào, trong điều kiện nào, giá trị chờ đợi bao nhiêu. Agent ghi lại và đối chiếu ngưỡng.",
     "T0", "Thực hiện phép đo",
     "Sụt áp khi động cơ tăng tốc là nguyên nhân reset ngẫu nhiên rất hay bị bỏ sót"),

    ("N-085", G8, "Chẩn đoán sâu bằng công cụ gỡ lỗi",
     "Trạng thái thanh ghi/bộ nhớ tại điểm dừng; vết thực thi",
     "Mạch nạp có hỗ trợ gỡ lỗi",
     "Đặt điểm dừng, đọc thanh ghi lúc chạy, lấy vết. Chỉ cần khi hai kênh không đủ kết luận — nên xếp sau, không phải công cụ đầu tiên.",
     "T1", "Điều khiển phiên gỡ lỗi",
     "Nhảy vào gỡ lỗi quá sớm tốn thời gian hơn đọc kỹ telemetry"),

    ("N-086", G8, "Kiểm độ bền dài hạn",
     "Kết quả chạy liên tục nhiều giờ: có reset không, có trôi không",
     "Thiết bị, kênh đo",
     "Chạy dài, ghi liên tục, phát hiện reset (bộ đếm thời gian chạy nhảy về 0), trôi giá trị, rò bộ nhớ. Nêu rõ đã chạy bao lâu — 10 phút không kết luận được cho 10 giờ.",
     "T3", "Bố trí điều kiện chạy dài",
     "Lỗi tích lũy (trôi, tràn bộ đếm) chỉ lộ ra sau thời gian dài"),

    # ---------------- GĐ9 ----------------
    ("N-090", G9, "Đối chiếu tiêu chí nghiệm thu",
     "Bảng: tiêu chí — số đo — đạt/không",
     "Số đo thật, tiêu chí đã chốt",
     "Rút số đo từ telemetry theo tiêu chí đã khai TỪ TRƯỚC. Thiếu một số đo đã khai là LỖI. Vượt ngưỡng thì không phong hạng.",
     "T2", "Nghiệm thu tại gate",
     "Tiêu chí viết sau khi nhìn số thì phép đối chiếu không còn nghĩa gì"),

    ("N-091", G9, "Phong hạng chất lượng bản mã",
     "Hạng: dịch được / kiểm bằng mô phỏng / kiểm trên phần cứng",
     "Bằng chứng tương ứng từng hạng",
     "Hạng cao nhất đòi người duyệt VÀ số đo thật VÀ bản đang phong hạng phải là bản ĐANG CHẠY trên thiết bị (đối chiếu nhật ký nạp).",
     "T2", "Duyệt phong hạng",
     "Phong hạng cho bản chưa từng chạy trên chip làm hỏng mọi lần quay lui về sau"),

    ("N-092", G9, "Giữ bản chạy tốt và quay lui",
     "Bản known-good; khả năng quay về nó",
     "Lịch sử phong hạng",
     "Chỉ cập nhật bản known-good tại gate nghiệm thu. Quay lui không được làm mất bản known-good đang giữ.",
     "T3", "Quyết định quay lui",
     "Không có đường lui thì một bản hỏng làm dừng cả kỳ thực nghiệm"),

    ("N-093", G9, "Xuất báo cáo và dấu vết",
     "Báo cáo: chỉ số, nhật ký quyết định, nguồn tri thức đã dùng",
     "Toàn bộ nhật ký",
     "Xuất bảng chỉ số cho báo cáo; mỗi kết quả truy được về prompt, model, phiên bản ràng buộc, trích đoạn đã dùng.",
     "T3", "—",
     "Kết quả không truy vết được thì không bảo vệ được"),

    ("N-094", G9, "Bàn giao tài liệu vận hành",
     "Hướng dẫn nạp, đo, chẩn đoán, và giới hạn đã biết",
     "Toàn bộ dự án",
     "Sinh tài liệu từ dữ liệu đã có (bảng chân, bảng việc, kịch bản chẩn đoán, giả định chưa kiểm). Nêu rõ điều hệ thống KHÔNG làm được.",
     "T1", "Duyệt tài liệu bàn giao",
     "Bàn giao thiếu giới hạn đã biết là đẩy rủi ro sang người tiếp nhận"),

    # ---------------- GĐ10 ----------------
    ("N-100", G10, "Đánh giá ảnh hưởng khi tài liệu đổi",
     "Danh sách module cần xem lại",
     "Trích đoạn mới, errata mới",
     "Truy ngược từ trích đoạn ra mã theo ba đường: đồ thị, trích dẫn trong mã, dấu vết commit. Đánh dấu cần xem lại, không tự sửa.",
     "T1", "Quyết định sửa hay không",
     "Không truy ngược được thì tài liệu đổi mà mã đứng yên"),

    ("N-101", G10, "Đánh giá ảnh hưởng khi linh kiện đổi",
     "So sánh linh kiện cũ/mới, danh sách chỗ phải sửa",
     "Mã linh kiện thay thế",
     "So bảng thanh ghi, dải hoạt động, sơ đồ chân giữa hai linh kiện; nêu đích danh khác biệt chạm vào mã nào.",
     "T1", "Chọn linh kiện thay thế",
     "Linh kiện 'tương đương' thường khác ở đúng chỗ dự án đang dùng"),

    ("N-102", G10, "Chẩn đoán sự cố ngoài hiện trường",
     "Nguyên nhân, và bản sửa",
     "Mô tả triệu chứng, log nếu có",
     "Chọn kịch bản chẩn đoán từ mô tả triệu chứng; dựng lại điều kiện; dùng hai kênh. Nếu không dựng lại được thì nói rõ là chưa kết luận được.",
     "T1", "Cung cấp quan sát hiện trường",
     "Kết luận vội cho sự cố không dựng lại được thường sửa nhầm chỗ"),

    ("N-103", G10, "Cập nhật firmware cho thiết bị đã triển khai",
     "Quy trình cập nhật an toàn, có đường lui",
     "Bản mới đã nghiệm thu",
     "Kiểm tương thích ngược; có cơ chế quay lui khi cập nhật hỏng; không cập nhật hàng loạt trước khi thử trên một thiết bị.",
     "T2", "Duyệt triển khai",
     "Cập nhật hỏng trên thiết bị ở xa là thiết bị mất luôn"),

    # ---------------- Xuyên suốt ----------------
    ("N-900", XS, "Giữ ranh giới tổng quát / nền tảng / dự án",
     "Engine không chứa tri thức của một họ chip nào",
     "Toàn bộ mã",
     "Quét tự động mỗi lần commit: tên chip, tên thanh ghi, tên công cụ của một hãng không được nằm trong lõi. Thiếu năng lực thì mở rộng interface, không thêm nhánh rẽ.",
     "T4", "—",
     "Mất ranh giới thì sản phẩm thành công cụ cho đúng một con chip"),

    ("N-901", XS, "Bảo vệ khóa và bí mật",
     "Khóa không lọt vào log, commit, prompt, thông báo lỗi",
     "Biến môi trường",
     "Khóa chỉ đọc từ biến môi trường; che ở mọi lối ra; test tự động kiểm điều này.",
     "T4", "—",
     "Khóa lọt vào kho Git là sự cố không thu hồi được"),

    ("N-902", XS, "Không bao giờ hành động ngoài ý muốn người",
     "Mọi việc chạm vào máy hoặc thiết bị đều có xác nhận",
     "—",
     "Cài đặt, nạp, xóa, gửi ra ngoài đều đòi xác nhận. Phiên không có người KHÔNG được coi là người đã đồng ý. Không có cờ dòng lệnh nào bỏ qua được.",
     "T2", "Xác nhận từng lần",
     "Một cờ bỏ qua tồn tại là một cờ sẽ được dùng lúc vội"),

    ("N-903", XS, "Nói rõ mức tin cậy của mọi điều mình nói",
     "Mỗi kết luận kèm nguồn hoặc kèm nhãn 'chưa chắc'",
     "—",
     "Phân biệt ĐÃ KIỂM / SUY RA / GIẢ ĐỊNH ở mọi đầu ra. Không đọc được thì nói không đọc được, thay vì trả một câu trả lời trông giống như đã kiểm.",
     "T4", "—",
     "Một dòng 'không khớp' sai tệ hơn một dòng 'không nhận diện được' đúng"),

    ("N-904", XS, "Quản lý chi phí gọi mô hình",
     "Số token và chi phí theo module, có trần",
     "Nhật ký gọi mô hình",
     "Kiểm ngân sách TRƯỚC khi gọi; ghi số token mỗi lần; cảnh báo khi một module ăn quá phần của nó.",
     "T3", "Duyệt khi vượt trần",
     "Chi phí không đo được thì không tối ưu được, và dễ vượt ngoài dự tính"),

    ("N-905", XS, "Ghi nhận mọi sai lệch so với thiết kế",
     "Nhật ký sai lệch: lệch gì, vì sao, ảnh hưởng gì",
     "—",
     "Buộc phải lệch thì ghi lại ngay kèm lý do và cập nhật tài liệu tương ứng; không lệch ngầm.",
     "T3", "Duyệt sai lệch lớn",
     "Lệch ngầm làm tài liệu và mã kể hai câu chuyện khác nhau"),

    ("N-906", XS, "Tự đánh giá và cải tiến quy trình",
     "Chỉ số theo thời gian: tỉ lệ qua cổng, số lần tự sửa, lỗi bịa",
     "Nhật ký các vòng chạy",
     "Tổng hợp chỉ số qua các module; chỉ ra khâu nào hay hỏng nhất; đề xuất sửa prompt hoặc sửa quy tắc.",
     "T1", "Quyết định thay đổi quy trình",
     "Không đo thì mọi cải tiến chỉ là cảm giác"),

    ("N-907", XS, "Khôi phục sau gián đoạn",
     "Phiên tiếp tục đúng chỗ đã dừng",
     "Trạng thái dự án",
     "Ghi trạng thái nguyên tử, sống sót qua treo máy; khi mở lại thì nói rõ đang ở đâu và bước kế tiếp là gì.",
     "T4", "—",
     "Mất trạng thái giữa chừng làm hỏng cả chuỗi dấu vết"),
]



# --------------------------------------------------------------------------
# Đối chiếu với mã hiện có — thêm sau khi bảng nghiệp vụ đã được review
# --------------------------------------------------------------------------
#
# Cột đáng chú ý nhất KHÔNG phải "đã làm hay chưa" mà là khoảng cách giữa mức
# tự chủ ĐỀ XUẤT và mức tự chủ ĐẠT ĐƯỢC. Rất nhiều việc "đã có" nhưng đang
# chạy ở mức thấp hơn hẳn thiết kế: cơ chế thì đủ, phần chủ động thì thiếu.

# (mã, trạng thái, bằng chứng, còn thiếu, tự chủ đạt)
DU, PHAN, CHUA, COY = "Đủ", "Một phần", "Chưa có", "Cố ý không làm"

DOI_CHIEU = [
 ("N-001", DU, "eaa/brief.py QUESTIONS + eaa brief · TC-49c", "", "T1"),
 ("N-002", DU, "eaa/brief.py probe_hardware · TC-49a", "", "T4"),
 ("N-003", DU, "eaa/brief.py identify_board · TC-49b", "", "T1"),
 ("N-004", DU, "eaa/docplan.py plan_documents · TC-55a,b", "", "T1"),
 ("N-005", DU, "eaa/brief.py ProjectDraft.gia_dinh · TC-49d", "", "T3"),
 ("N-006", DU, "eaa/propose.py ScopeProposal · TC-54a", "", "T1"),
 ("N-010", DU, "eaa/propose.py ConstraintProposal · TC-54b", "", "T1"),
 ("N-011", DU, "eaa/propose.py AcceptanceProposal · TC-54c,d", "", "T1"),
 ("N-012", DU, "eaa/options.py + eaa decide · TC-46", "", "T1"),
 ("N-013", DU, "eaa/graph.py check_module · TC-18", "", "T3"),
 ("N-014", DU, "eaa/propose.py PinMapProposal + pin_functions · TC-54e,f", "", "T1"),
 ("N-015", DU, "eaa/budget.py ResourceBudget · TC-53a,b,c,g", "", "T1"),
 ("N-016", DU, "eaa/safety.py + eaa safety propose · TC-51 (29 test)", "", "T1"),
 ("N-017", DU, "eaa/safety.py SafeState · TC-51c, TC-51d", "", "T1"),
 ("N-020", DU, "eaa/doctor.py + toolsearch.derive_requirements · TC-34, TC-39", "", "T4"),
 ("N-021", DU, "eaa/toolsearch.py · TC-39", "", "T1"),
 ("N-022", DU, "eaa doctor --fix · TC-34, TC-37", "", "T2"),
 ("N-023", DU, "eaa/doctor.py EnvLock · TC-36", "", "T3"),
 ("N-024", DU, "eaa/vcs.py · TC-01", "", "T3"),
 ("N-030", DU, "eaa/docplan.py plan_pages · TC-55d,e", "", "T1"),
 ("N-031", DU, "eaa/ingest.py → G2 · TC-22", "", "T2"),
 ("N-032", DU, "eaa/graph.py · TC-18", "", "T4"),
 ("N-033", DU, "eaa/readiness.py conflict + chuẩn hóa số · TC-26", "", "T1"),
 ("N-034", DU, "eaa/readiness.py · TC-24", "", "T3"),
 ("N-035", DU, "eaa/gapsearch.py + eaa resolve · TC-48 (26 test)", "", "T2"),
 ("N-036", DU, "eaa/lifecycle.py (3 đường truy ngược) · TC-29", "", "T3"),
 ("N-037", DU, "eaa/docplan.py ErrataAnalysis · TC-55f,g,h", "", "T1"),
 ("N-040", DU, "eaa/decompose.py + eaa plan propose · TC-50 (32 test)", "", "T1"),
 ("N-041", DU, "eaa/interfaces.py + khuôn interfaces của pack · TC-56a..e", "", "T2"),
 ("N-042", DU, "decompose.order()/parallel_groups() · TC-50b", "", "T3"),
 ("N-043", DU, "decompose: chu kỳ + tải CPU + 3 phép kiểm · TC-50c", "", "T1"),
 ("N-050", DU, "eaa/composer.py + orchestrator · TC-04, TC-17", "", "T2"),
 ("N-051", DU, "eaa/tools/compile.py CompileGate · TC-40", "", "T4"),
 ("N-052", DU, "eaa/tools/static.py · TC-07", "", "T4"),
 ("N-053", DU, "eaa/tools/unittests.py host_gaps · TC-56f,g", "", "T4"),
 ("N-054", DU, "eaa/tools/compile.py SizeGate + size_scope · TC-40d", "", "T4"),
 ("N-055", DU, "eaa/orchestrator.py · TC-06, TC-19", "", "T3"),
 ("N-056", DU, "eaa/gates.py + vcs.py content_digest · TC-01", "", "T2"),
 ("N-057", DU, "eaa/ledger.py · TC-10", "", "T3"),
 ("N-060", DU, "eaa/propose.py PlantModelProposal · TC-57f,g", "", "T1"),
 ("N-061", DU, "eaa/tools/sim.py, sim_runner.py · TC-12", "", "T4"),
 ("N-062", DU, "eaa sim --sweep · TC-12", "", "T1"),
 ("N-063", DU, "sim_runner FaultSpec + require_safe_state · TC-57a..e", "", "T3"),
 ("N-064", COY, "—", "Cố ý để ngoài chuỗi tự động: nối vào mà không thật sự chạy artifact thì cổng sẽ 'đạt' mà chẳng kiểm gì.", "T0"),
 ("N-070", DU, "eaa/firmware.py + eaa build · TC-41", "", "T2"),
 ("N-071", DU, "budget.derived + stack_headroom_bytes_min · TC-53d", "", "T4"),
 ("N-072", DU, "eaa/flash.py preflight · TC-42b", "", "T2"),
 ("N-073", DU, "eaa/serialport.py match_confirmed · TC-47d", "", "T3"),
 ("N-074", DU, "eaa/flash.py + FlashLog · TC-42d", "", "T2"),
 ("N-075", DU, "eaa/flash.py VerifyResult + năng lực flash_verify · TC-52", "", "T3"),
 ("N-080", DU, "eaa/telemetry.py + eaa telemetry · TC-43", "", "T3"),
 ("N-081", DU, "đủ 6 kịch bản khai firmware_template · TC-58a", "", "T2"),
 ("N-082", DU, "eaa/diagnostics.py · TC-27", "", "T1"),
 ("N-083", DU, "DS-06 đo cả trường hợp xấu nhất và tải CPU · TC-58b", "", "T3"),
 ("N-084", DU, "ManualMeasurement + eaa diagnose measure · TC-58c,d", "", "T0"),
 ("N-085", CHUA, "—", "Ngoài phạm vi đề án (đã ghi từ trước).", "T0"),
 ("N-086", DU, "eaa/endurance.py + eaa endurance · TC-58e,f,g", "", "T3"),
 ("N-090", DU, "eaa/acceptance.py · TC-45", "", "T2"),
 ("N-091", DU, "eaa/versions.py + check_device_commit · TC-30, TC-45a", "", "T2"),
 ("N-092", DU, "eaa/versions.py + eaa rollback · TC-30", "", "T3"),
 ("N-093", DU, "eaa/kpi.py + llm/calllog.py + registry.py · TC-09, TC-15", "", "T3"),
 ("N-094", DU, "eaa/handover.py OperationsHandbook · TC-59a,b", "", "T1"),
 ("N-100", DU, "eaa/lifecycle.py · TC-29", "", "T1"),
 ("N-101", DU, "eaa/handover.py SwapAnalysis · TC-59c,d", "", "T1"),
 ("N-102", DU, "eaa/diagnostics.py FieldCase · TC-59e,f", "", "T1"),
 ("N-103", DU, "eaa/handover.py RolloutPlan · TC-59g,h", "", "T2"),
 ("N-900", DU, "TC-38 (quét mỗi commit) + TC-47a (không rẽ nhánh theo pack)", "", "T4"),
 ("N-901", DU, "eaa/llm/base.py mask_secrets · TC-14 + tests/conftest.py", "", "T4"),
 ("N-902", DU, "eaa/gates.py, doctor, flash — không cờ bỏ qua · TC-01, TC-42c", "", "T2"),
 ("N-903", PHAN, "eaa/confidence.py — bộ từ vựng chung · TC-60a,b,c", "Đã có MỘT bộ từ vựng và năm chỗ quy về nó (kiểm sau nạp, errata, tham số mô hình, bảng chân, chạy dài); các đầu ra còn lại vẫn chưa gắn nhãn.", "T3"),
 ("N-904", DU, "eaa/budget.py TokenBudget · TC-53e,f", "", "T3"),
 ("N-905", DU, "eaa/deviation.py + eaa deviations · TC-60d,e,f", "", "T3"),
 ("N-906", DU, "kpi.weak_points + eaa report review · TC-60g,h,i", "", "T1"),
 ("N-907", DU, "eaa/state.py ghi nguyên tử + eaa resume · TC-03", "", "T4"),
]


# --------------------------------------------------------------------------
# Câu hỏi Agent phải biết hỏi — danh mục khai thác yêu cầu
# --------------------------------------------------------------------------

CAU_HOI: list[tuple[str, str, str, str]] = [
    (G0, "Mục tiêu", "Thiết bị này phải làm được gì, đo bằng cách nào?",
     "Không có câu trả lời đo được thì không có tiêu chí nghiệm thu"),
    (G0, "Bo mạch", "Bo nào? (nếu đã cắm, Agent tự dò và chỉ hỏi xác nhận)",
     "Sai bo là sai toàn bộ thanh ghi và bản đồ bộ nhớ"),
    (G0, "Đối tượng", "Điều khiển cái gì: động cơ loại nào, cảm biến nào, dải bao nhiêu?",
     "Quyết định chu kỳ điều khiển và độ phân giải cần thiết"),
    (G0, "Nguồn", "Cấp nguồn thế nào? Nguồn động lực và nguồn điều khiển có tách không?",
     "Sụt áp khi tải là nguyên nhân reset ngẫu nhiên hay bị bỏ sót"),
    (G0, "Tài liệu", "Có datasheet, schematic, errata chưa? Rev silicon là gì?",
     "Errata là tài liệu hay bị quên nhất"),
    (G0, "Môi trường", "Chạy ở đâu: nhiệt độ, rung, nhiễu điện?",
     "Ảnh hưởng tới lựa chọn bộ lọc và ngưỡng cảnh báo"),
    (G1, "Thời gian thực", "Chu kỳ điều khiển bao nhiêu? Trễ tối đa chấp nhận được?",
     "Là quyết định vật lý, máy không suy ra được từ mã"),
    (G1, "An toàn", "Mất điều khiển thì đưa thiết bị về trạng thái nào?",
     "Không có chế độ an toàn thì lỗi phần mềm thành hỏng cơ khí"),
    (G1, "Ưu tiên", "Nếu phải chọn: chính xác hơn hay đáp ứng nhanh hơn?",
     "Định hướng mọi đánh đổi thiết kế về sau"),
    (G1, "Ràng buộc mềm", "Có phải theo chuẩn mã nguồn nào không (MISRA, coding style)?",
     "Thêm quy tắc vào cổng phân tích tĩnh"),
    (G4, "Kiến trúc", "Chấp nhận đưa thư viện của hãng vào không, hay bare-metal thuần?",
     "Quyết định kích thước firmware và mức phụ thuộc"),
    (G8, "Đo đạc", "Có đồng hồ đo, máy hiện sóng, tải giả không?",
     "Quyết định số đo nào lấy được, số nào phải bỏ"),
    (G9, "Nghiệm thu", "Ai ký nghiệm thu, và cần bằng chứng dạng gì?",
     "Định hình toàn bộ báo cáo và dấu vết phải giữ"),
]


# --------------------------------------------------------------------------
# Kiểu dáng
# --------------------------------------------------------------------------

MAU_TRANG_THAI = {
    "Đủ": "C6EFCE", "Một phần": "FFEB9C", "Chưa có": "FFC7CE", "Cố ý không làm": "D9E1F2"
}
MAU_CHU_TT = {
    "Đủ": "006100", "Một phần": "9C5700", "Chưa có": "9C0006", "Cố ý không làm": "1F4E79"
}
MAU_TU_CHU = {"T0": "F4B183", "T1": "FFE699", "T2": "C6E0B4", "T3": "BDD7EE", "T4": "D9D9D9"}
NEN_TIEU_DE = PatternFill("solid", fgColor="2F5597")
TRANG_DAM = Font(bold=True, color="FFFFFF")
VIEN = Border(*[Side(style="thin", color="BFBFBF")] * 4)
TREN_TRAI = Alignment(vertical="top", wrap_text=True)


def _tieu_de(ws, cot: list[str], rong: list[int]) -> None:
    ws.append(cot)
    for i, r in enumerate(rong, 1):
        ws.column_dimensions[get_column_letter(i)].width = r
    for o in ws[1]:
        o.font = TRANG_DAM
        o.fill = NEN_TIEU_DE
        o.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        o.border = VIEN
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


def dung_bang(dich: Path) -> Path:
    wb = Workbook()

    # ---- Sheet 1: Hướng dẫn đọc ----
    ws = wb.active
    ws.title = "Hướng dẫn đọc"
    ws.append(["Nghiệp vụ của Agent lập trình nhúng — bảng phân tích để review"])
    ws["A1"].font = Font(bold=True, size=16, color="2F5597")
    ws.append([
        "Bảng này KHÔNG nói mã nguồn hiện có làm được gì. Nó liệt kê những việc "
        "mà một Agent lập trình nhúng phải làm, và cách làm để tự chủ cao nhất."
    ])
    ws["A2"].font = Font(italic=True, color="7F7F7F")
    ws.append([
        "Đề án Thạc sĩ Kỹ thuật (Kỹ thuật Điện tử, PTIT) · "
        "Học viên: Vũ Trí Công · Giảng viên hướng dẫn: TS. Nguyễn Trung Hiếu"
    ])
    ws["A3"].font = Font(italic=True, color="7F7F7F")
    ws.append([])
    ws.append(["Thang mức tự chủ — cột quan trọng nhất của bảng"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=12, color="2F5597")
    ws.append(["Mức", "Nghĩa", "Khi nào dùng"])
    hang_tieu_de = ws.max_row
    for o in ws[hang_tieu_de]:
        o.font = TRANG_DAM
        o.fill = NEN_TIEU_DE
        o.border = VIEN
    for ma, nghia, khi_nao in THANG_TU_CHU:
        ws.append([ma, nghia, khi_nao])
        o = ws.cell(row=ws.max_row, column=1)
        o.fill = PatternFill("solid", fgColor=MAU_TU_CHU[ma])
        o.font = Font(bold=True)
    for i, r in enumerate([10, 42, 82], 1):
        ws.column_dimensions[get_column_letter(i)].width = r
    for hang in ws.iter_rows(min_row=hang_tieu_de, max_row=ws.max_row):
        for o in hang:
            o.alignment = TREN_TRAI
            o.border = VIEN

    ws.append([])
    ws.append(["Cách dùng bảng"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=12, color="2F5597")
    for dong in [
        "1. Đọc sheet 'Nghiệp vụ' — soát xem thiếu việc nào, hoặc việc nào không cần.",
        "2. Soát cột 'Mức tự chủ' — chỗ nào bạn muốn Agent tự chủ hơn (hoặc ít hơn) mức đề xuất.",
        "3. Soát cột 'Cách Agent làm' — đây là chỗ quyết định Agent thông minh tới đâu.",
        "4. Sheet 'Câu hỏi phải biết hỏi' — danh mục khai thác yêu cầu; thiếu câu nào thì bổ sung.",
        "5. Sau khi bạn duyệt, mới đối chiếu với mã hiện có để biết đã làm được bao nhiêu.",
    ]:
        ws.append([dong])
    ws.append([])
    ws.append([
        "Ghi chú: mức tự chủ ở đây là ĐỀ XUẤT của tôi, không phải kết luận. "
        "Nguyên tắc dùng để đề xuất: việc nào sai thì hỏng thật (nạp, cài, xóa, "
        "phong hạng) đều không quá T2."
    ])
    ws[f"A{ws.max_row}"].alignment = TREN_TRAI
    ws[f"A{ws.max_row}"].font = Font(italic=True)

    # ---- Sheet 2: Nghiệp vụ ----
    ws2 = wb.create_sheet("Nghiệp vụ")
    _tieu_de(
        ws2,
        ["Mã", "Giai đoạn", "Nghiệp vụ", "Kết quả cần đạt", "Đầu vào cần có",
         "CÁCH AGENT LÀM (để tự chủ cao)", "Mức tự chủ", "Người quyết gì",
         "Rủi ro nếu làm sai"],
        [8, 26, 34, 40, 30, 74, 11, 26, 48],
    )
    for r in NGHIEP_VU:
        ws2.append(list(r))
    for hang in ws2.iter_rows(min_row=2):
        for o in hang:
            o.alignment = TREN_TRAI
            o.border = VIEN
        o_tc = hang[6]
        if o_tc.value in MAU_TU_CHU:
            o_tc.fill = PatternFill("solid", fgColor=MAU_TU_CHU[o_tc.value])
            o_tc.font = Font(bold=True)
            o_tc.alignment = Alignment(vertical="center", horizontal="center")
        hang[0].font = Font(bold=True)
    ws2.auto_filter.ref = f"A1:I{ws2.max_row}"

    # ---- Sheet 3: Câu hỏi phải biết hỏi ----
    ws3 = wb.create_sheet("Câu hỏi phải biết hỏi")
    _tieu_de(
        ws3,
        ["Giai đoạn", "Chủ đề", "Câu hỏi", "Vì sao phải hỏi"],
        [26, 18, 62, 62],
    )
    for r in CAU_HOI:
        ws3.append(list(r))
    for hang in ws3.iter_rows(min_row=2):
        for o in hang:
            o.alignment = TREN_TRAI
            o.border = VIEN

    # ---- Sheet 4: Thống kê ----
    ws4 = wb.create_sheet("Thống kê")
    ws4.append(["Số nghiệp vụ theo giai đoạn"])
    ws4["A1"].font = Font(bold=True, size=12, color="2F5597")
    ws4.append(["Giai đoạn", "Số việc", "T0", "T1", "T2", "T3", "T4"])
    for o in ws4[2]:
        o.font = TRANG_DAM
        o.fill = NEN_TIEU_DE
        o.border = VIEN

    thu_tu: list[str] = []
    for r in NGHIEP_VU:
        if r[1] not in thu_tu:
            thu_tu.append(r[1])
    for gd in thu_tu:
        hang = [r for r in NGHIEP_VU if r[1] == gd]
        ws4.append([gd, len(hang)] + [sum(1 for r in hang if r[6] == m) for m in
                                      ("T0", "T1", "T2", "T3", "T4")])
    ws4.append(["TỔNG", len(NGHIEP_VU)] + [sum(1 for r in NGHIEP_VU if r[6] == m) for m in
                                            ("T0", "T1", "T2", "T3", "T4")])
    for o in ws4[ws4.max_row]:
        o.font = Font(bold=True)
        o.fill = PatternFill("solid", fgColor="EDEDED")
    for i, r in enumerate([30, 10, 8, 8, 8, 8, 8], 1):
        ws4.column_dimensions[get_column_letter(i)].width = r
    for hang in ws4.iter_rows(min_row=2):
        for o in hang:
            o.border = VIEN
            o.alignment = Alignment(vertical="center", wrap_text=True)

    # ---- Sheet 5: Đối chiếu với mã hiện có ----
    tra = {r[0]: r for r in DOI_CHIEU}
    ws5 = wb.create_sheet("Đối chiếu")
    _tieu_de(
        ws5,
        ["Mã", "Giai đoạn", "Nghiệp vụ", "Trạng thái", "Tự chủ\nĐỀ XUẤT",
         "Tự chủ\nĐẠT", "Khoảng\ncách", "Bằng chứng trong mã", "Còn thiếu gì"],
        [8, 26, 34, 14, 10, 9, 9, 52, 66],
    )
    bac = {"T0": 0, "T1": 1, "T2": 2, "T3": 3, "T4": 4}
    for nv in NGHIEP_VU:
        d = tra[nv[0]]
        khoang = bac[nv[6]] - bac[d[4]]
        ws5.append([nv[0], nv[1], nv[2], d[1], nv[6], d[4],
                    f"−{khoang}" if khoang > 0 else "—", d[2], d[3]])
    for hang in ws5.iter_rows(min_row=2):
        for o in hang:
            o.alignment = TREN_TRAI
            o.border = VIEN
        hang[0].font = Font(bold=True)
        tt = hang[3]
        if tt.value in MAU_TRANG_THAI:
            tt.fill = PatternFill("solid", fgColor=MAU_TRANG_THAI[tt.value])
            tt.font = Font(bold=True, color=MAU_CHU_TT[tt.value])
            tt.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        for i in (4, 5):
            o = hang[i]
            if o.value in MAU_TU_CHU:
                o.fill = PatternFill("solid", fgColor=MAU_TU_CHU[o.value])
                o.alignment = Alignment(vertical="center", horizontal="center")
        kc = hang[6]
        kc.alignment = Alignment(vertical="center", horizontal="center")
        if kc.value != "—":
            kc.font = Font(bold=True, color="9C0006")
    ws5.auto_filter.ref = f"A1:I{ws5.max_row}"

    # ---- Thống kê: thêm phần đối chiếu ----
    ws4.append([])
    ws4.append(["Đối chiếu với mã hiện có"])
    ws4[f"A{ws4.max_row}"].font = Font(bold=True, size=12, color="2F5597")
    ws4.append(["Giai đoạn", "Việc", "Đủ", "Một phần", "Chưa có", "Cố ý không", "Có khoảng cách tự chủ"])
    for o in ws4[ws4.max_row]:
        o.font = TRANG_DAM
        o.fill = NEN_TIEU_DE
        o.border = VIEN
    for gd in thu_tu:
        hang = [r for r in NGHIEP_VU if r[1] == gd]
        d = [tra[r[0]] for r in hang]
        lech = sum(1 for r in hang if bac[r[6]] > bac[tra[r[0]][4]])
        ws4.append([gd, len(hang)] + [sum(1 for x in d if x[1] == t) for t in
                    ("Đủ", "Một phần", "Chưa có", "Cố ý không làm")] + [lech])
    tat_ca = [tra[r[0]] for r in NGHIEP_VU]
    lech_tong = sum(1 for r in NGHIEP_VU if bac[r[6]] > bac[tra[r[0]][4]])
    ws4.append(["TỔNG", len(NGHIEP_VU)] + [sum(1 for x in tat_ca if x[1] == t) for t in
                ("Đủ", "Một phần", "Chưa có", "Cố ý không làm")] + [lech_tong])
    for o in ws4[ws4.max_row]:
        o.font = Font(bold=True)
        o.fill = PatternFill("solid", fgColor="EDEDED")
    for hang in ws4.iter_rows(min_row=2):
        for o in hang:
            o.border = VIEN

    dich.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dich)
    return dich


if __name__ == "__main__":
    dich = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("docs/EAA_Nghiep_vu_Agent.xlsx")
    print(f"Đã ghi: {dung_bang(dich)}")
