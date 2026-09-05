#!/usr/bin/env python3
"""Sinh `docs/EAA_Viec_phai_lam.xlsx` — việc gì · sở cứ · làm thế nào.

Bảng này khác `EAA_So_sanh_Agent_nhung.xlsx` ở chỗ: bảng kia so ta với thị
trường, bảng này nói **phải làm gì** vì kết quả so ấy — và với mỗi việc, nó
buộc phải nêu **sở cứ** cùng **hạng của sở cứ**.

Vì sao cột "hạng sở cứ" là cột quan trọng nhất
-----------------------------------------------

Một danh sách việc phải làm rất dễ trở thành danh sách những thứ người viết
thấy hay. Cách duy nhất chặn được là bắt mỗi việc phải khai nó dựa vào đâu, và
khai chỗ dựa ấy thuộc hạng nào:

* **ĐO** — con số chạy lại được trên chính kho này;
* **SỔ SAI LỆCH** — một lỗi đã xảy ra thật, có mục ghi và bài kiểm canh;
* **VĂN LIỆU** — kết quả đo trong một bài báo có ablation;
* **KHAI (đối thủ)** — thứ nhà cung cấp công bố, ta không chạy thử được;
* **SUY RA** — lập luận của ta, chưa có số.

Hai việc trong bảng có sở cứ hạng SUY RA. Chúng vẫn nằm trong bảng, nhưng nằm
kèm nhãn — vì một việc dựa trên lập luận suông không được lẫn vào việc dựa trên
một lỗi đã xảy ra thật.

Chạy: python3 scripts/lam_bang_viec_phai_lam.py
"""

from __future__ import annotations

import pathlib
import sys

RA = pathlib.Path("docs/EAA_Viec_phai_lam.xlsx")

# ── hạng sở cứ ───────────────────────────────────────────────────────────────

DO = "ĐO"
SO = "SỔ SAI LỆCH"
VL = "VĂN LIỆU"
KHAI = "KHAI (đối thủ)"
SR = "SUY RA"

# ── sổ sở cứ ─────────────────────────────────────────────────────────────────
# (mã, hạng, nội dung, tra ở đâu / chạy lại bằng gì)

SO_CU: list[tuple[str, str, str, str]] = [
    ("SC-01", DO,
     "3 trên 12 lần từ chối tại G3 (đếm lại nhật ký được 13 — xem SL-179) là mã TỰ CHỈNH CHO VỪA ĐỒ ĐO của chính nó, "
     "và cả ba qua sạch bốn cổng máy. Ghi ngày 03/09, trước khi bộ dò nào tồn tại",
     "docs/TIEP_TUC_TU_DAY.md §'1 — N-908'"),
    ("SC-02", DO,
     "Thông báo lỗi nói được việc phải làm: 150/182 = 82%. Mốc trước khi sửa "
     "25/182 = 14% (bản benchmark đầu công bố nhầm 36%)",
     "pytest tests/test_tc145_thong_bao_loi_noi_duoc_viec_phai_lam.py"),
    ("SC-03", DO,
     "Bốn trục đo mới của đề án (độ nhạy bài kiểm · vá chỉnh đồ đo · mất việc "
     "im lặng · truy về được) đều đang báo CHƯA ĐO ĐƯỢC — cái thước đã có, "
     "chưa đo gì",
     "eaa report bench · eaa/bench.py · SL-177"),
    ("SC-04", DO,
     "Kho firmware có 37 commit, 19 lượt sinh, 7 module ⇒ 12 lượt SINH LẠI. "
     "Toàn bộ diễn ra 01–03/09; cả bốn bộ dò ra đời 04/09. Lịch sử không thể "
     "bị chính bộ dò làm nhiễu",
     "git -C projects/robot_balance/firmware log --oneline"),
    ("SC-05", VL,
     "arXiv 2603.19583 — tri thức do NGƯỜI nén theo từng ngoại vi nâng kết quả "
     "lên gần trần. Đây là mốc DUY NHẤT trong khảo sát có ablation đo được",
     "docs/KHAO_SAT_AGENT_NHUNG.md · docs/EAA_Benchmark_San_pham.md Bảng 1"),
    ("SC-06", SO,
     "SL-125 — hồ sơ phần cứng gõ tay lệch với mạch thật. Lỗi đã xảy ra, không "
     "phải lỗi giả định",
     "docs/SAI_LECH_THIET_KE.md mục SL-125"),
    ("SC-07", DO,
     "`eaa observe` chạy trên dự án thật: 9/9 module chưa khai dấu hiệu quan "
     "sát được, và KHÔNG kênh quan sát nào trong hồ sơ phần cứng",
     "eaa observe · SL-175 · TC-135"),
    ("SC-08", KHAI,
     "Embedder cược vào vòng kín trên silicon thật — quan sát phần cứng là "
     "kênh phản hồi CHÍNH, không phải kênh phụ",
     "docs/KHAO_SAT_AGENT_NHUNG.md"),
    ("SC-09", DO,
     "Năm lỗi chỉ phần cứng mới chỉ ra được, và tốc độ bootloader đo trên bo là "
     "57600 chứ không phải 115200 như tài liệu — số đo trên bo cãi lại tài liệu",
     "git log --oneline (abae855, 89c5177)"),
    ("SC-10", DO,
     "Cảnh báo TRÔI hồ sơ phần cứng vẫn đang bật; G1 chưa được duyệt lại. Đây "
     "là việc của NGƯỜI, hệ không được tự bấm",
     "eaa status · docs/TIEP_TUC_TU_DAY.md §'4 — Duyệt lại G1'"),
    ("SC-11", SO,
     "SL-174 — `eaa/dimension.py` CỐ Ý chưa nối vào bản đồ thanh ghi. Phép soi "
     "thứ nguyên đang tra sổ số đo, chưa tra được bề rộng trường bit",
     "docs/SAI_LECH_THIET_KE.md mục SL-174 · eaa/dimension.py"),
    ("SC-12", KHAI,
     "Parasoft · LDRA · QA Systems bán bằng chứng chứng nhận: truy vết hai "
     "chiều yêu cầu ↔ ca kiểm ↔ mã. Đây là mốc gần ta nhất về triết lý",
     "docs/EAA_Benchmark_San_pham.md Bảng 1"),
    ("SC-13", DO,
     "ISR bước 50 kHz trên AVR 16 MHz — chưa ai đo nó ăn bao nhiêu phần trăm "
     "CPU. Một ràng buộc thời gian thực chưa có số",
     "docs/TIEP_TUC_TU_DAY.md §'5 — Đo giữ nhịp trên bo'"),
    ("SC-14", DO,
     "Bảng năng lực sau rà soát vòng hai: 137 ĐỦ · 3 MỘT PHẦN · 4 CHƯA trên "
     "146 dòng. Bốn dòng còn lại: N-036, N-100, N-908, N-909",
     "docs/RA_SOAT_NANG_LUC_04_09.md §7.3"),
    ("SC-15", SR,
     "Cổng nối tiếp là tài nguyên độc chiếm. Hai phiên cùng mở một cổng thì "
     "phiên sau đọc rác — nhưng chưa có ai chạy hai phiên nên CHƯA đo được",
     "lập luận, chưa có số — nhãn SUY RA là cố ý"),
    ("SC-16", SR,
     "Kỹ sư nhúng ngồi trong IDE cả ngày; mọi lệnh của ta đều ở dòng lệnh. "
     "Chưa đo số bước từ mở IDE tới lượt sinh đầu tiên",
     "lập luận, chưa có số — nhãn SUY RA là cố ý"),
    ("SC-17", DO,
     "Lớp truy cập mạng phân hai hạng nguồn theo URL CUỐI sau chuyển hướng "
     "(SL-71..80). Danh sách miền chính chủ hiện nghiêng về pack thứ nhất",
     "eaa/web.py · pytest tests/test_tc65*.py"),
    ("SC-18", DO,
     "56 lệnh cấp một · 103 lệnh đầy đủ · 0 lệnh thiếu dòng trợ giúp · 4 mã "
     "thoát phân biệt. Mặt CLI đã kín; mặt IDE chưa có",
     "eaa --help · eaa/__init__.py"),
]

# ── việc phải làm ────────────────────────────────────────────────────────────
# (mã, việc, hạng việc, sở cứ, làm thế nào, tệp, xong đo bằng gì, bài canh,
#  phụ thuộc, rủi ro, công, ưu tiên, trạng thái)

BIEN_SO = "Biến lời khai thành SỐ"
CHAN_LOI = "Chặn một hạng lỗi"
MO_RONG = "Mở rộng mặt tiếp xúc"
CUA_NGUOI = "Việc của NGƯỜI, hệ không tự bấm"

VIEC: list[tuple[str, ...]] = [
    ("V1",
     "Nâng tỉ lệ thông báo lỗi nói được VIỆC PHẢI LÀM lên ≥ 80%",
     MO_RONG,
     "SC-02",
     "1) Đo lại mốc thật bằng phép đếm ngoặc cân, KHÔNG bằng biểu thức một dạng. "
     "2) Gom gợi ý vào MỘT bảng `GOI_Y_KHI_HONG` chứ không rải vào 182 chuỗi. "
     "3) Gắn bảng ở `main()` — nơi biết người dùng vừa gõ lệnh nào — chứ không "
     "gắn ở chỗ ném lỗi, để lỗi từ hàm phụ trợ vẫn nhận đúng gợi ý. "
     "4) Thông báo đã tự nêu một lệnh thì KHÔNG gắn thêm.",
     "eaa/cli.py",
     "150/182 = 82%, và báo kèm mốc cũ 14%",
     "TC-145 (11 bài, 4 đột biến đều bị bắt)",
     "—",
     "Cám dỗ nới định nghĩa phép đo cho dễ đạt: đổi 'nêu được lệnh cụ thể' "
     "thành 'có tính hành động' thì 14% lên 60% mà không sửa dòng nào",
     "1 ngày",
     "1",
     "XONG (SL-178)"),

    ("V2",
     "Dựng BỘ NHIỆM VỤ cho thước đo, chạy trên ít nhất hai Platform Pack",
     BIEN_SO,
     "SC-03",
     "1) Chọn N nhiệm vụ sinh module, mỗi nhiệm vụ có tiêu chí đúng/sai máy "
     "chấm được. 2) Chạy k lượt mỗi nhiệm vụ để tính pass@k đúng công thức "
     "không chệch. 3) Hạng kết cục lấy từ BÁO CÁO CỦA CHUỖI CỔNG, không lấy "
     "từ trường `status` tự khai. 4) Kết quả trên bo thật và trên máy chủ "
     "đứng RIÊNG, không gộp. 5) Bốn trục mới lấy số từ bộ dò đã có, "
     "`bench.py` chỉ gom.",
     "eaa/bench.py · projects/*/ · packs/avr · packs/stm32",
     "`eaa report bench` ra số thật cho cả bốn trục, trên ≥2 pack",
     "TC-138 đã canh cái thước; cần thêm bài canh bộ nhiệm vụ",
     "Cần ngân sách gọi mô hình thật",
     "Trộn hai hạng bằng chứng (bo thật / máy chủ) rồi báo một con số. "
     "Và cám dỗ chọn nhiệm vụ mà ta biết mình làm tốt",
     "3–5 ngày + token",
     "2",
     "CHỜ NGÂN SÁCH"),

    ("V3",
     "Đo NGƯỢC 12 lượt sinh lại module của lịch sử dự án bằng bốn bộ dò",
     BIEN_SO,
     "SC-04, SC-01",
     "1) Dựng 12 cặp (bản đã merge, bản sinh mới) từ kho firmware lồng. "
     "2) CHỐT CHÂN LÝ NỀN TRƯỚC khi chạy bộ dò, chỉ lấy từ sổ sai lệch và "
     "nhật ký commit viết tại thời điểm ấy — đặc biệt SC-01 đã nêu sẵn 3/12. "
     "3) Chạy bốn bộ dò trên từng cặp. 4) Đếm: tìm ra / BỎ SÓT / báo nhầm. "
     "5) Ghi cả phần bỏ sót vào sổ sai lệch, vì đó mới là phần có giá trị.",
     "scripts/do_nguoc_lich_su.py (mới) · eaa/contract.py · eaa/sensitivity.py "
     "· eaa/instrument.py · eaa/tools/regcheck.py",
     "ĐÃ ĐO: bắt được 3/13 · trúng một phần 1 · kêu trật 7 · BỎ SÓT 0. "
     "Và contract kêu sớm hơn người 54 phút ở ca imu_start_read",
     "TC-146 (8 bài) — canh cả việc chân lý nền KHÔNG bị sửa sau khi chốt",
     "Không tốn token API",
     "Tôi vừa làm bộ dò vừa chấm bộ dò. Nếu chân lý nền được chọn SAU khi "
     "thấy kết quả thì con số đi ra vô nghĩa — nên nó phải được chốt trước "
     "và ghi ra tệp trước",
     "2 ngày",
     "3",
     "XONG (SL-179)"),

    ("V4",
     "Kỹ năng phần cứng theo từng NGOẠI VI, nạp vào prompt và duyệt qua G2",
     CHAN_LOI,
     "SC-05",
     "1) Định dạng một 'kỹ năng': tên ngoại vi, thứ tự thao tác bắt buộc, bẫy "
     "thường gặp, trích dẫn nguồn. 2) Kỹ năng KHÁC chunk tài liệu: chunk là "
     "trích đoạn, kỹ năng là thủ tục đã được người nén. 3) Vào kho qua G2 như "
     "mọi tri thức khác — không có cửa sau. 4) Bật/tắt được, để còn so trước "
     "sau trên cùng bộ nhiệm vụ.",
     "eaa/skills.py (mới) · eaa/composer.py · packs/*/",
     "So pass@k TRƯỚC và SAU khi bật lớp kỹ năng, trên cùng bộ nhiệm vụ V2",
     "TC-139 (đã đặt chỗ trong kế hoạch, chưa viết)",
     "V2 — không có bộ nhiệm vụ thì không so được trước/sau",
     "Kỹ năng viết tay dễ thành hằng số phần cứng lẻn vào `eaa/`. TC-38 phải "
     "chặn: kỹ năng thuộc pack, không thuộc engine",
     "3–4 ngày",
     "4",
     "CHƯA"),

    ("V5",
     "Đọc netlist và ĐỐI CHIẾU với hồ sơ phần cứng tại G1",
     CHAN_LOI,
     "SC-06",
     "1) Đọc netlist/sơ đồ nguyên lý ra một mô hình trung lập. 2) KHÔNG thay "
     "hồ sơ phần cứng — chỉ đối chiếu và nêu chỗ lệch. 3) Chỗ lệch trình ở "
     "G1 cho người phân xử; hệ không tự chọn bên nào đúng. 4) Netlist là "
     "nguồn SUY RA, không phải ĐÃ KIỂM: nó là ý định thiết kế, không phải "
     "mạch đã hàn.",
     "eaa/netlist.py (mới) · eaa/gates.py (G1)",
     "Số chỗ lệch tìm được trên dự án thật",
     "TC-140 (đã đặt chỗ, chưa viết)",
     "—",
     "Tự động 'sửa' hồ sơ theo netlist là làm mất chính thứ G1 sinh ra để "
     "hỏi. Chỉ được nêu lệch",
     "3 ngày",
     "5",
     "CHƯA"),

    ("V6",
     "Nối bộ soi THỨ NGUYÊN vào bản đồ thanh ghi",
     CHAN_LOI,
     "SC-11",
     "1) `dimension.py` hiện tra sổ số đo để lấy đơn vị. 2) Nối thêm nguồn "
     "thứ hai: bề rộng và mặt nạ trường bit từ `regmap.py`. 3) Bắt hạng lỗi "
     "'gán giá trị vượt bề rộng trường' mà chú thích vẫn khai là đúng. "
     "4) Giữ nguyên luật: không tra được thì trả BA TRẠNG THÁI, không ép về "
     "đúng/sai.",
     "eaa/dimension.py · eaa/regmap.py",
     "Số chú thích sai thứ nguyên bắt thêm được trên 12 lượt của V3",
     "TC-134 mở rộng",
     "V3 cho sẵn tập mã để thử",
     "Bộ soi hay báo nhầm sớm muộn cũng bị tắt đi, và lúc ấy nó không bảo vệ "
     "được gì — giữ ba trạng thái là để tránh đúng chuyện đó",
     "1–2 ngày",
     "6",
     "CHƯA"),

    ("V7",
     "Trọng tài phần cứng: khoá cổng nối tiếp giữa các phiên",
     CHAN_LOI,
     "SC-15",
     "1) Khoá theo tệp, có tên phiên và dấu thời gian. 2) Phiên thứ hai bị "
     "chặn kèm thông báo nói RÕ ai đang giữ và làm gì để lấy lại. 3) Khoá "
     "chết phải tự hết hạn — một khoá không bao giờ nhả là một lỗi tệ hơn "
     "lỗi nó chặn.",
     "eaa/serialport.py · eaa/telemetry.py",
     "Chạy hai phiên đồng thời: phiên hai bị chặn với thông báo rõ",
     "Cần TC mới",
     "—",
     "Sở cứ hạng SUY RA — chưa ai chạy hai phiên nên chưa có số. Việc này "
     "vào bảng kèm nhãn ấy, không giả vờ là đã đo",
     "1 ngày",
     "7",
     "CHƯA"),

    ("V8",
     "Đóng nốt hai dòng CHƯA và hai dòng MỘT PHẦN của bảng năng lực",
     BIEN_SO,
     "SC-14",
     "1) N-036 và N-100 đã có cửa vào (`eaa knowledge stale/supersede/"
     "deprecate`, SL-172) — kiểm lại xem còn nhánh nào chưa chạy được. "
     "2) N-908 và N-909 ở mức MỘT PHẦN: phần còn thiếu là phần đòi biết bài "
     "toán. 3) Với hai dòng ấy, quyết một trong hai: lấp nốt, HOẶC chốt rằng "
     "nhánh ấy thuộc về người và ghi lý do — chứ không để treo.",
     "eaa/lifecycle.py · eaa/sensitivity.py · eaa/instrument.py · scripts/lam_bang_nghiep_vu.py",
     "Sheet Khoảng trống còn bao nhiêu dòng, và mỗi dòng có một quyết định",
     "TC-129, TC-131, TC-132 đã có",
     "V3 cho bằng chứng về phần N-908 còn thiếu",
     "Lấp cho đủ số là tệ hơn để trống có lý do. Một dòng chốt 'thuộc về "
     "người' vẫn là một dòng đã đóng",
     "2 ngày",
     "8",
     "CHƯA"),

    ("V9",
     "Truy vết HAI CHIỀU: yêu cầu ↔ ca kiểm ↔ mã ↔ quyết định gate",
     MO_RONG,
     "SC-12",
     "1) Ta đã có một nửa: mỗi SL có TC canh, mỗi lượt sinh có prompt hash và "
     "chunk id. 2) Còn thiếu chiều NGƯỢC: từ một yêu cầu, hỏi ra mọi TC và "
     "mọi dòng mã phục vụ nó. 3) Dựng chỉ mục từ dữ liệu ĐÃ CÓ trong Git, "
     "không bắt người khai thêm. 4) Xuất ra bảng mà người kiểm chứng nhận "
     "đọc được.",
     "eaa/trace.py (mới) · docs/ · eaa/vcs.py",
     "Từ một mã yêu cầu bất kỳ, ra được danh sách TC và commit — kiểm tay 10 mẫu",
     "Cần TC mới",
     "—",
     "Đây là mốc gần ta nhất về triết lý, nên cũng là chỗ dễ khai quá lời "
     "nhất. Ta KHÔNG làm chứng nhận; ta làm truy vết. Phải nói rõ khác biệt",
     "3 ngày",
     "9",
     "CHƯA"),

    ("V10",
     "Ba việc chờ NGƯỜI bấm — không được ủy quyền cho hệ",
     CUA_NGUOI,
     "SC-10, SC-07",
     "1) Duyệt lại G1 cho hồ sơ phần cứng đang bật cảnh báo TRÔI. "
     "2) Khai trường `observable:` trong `hardware_profile.yaml` — hiện 9/9 "
     "module chưa khai dấu hiệu quan sát được. "
     "3) Ghi `ACCEL_BALANCE_OFFSET = -535` vào sổ số đo phần cứng nếu đó là "
     "số đo đã duyệt.",
     "projects/robot_balance/hardware_profile.yaml · projects/robot_balance/board_facts.jsonl",
     "Cảnh báo TRÔI tắt; `eaa observe` không còn báo 0 kênh quan sát",
     "TC-135 đã canh phần hệ",
     "—",
     "Hệ tự bấm ba việc này là phá đúng bất biến 5 cổng người duyệt. Chúng "
     "nằm trong bảng để KHÔNG bị quên, không phải để được tự động hoá",
     "30 phút của người",
     "10",
     "CHỜ NGƯỜI"),

    ("V11",
     "Mở rộng danh sách miền CHÍNH CHỦ sang nhà sản xuất của pack thứ hai",
     MO_RONG,
     "SC-17",
     "1) Danh sách miền chính chủ hiện nghiêng về pack thứ nhất. 2) Thêm "
     "miền của hãng chip pack STM32. 3) KHÔNG nới luật: hạng vẫn tính theo "
     "URL CUỐI sau chuyển hướng, và trang chính chủ vẫn là SUY RA chứ không "
     "phải ĐÃ KIỂM. 4) Thêm miền là thêm DỮ LIỆU, không sửa mã cưỡng chế.",
     "eaa/web.py (bảng miền) · packs/stm32/",
     "Số datasheet chính chủ nạp được cho pack thứ hai",
     "TC-65, TC-66 đã canh luật; chỉ thêm dữ liệu",
     "—",
     "Nới một miền vì tiện là cách danh sách trắng biến thành trang trí. "
     "Thêm miền phải qua cùng cửa như mọi tri thức",
     "0,5 ngày",
     "11",
     "CHƯA"),

    ("V12",
     "Đo tải CPU của ISR bước 50 kHz trên bo thật",
     BIEN_SO,
     "SC-13",
     "1) Một ràng buộc thời gian thực đang không có số. 2) Đo bằng chân GPIO "
     "lật đầu/cuối ISR, đọc bằng máy hiện sóng hoặc bộ đếm. 3) Số đo vào sổ "
     "`board_facts.jsonl` qua G4, rồi chảy ngược vào prompt ở lớp K8. "
     "4) Nếu tải vượt ngưỡng thì đó là ràng buộc thiết kế, không phải lỗi mã.",
     "projects/robot_balance/ · projects/robot_balance/board_facts.jsonl",
     "Một con số phần trăm CPU, có xuất xứ và đơn vị, đã duyệt qua G4",
     "TC-133 đã canh đường số đo vào prompt",
     "Cần bo và máy đo",
     "Số đo trên bo đã cãi lại tài liệu một lần (bootloader 57600 chứ không "
     "115200). Đo bằng suy luận thay vì bằng que đo là lặp lại chính lỗi ấy",
     "0,5 ngày + thiết bị",
     "12",
     "CHƯA"),

    ("V13",
     "Lớp mỏng tích hợp IDE",
     MO_RONG,
     "SC-16, SC-18",
     "1) Mặt CLI đã kín: 103 lệnh, 0 lệnh thiếu trợ giúp, 4 mã thoát phân "
     "biệt. 2) Lớp IDE gọi THẲNG các lệnh ấy, không dựng đường đi thứ hai — "
     "một đường thứ hai tới merge là phá bất biến số một. 3) Giữ mỏng: hiện "
     "trạng thái, mở gate, xem báo cáo. Không sinh mã trong IDE.",
     "tích hợp riêng, không đụng eaa/",
     "Số bước từ mở IDE tới lượt sinh đầu tiên",
     "Bất biến 'không có nhánh thứ hai dẫn tới merge' đã có TC-01, TC-02",
     "—",
     "Sở cứ hạng SUY RA. Và rủi ro thật: một lớp IDE 'tiện' rất dễ mọc ra "
     "đường tắt vòng qua gate",
     "4–5 ngày",
     "13",
     "CHƯA"),

    ("V14",
     "Điều khiển máy đo và phiên gỡ lỗi tự động",
     MO_RONG,
     "SC-08, SC-09",
     "1) Đây là khoảng cách phần cứng lớn nhất so với mốc Embedder. 2) Tốn "
     "THIẾT BỊ chứ không chỉ tốn mã. 3) Chỉ có nghĩa khi đã có thước (V2) để "
     "chứng minh nó cải thiện được gì — làm trước V2 là thêm tính năng không "
     "chứng minh được.",
     "eaa/instrumentctl.py (mới) · packs/*/",
     "Số kịch bản chẩn đoán chạy được không cần người cầm que đo",
     "Cần TC mới",
     "V2 — nếu không thì không đo được nó có cải thiện gì",
     "Làm sớm là đúng cái bẫy bảng này sinh ra để tránh: thêm năng lực không "
     "ai có, rồi không có số nào chứng minh nó tốt hơn",
     "1–2 tuần + thiết bị",
     "14",
     "CHƯA"),
]

COT = [
    ("Mã", 7),
    ("Việc — làm gì", 46),
    ("Hạng việc", 17),
    ("Sở cứ", 11),
    ("Làm NHƯ THẾ NÀO", 68),
    ("Đụng tệp nào", 30),
    ("Xong thì ĐO bằng gì", 34),
    ("Bài kiểm canh", 26),
    ("Phụ thuộc", 18),
    ("Rủi ro — cái dễ làm hỏng", 44),
    ("Công", 14),
    ("Ưu tiên", 8),
    ("Trạng thái", 16),
]

# ── bốn mốc đối thủ ──────────────────────────────────────────────────────────

MOC: list[tuple[str, str, str, str]] = [
    ("Embedder",
     "Vòng kín trên silicon thật — quan sát phần cứng là kênh phản hồi CHÍNH",
     "V12, V14",
     "Ta có đường đo trên bo và số đo chảy ngược vào prompt (K8), nhưng chưa "
     "tự điều khiển máy đo. Khoảng cách thật, và nó tốn thiết bị"),
    ("Skilled AI Agents (arXiv 2603.19583)",
     "Tri thức do NGƯỜI nén lại theo từng ngoại vi",
     "V4",
     "Mốc DUY NHẤT có ablation đo được — nên đây là việc duy nhất trong bảng "
     "mà ta biết trước nó nâng được kết quả bao nhiêu"),
    ("STM32 Sidekick · MCP Espressif",
     "Thẩm quyền trên tài liệu của hãng",
     "V11",
     "Ta hơn ở chỗ NÓI RA mức tin cậy (23 lớp mang nhãn, TC-63) trong khi "
     "Sidekick dặn người dùng tự đối chiếu. Ta kém ở BỀ RỘNG nguồn"),
    ("Parasoft · LDRA · QA Systems",
     "Bằng chứng chứng nhận, truy vết hai chiều",
     "V9",
     "Gần ta nhất về triết lý. Ta có một nửa chiều truy vết; ta KHÔNG làm "
     "chứng nhận và không nên khai là có"),
]

# ── bảng này KHÔNG nói gì ────────────────────────────────────────────────────

KHONG_NOI: list[str] = [
    "Bảng này KHÔNG cho điểm tổng. Sở cứ thuộc năm hạng khác nhau, và cộng "
    "chúng lại thành một điểm là làm ra một con số sai theo hướng có lợi cho "
    "người viết — tức hướng khó tự phát hiện nhất.",
    "Cột Công là ƯỚC LƯỢNG, không phải cam kết. Nó chưa từng được đo lại sau "
    "khi làm xong việc nào.",
    "Hai việc (V7 và V13) dựa trên sở cứ hạng SUY RA — "
    "lập luận của ta, chưa có số. Chúng nằm trong bảng kèm nhãn, không lẫn vào "
    "việc dựa trên lỗi đã xảy ra thật.",
    "Ba mốc đối thủ được mô tả từ tài liệu họ CÔNG BỐ. Không sản phẩm nào "
    "trong số đó được chạy thử trên cùng bài toán với ta.",
    "Thứ tự ưu tiên xếp theo 'nuôi được chiều sâu bao nhiêu', không theo 'dễ "
    "làm bao nhiêu'. V2 và V3 đứng trên vì không có số thì mọi việc sau đều "
    "không chứng minh được gì.",
    "Một việc XONG trong cột Trạng thái nghĩa là có bài kiểm canh nó, không "
    "phải là đã chạy thử một lần.",
]


def main() -> int:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Cần openpyxl: pip install openpyxl", file=sys.stderr)
        return 1

    DAM = Font(bold=True, color="FFFFFF")
    NEN_TIEU_DE = PatternFill("solid", fgColor="1F3864")
    TREN = Alignment(vertical="top", wrap_text=True)
    GIUA = Alignment(horizontal="center", vertical="top", wrap_text=True)
    VIEN = Border(*(Side(style="thin", color="BFBFBF"),) * 4)

    MAU_HANG = {
        DO: PatternFill("solid", fgColor="C6EFCE"),
        SO: PatternFill("solid", fgColor="DDEBF7"),
        VL: PatternFill("solid", fgColor="E4DFEC"),
        KHAI: PatternFill("solid", fgColor="FFF2CC"),
        SR: PatternFill("solid", fgColor="FFC7CE"),
    }
    MAU_TT = {
        "XONG (SL-178)": PatternFill("solid", fgColor="C6EFCE"),
        "ĐANG LÀM": PatternFill("solid", fgColor="FFF2CC"),
        "CHỜ NGÂN SÁCH": PatternFill("solid", fgColor="E7E6E6"),
        "CHỜ NGƯỜI": PatternFill("solid", fgColor="DDEBF7"),
        "CHƯA": PatternFill("solid", fgColor="FFC7CE"),
    }

    wb = Workbook()

    def dau_bang(ws, cot: list[tuple[str, int]]) -> None:
        for i, (ten, rong) in enumerate(cot, 1):
            o = ws.cell(row=1, column=i, value=ten)
            o.font, o.fill, o.alignment, o.border = DAM, NEN_TIEU_DE, GIUA, VIEN
            ws.column_dimensions[get_column_letter(i)].width = rong
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

    # ── sheet 1: đọc trước ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Đọc trước"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 112
    doc: list[tuple[str, str]] = [
        ("EAA — việc phải làm để tốt hơn thị trường", ""),
        ("", ""),
        ("Bảng này trả lời gì",
         "Ba câu, cho từng việc: LÀM GÌ · DỰA VÀO ĐÂU · LÀM THẾ NÀO. "
         "Sinh ra từ bản benchmark sản phẩm và bản rà soát năng lực, "
         "không phải từ ý thích."),
        ("Cột quan trọng nhất",
         "Cột 'Sở cứ' và sheet 'Sở cứ'. Một danh sách việc phải làm rất dễ "
         "thành danh sách những thứ người viết thấy hay; cách duy nhất chặn "
         "được là bắt mỗi việc khai nó dựa vào đâu, và chỗ dựa ấy thuộc hạng nào."),
        ("Năm hạng sở cứ",
         "ĐO — con số chạy lại được trên chính kho này.  "
         "SỔ SAI LỆCH — một lỗi đã xảy ra thật, có mục ghi và bài kiểm canh.  "
         "VĂN LIỆU — kết quả trong một bài báo có ablation.  "
         "KHAI (đối thủ) — nhà cung cấp công bố, ta không chạy thử được.  "
         "SUY RA — lập luận của ta, chưa có số."),
        ("Vì sao không có điểm tổng",
         "Cộng năm hạng bằng chứng khác nhau thành một con số là làm ra một "
         "con số sai theo hướng có lợi cho người viết — tức hướng khó tự phát "
         "hiện nhất."),
        ("Thứ tự ưu tiên xếp theo gì",
         "Theo 'nuôi được chiều sâu bao nhiêu', không theo 'dễ làm bao nhiêu'. "
         "V2 và V3 đứng trên vì không có số thì mọi việc sau đều không chứng "
         "minh được gì."),
        ("", ""),
        ("Sinh lại bảng này", "python3 scripts/lam_bang_viec_phai_lam.py"),
        ("Đọc cùng",
         "docs/EAA_Benchmark_San_pham.docx · docs/RA_SOAT_NANG_LUC_04_09.md · "
         "docs/KE_HOACH_VUOT_LEN.md · docs/SAI_LECH_THIET_KE.md"),
    ]
    for r, (a, b) in enumerate(doc, 1):
        ws.cell(row=r, column=1, value=a).font = Font(bold=True, size=14 if r == 1 else 11)
        ws.cell(row=r, column=1).alignment = TREN
        ws.cell(row=r, column=2, value=b).alignment = TREN

    # ── sheet 2: việc phải làm ───────────────────────────────────────────────
    ws = wb.create_sheet("Việc phải làm")
    dau_bang(ws, COT)
    for r, hang in enumerate(VIEC, 2):
        for c, gia_tri in enumerate(hang, 1):
            o = ws.cell(row=r, column=c, value=gia_tri)
            o.alignment = GIUA if c in (1, 11, 12) else TREN
            o.border = VIEN
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=13).fill = MAU_TT.get(hang[12], PatternFill())
        ws.row_dimensions[r].height = 118
    ws.auto_filter.ref = f"A1:M{len(VIEC) + 1}"

    # ── sheet 3: sổ sở cứ ────────────────────────────────────────────────────
    ws = wb.create_sheet("Sở cứ")
    dau_bang(ws, [("Mã", 8), ("Hạng", 16), ("Sở cứ nói gì", 84),
                  ("Tra ở đâu / chạy lại bằng gì", 56), ("Việc dùng nó", 18)])
    dung: dict[str, list[str]] = {}
    for v in VIEC:
        for ma in v[3].split(", "):
            dung.setdefault(ma, []).append(v[0])
    for r, (ma, hang, noi_dung, tra) in enumerate(SO_CU, 2):
        for c, gia_tri in enumerate(
            (ma, hang, noi_dung, tra, ", ".join(dung.get(ma, ["—"]))), 1
        ):
            o = ws.cell(row=r, column=c, value=gia_tri)
            o.alignment = GIUA if c in (1, 2, 5) else TREN
            o.border = VIEN
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=2).fill = MAU_HANG.get(hang, PatternFill())
        ws.row_dimensions[r].height = 58
    ws.auto_filter.ref = f"A1:E{len(SO_CU) + 1}"

    # ── sheet 4: mốc đối thủ ─────────────────────────────────────────────────
    ws = wb.create_sheet("Mốc đối thủ")
    dau_bang(ws, [("Sản phẩm", 34), ("Cốt lõi nó cược vào", 52),
                  ("Việc của ta trả lời nó", 20), ("Ta đứng đâu — thẳng thắn", 76)])
    for r, hang in enumerate(MOC, 2):
        for c, gia_tri in enumerate(hang, 1):
            o = ws.cell(row=r, column=c, value=gia_tri)
            o.alignment = GIUA if c == 3 else TREN
            o.border = VIEN
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.row_dimensions[r].height = 68

    # ── sheet 5: bảng này KHÔNG nói gì ───────────────────────────────────────
    ws = wb.create_sheet("KHÔNG nói gì")
    dau_bang(ws, [("#", 6), ("Điều bảng này KHÔNG chứng minh", 122)])
    for r, dong in enumerate(KHONG_NOI, 2):
        ws.cell(row=r, column=1, value=r - 1).alignment = GIUA
        o = ws.cell(row=r, column=2, value=dong)
        o.alignment, o.border = TREN, VIEN
        ws.row_dimensions[r].height = 50

    RA.parent.mkdir(parents=True, exist_ok=True)
    wb.save(RA)

    hang_dem: dict[str, int] = {}
    for _, h, _, _ in SO_CU:
        hang_dem[h] = hang_dem.get(h, 0) + 1
    print(f"Đã ghi {RA}  ({RA.stat().st_size:,} byte)")
    print(f"  việc  : {len(VIEC)} — "
          + " · ".join(f"{t}: {sum(1 for v in VIEC if v[12] == t)}"
                       for t in ("XONG (SL-178)", "ĐANG LÀM", "CHỜ NGÂN SÁCH",
                                 "CHỜ NGƯỜI", "CHƯA")))
    print(f"  sở cứ : {len(SO_CU)} — "
          + " · ".join(f"{k}: {v}" for k, v in hang_dem.items()))
    print(f"  sheet : {len(wb.sheetnames)} — {' · '.join(wb.sheetnames)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
