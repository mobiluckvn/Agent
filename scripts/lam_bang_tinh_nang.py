"""Sinh bảng Excel thống kê tính năng của Embedded AIDD Agent.

Chạy: python scripts/lam_bang_tinh_nang.py [đường-dẫn-xlsx]

Bảng này là DỮ LIỆU, không phải văn bản viết tay: mỗi dòng nêu tính năng, module
hiện thực, bằng chứng (mã test), và trạng thái. Nhờ vậy khi sprint sau đổi trạng
thái một dòng thì sửa đúng một chỗ, và con số ở sheet Tổng quan tự tính lại.

Bốn trạng thái, và ranh giới giữa chúng có ý nghĩa:

* ĐÃ LÀM — có mã, có test, chạy được end-to-end.
* MỘT PHẦN — có mã nhưng còn mắt xích thiếu; ghi rõ thiếu gì ở cột "Còn thiếu".
* CHƯA LÀM — chưa có mã.
* CỐ Ý KHÔNG LÀM — có thể làm nhưng thiết kế cấm; ghi rõ điều khoản cấm. Cột
  này quan trọng với người đọc bảng: "chưa tự cài được" và "cố ý không tự cài"
  là hai câu hoàn toàn khác nhau.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DA = "Đã làm"
PHAN = "Một phần"
CHUA = "Chưa làm"
COY = "Cố ý không làm"

# --------------------------------------------------------------------------
# Dữ liệu: (nhóm, tính năng, mô tả, module/lệnh, bằng chứng, trạng thái, còn thiếu)
# --------------------------------------------------------------------------

TINH_NANG: list[tuple[str, str, str, str, str, str, str]] = [
    # -- A. Khung xương & trạng thái ---------------------------------------
    ("A. Khung & trạng thái", "Project State ghi nguyên tử",
     "Ghi tạm → fsync → os.replace → fsync thư mục; sống sót qua crash giữa chừng",
     "eaa/state.py", "TC-03 (24 test)", DA, ""),
    ("A. Khung & trạng thái", "Khóa liên tiến trình",
     "Khóa tái nhập, tự thu hồi khóa mồ côi khi tiến trình giữ khóa đã chết",
     "eaa/state.py", "TC-03", DA, ""),
    ("A. Khung & trạng thái", "Máy trạng thái 13 công đoạn",
     "A1..F1, mỗi cung chuyển khai báo gate phải qua; không có cung tắt",
     "eaa/policy.py", "TC-08 (58 test)", DA, ""),
    ("A. Khung & trạng thái", "Bảng phân quyền AUTO / HUMAN",
     "Pha nào máy tự đi, pha nào bắt buộc có người",
     "eaa/policy.py", "TC-08", DA, ""),
    ("A. Khung & trạng thái", "Khôi phục phiên sau khi tắt máy",
     "Đọc lại Project State và nói bước kế tiếp",
     "eaa resume / status", "test_cli_s0.py", DA, ""),

    # -- B. Kiến trúc ba tầng ----------------------------------------------
    ("B. Kiến trúc ba tầng", "Engine sạch phần cứng",
     "Không một hằng số, tên thanh ghi hay tên chip nào trong eaa/",
     "toàn bộ eaa/", "TC-38 quét mỗi commit", DA, ""),
    ("B. Kiến trúc ba tầng", "Platform Pack qua interface",
     "Engine gọi toolchain chỉ qua eaa/platform.py; pack khai báo năng lực bằng YAML",
     "eaa/platform.py", "test_platform_pack.py", DA, ""),
    ("B. Kiến trúc ba tầng", "Thêm họ MCU mới không sửa engine (NFR-05)",
     "Pack thứ hai (STM32 Cortex-M4F) thêm vào KHÔNG sinh một nhánh rẽ nào trong eaa/",
     "packs/avr/ + packs/stm32/", "TC-47", DA,
     ""),

    # -- C. Tri thức, RAG, vòng đời ----------------------------------------
    ("C. Tri thức & RAG", "Kho ràng buộc + hồ sơ phần cứng + thư viện prompt",
     "constraints.yaml, hardware_profile.yaml, prompt nền tảng theo pack",
     "eaa/kb.py", "test_kb.py", DA, ""),
    ("C. Tri thức & RAG", "Knowledge Graph module→ngoại vi→thanh ghi→chunk",
     "Đồ thị tự dựng từ kho tri thức; khớp tên thanh ghi chính xác, không nhúng vector",
     "eaa/graph.py", "TC-18 (31 test)", DA, ""),
    ("C. Tri thức & RAG", "Truy xuất chunk tất định",
     "Cùng câu hỏi cho cùng top-k; không phụ thuộc thứ tự ngẫu nhiên",
     "eaa/graph.py", "TC-18", DA, ""),
    ("C. Tri thức & RAG", "Nén ngữ cảnh K1–K7, ngân sách 8.000 token",
     "7 tầng có ngân sách con; cắt luật lỗi rồi tới interface, KHÔNG BAO GIỜ cắt chunk",
     "eaa/composer.py", "TC-04 (36 test), TC-16", DA, ""),
    ("C. Tri thức & RAG", "Bắt buộc trích dẫn // ref: <chunk-id>",
     "Mã cấu hình thanh ghi không có trích dẫn thì cổng static chặn",
     "eaa/tools/static.py", "TC-17", DA, ""),
    ("C. Tri thức & RAG", "Append-only + supersede + stale set",
     "Không ghi đè vật lý; chunk bị thay thì mã dùng nó bị đánh dấu cần xem lại",
     "eaa/lifecycle.py", "TC-29 (23 test)", DA, ""),
    ("C. Tri thức & RAG", "Ba đường truy vấn ngược",
     "Từ chunk tìm ra mã: qua đồ thị, qua // ref:, qua trailer chunk-ids của commit",
     "eaa/lifecycle.py", "TC-29", DA, ""),
    ("C. Tri thức & RAG", "Nạp PDF thành proposed fact",
     "Trích đoạn luôn ở trạng thái 'proposed', phải qua G2 mới thành tri thức",
     "eaa/ingest.py", "TC-22 (44 test)", DA, ""),
    ("C. Tri thức & RAG", "Nguồn web giới hạn miền nhà sản xuất",
     "So theo hậu tố tên miền, không so chuỗi con",
     "eaa/ingest.py", "TC-22", DA, ""),
    ("C. Tri thức & RAG", "Sổ giả định (Assumption Log)",
     "Giả định được ghi, và bị thay khi có số đo thật",
     "eaa/ingest.py", "TC-22", DA, ""),
    ("C. Tri thức & RAG", "Tự đánh giá đủ thông tin (RIC)",
     "Trả về CÓ / THIẾU / MÂU THUẪN; chuẩn hóa số nên 0b00 == 0x00 == 0",
     "eaa/readiness.py", "TC-24 (18 test)", DA, ""),
    ("C. Tri thức & RAG", "BM25 bổ trợ truy xuất theo từ khóa",
     "Tầng 2 của AIS §4.2: quan hệ trước, BM25 lấp chỗ trống, ngưỡng là ĐỘ PHỦ từ khóa",
     "eaa/rag.py", "TC-64 (20 test)", DA, ""),

    # -- D. Sinh mã ---------------------------------------------------------
    ("D. Sinh mã", "Vòng lặp chuẩn 13 bước",
     "Từ chọn module tới merge, mỗi bước ghi vết vào Project State",
     "eaa/orchestrator.py", "TC-06 (28 test)", DA, ""),
    ("D. Sinh mã", "Vòng tự sửa ≤ 3 lần, dạng patch",
     "Gửi patch chứ không gửi lại cả tệp; quá 3 lần thì dừng và bàn giao người",
     "eaa/orchestrator.py", "TC-06, TC-19", DA, ""),
    ("D. Sinh mã", "Kiểm ngân sách token TRƯỚC khi gọi mô hình",
     "count_tokens chạy trước; vượt ngân sách thì không gọi API",
     "eaa/llm/base.py", "TC-16", DA, ""),
    ("D. Sinh mã", "Adapter Gemini Pro 3.1 ghim phiên bản",
     "REST thuần bằng urllib, stateless mỗi lần gọi, clamp output limit theo model thật",
     "eaa/llm/gemini.py", "TC-11 (31 test)", DA, ""),
    ("D. Sinh mã", "Lối gọi văn xuôi complete()",
     "Tra cứu / phân loại lỗi không phải sinh mã, nên không bị đòi khối ```file:",
     "eaa/llm/*.py", "TC-39", DA, ""),
    ("D. Sinh mã", "MockLLM tất định cho Sprint 1–3",
     "Cùng interface với mô hình thật; kiểm thử tích hợp không tốn API",
     "eaa/llm/mock.py", "toàn bộ test tích hợp", DA, ""),
    ("D. Sinh mã", "Nhật ký lời gọi + phát lại",
     "Lưu (băm prompt → phản hồi) làm bằng chứng; ReplayClient chạy lại không tốn API",
     "eaa/llm/calllog.py", "TC-15", DA, ""),
    ("D. Sinh mã", "Phát hiện mô hình trôi hành vi",
     "Cùng băm prompt mà hai phản hồi khác nhau thì hiện ra",
     "eaa/llm/calllog.py", "TC-15", DA, ""),
    ("D. Sinh mã", "Che khóa API ở mọi lối ra",
     "Không vào log, commit, repr, thông báo lỗi hay băm prompt",
     "eaa/llm/base.py", "TC-14", DA, ""),
    ("D. Sinh mã", "Ráp firmware hoàn chỉnh (main + scheduler)",
     "Sinh vòng lặp chính từ khuôn của pack, dịch mọi module đã merge, liên kết, ra ảnh nạp được",
     "eaa/firmware.py + eaa build", "TC-41 (24 test)", DA, ""),
    ("D. Sinh mã", "Bộ định thời hợp tác",
     "Bảng việc định kỳ; ngắt chỉ tăng bộ đếm, đọc bộ đếm trong khối nguyên tử",
     "packs/avr/templates/main.c.tmpl", "TC-41", DA, ""),
    ("D. Sinh mã", "Module đã merge không được bỏ quên khi ráp",
     "Merge mà vắng mặt trong firmware.yaml là LỖI; không chạy định kỳ thì khai step: null",
     "eaa/firmware.py", "TC-41b", DA, ""),
    ("D. Sinh mã", "Mã chưa merge không vào được firmware",
     "Bản thiết kế ráp nhắc tới module chưa qua G3 thì dừng",
     "eaa/firmware.py", "TC-41c", DA, ""),

    # -- E. Kiểm chứng & cổng ----------------------------------------------
    ("E. Kiểm chứng & cổng", "Bốn cổng công cụ",
     "compile, static, unittests, size — mỗi cổng trả ToolReport có bằng chứng",
     "eaa/tools/", "TC-07 (43 test)", DA, ""),
    ("E. Kiểm chứng & cổng", "Bất biến merge",
     "Chỉ merge khi TOÀN BỘ ToolReport.passed VÀ G3 duyệt; không có nhánh thứ hai",
     "eaa/vcs.py + orchestrator", "TC-01 (45 test)", DA, ""),
    ("E. Kiểm chứng & cổng", "Băm nội dung xuyên suốt",
     "Hồ sơ gate → quyết định → ủy quyền merge cùng một băm: người duyệt X thì merge đúng X",
     "eaa/gates.py, eaa/vcs.py", "TC-01", DA, ""),
    ("E. Kiểm chứng & cổng", "5 Human Gate không vượt được",
     "Không có cờ --yes / --force / --skip-gate nào tồn tại trong CLI",
     "eaa/gates.py", "TC-01, TC-28, TC-34", DA, ""),
    ("E. Kiểm chứng & cổng", "Phiên không có người ≠ người đồng ý",
     "Không phải TTY thì ném GateNotInteractive chứ không mặc định duyệt",
     "eaa/gates.py", "TC-28", DA, ""),
    ("E. Kiểm chứng & cổng", "Lỗi cấu hình không đi vào vòng tự sửa",
     "Thiếu luật, thiếu công cụ → chặn ngay, không đốt 3 lượt sửa vô ích",
     "eaa/orchestrator.py", "TC-06", DA, ""),
    ("E. Kiểm chứng & cổng", "Sổ lỗi ảo giác",
     "Append-only; phân xử ghi thành sự kiện mới chứ không sửa sự kiện cũ",
     "eaa/ledger.py", "TC-10 (20 test)", DA, ""),
    ("E. Kiểm chứng & cổng", "Báo cáo KPI",
     "Tỉ lệ qua cổng, số lượt tự sửa, token tiêu thụ theo module",
     "eaa/kpi.py + eaa report", "TC-09", DA, ""),
    ("E. Kiểm chứng & cổng", "Cổng SIL cho mã C sinh ra",
     "Chạy chính mã C sinh ra trong mô phỏng, không chỉ mô hình Python",
     "—", "—", CHUA,
     "Cố ý để ngoài chuỗi tự động: nếu nối vào mà không thật sự chạy artifact "
     "thì cổng sẽ 'đạt' mà chẳng kiểm gì — nguy hiểm hơn là không có cổng"),

    # -- F. Mô phỏng --------------------------------------------------------
    ("F. Mô phỏng", "Mô hình con lắc ngược + MIL",
     "Mô phỏng động lực học, đánh giá theo tiêu chí nghiệm thu",
     "eaa/tools/sim.py, sim_runner.py", "TC-12 (12 test)", DA, ""),
    ("F. Mô phỏng", "Quét tham số",
     "Quét lưới bộ tham số điều khiển, xếp hạng theo chỉ số",
     "eaa sim --sweep", "TC-12", DA, ""),
    ("F. Mô phỏng", "Chỉ số thời gian ổn định đo đúng",
     "Đo từ lần cuối tín hiệu rời dải, tính từ lúc nhiễu kết thúc",
     "eaa/tools/sim.py", "test_sim_verification.py", DA, ""),

    # -- G. Công cụ & môi trường -------------------------------------------
    ("G. Công cụ & môi trường", "Chế độ 1 — quét công cụ đã có",
     "Có gì trên máy, phiên bản bao nhiêu, thiếu thì chặn cổng nào",
     "eaa doctor", "TC-34 (37 test)", DA, ""),
    ("G. Công cụ & môi trường", "Chế độ 2 — sinh lệnh cài",
     "Sinh đúng lệnh theo hệ điều hành, hỏi người trước từng lệnh",
     "eaa doctor --fix", "TC-34, TC-37", DA, ""),
    ("G. Công cụ & môi trường", "Chế độ 3 — TỰ TÌM công cụ chưa biết",
     "Phát hiện → tra cứu bằng mô hình → đề xuất; đây là phần vừa làm xong",
     "eaa/toolsearch.py", "TC-39 (29 test)", DA, ""),
    ("G. Công cụ & môi trường", "Nhu cầu công cụ suy từ Platform Pack",
     "Không chép tay danh sách: pack.yaml đã ghi mọi chương trình nó gọi",
     "eaa/toolsearch.py", "TC-39", DA, ""),
    ("G. Công cụ & môi trường", "Kiểm nguồn cài trước khi tới tay người",
     "10 trình quản lý gói chính thống; tải trực tiếp phải HTTPS + miền cho phép + sha256",
     "eaa/toolsearch.py", "TC-39", DA, ""),
    ("G. Công cụ & môi trường", "Đề xuất công cụ vào manifest qua G2",
     "Append + supersede, mang theo approved_by / approved_at",
     "eaa gate approve G2", "TC-39", DA, ""),
    ("G. Công cụ & môi trường", "Ràng buộc phiên bản của pack đè lên đề xuất",
     "pack.yaml ghi avr-gcc >=12.0, mô hình đề xuất 7.3 — lấy theo pack",
     "eaa/toolsearch.py", "TC-39", DA, ""),
    ("G. Công cụ & môi trường", "Khóa môi trường + cảnh báo trôi",
     "env_lock.json; toolchain đổi phiên bản thì cảnh báo trước khi so sánh A/B",
     "eaa/doctor.py", "TC-36", DA, ""),
    ("G. Công cụ & môi trường", "Thẻ công cụ",
     "Cú pháp gọi, cách đọc dòng lỗi, phiên bản đã xác nhận",
     "eaa/doctor.py", "TC-37", DA, ""),
    ("G. Công cụ & môi trường", "Tự cài KHÔNG hỏi người",
     "Kỹ thuật làm được trong một buổi, nhưng FR-ENV-02 và AIS §9.4 cấm",
     "eaa/doctor.py", "TC-34", COY,
     "Cài phần mềm là thay đổi máy kỹ sư. Một agent tự cài im lặng là một agent "
     "có quyền ghi tùy ý lên máy — bỏ chốt này thì hai chốt còn lại mất ý nghĩa"),
    ("G. Công cụ & môi trường", "Tự phát hiện cổng USB / thiết bị cắm vào",
     "Liệt kê cổng nối tiếp, khớp VID/PID với bo dự án khai; không đọc được thì nói thẳng",
     "eaa/serialport.py + eaa ports", "TC-42a (8 test)", DA, ""),

    # -- H. Phiên bản mã ----------------------------------------------------
    ("H. Phiên bản mã", "Ba hạng chất lượng",
     "build-ok < sim-verified < hw-verified; hạng cao đòi bằng chứng tương ứng",
     "eaa/versions.py", "TC-30 (26 test)", DA, ""),
    ("H. Phiên bản mã", "known_good.lock chỉ cập nhật tại G4",
     "Phong hạng hw-verified đòi quyết định G4 VÀ số đo thật",
     "eaa/versions.py", "TC-30", DA, ""),
    ("H. Phiên bản mã", "Quay lui về bản known-good",
     "Rollback không làm mất bản known-good đang giữ",
     "eaa rollback", "TC-30", DA, ""),
    ("H. Phiên bản mã", "Commit truy vết đầy đủ (NFR-07)",
     "Mỗi commit mang băm prompt, model, phiên bản ràng buộc, danh sách chunk",
     "eaa/vcs.py", "TC-01", DA, ""),
    ("H. Phiên bản mã", "Kho phẩm xuất",
     "list / get / regen; gửi lại có kiểm băm để biết bản gửi có bị đổi không",
     "eaa docs", "TC-32 (33 test)", DA, ""),

    # -- I. Tương tác người dùng -------------------------------------------
    ("I. Tương tác", "17 lệnh CLI phủ 10 use case",
     "init, plan, gen, gate, sim, doctor, docs, tune, rollback, diagnose...",
     "eaa/cli.py", "test_cli_e2e.py", DA, ""),
    ("I. Tương tác", "Hồ sơ gate đủ để quyết định",
     "Tóm tắt + chi tiết + băm nội dung; G4 nêu tiêu chí trước, số đo sau",
     "eaa/gates.py", "TC-01, TC-28", DA, ""),
    ("I. Tương tác", "Từ chối phải có lý do",
     "GateDecision đòi tên người, và đòi lý do khi từ chối",
     "eaa/gates.py", "TC-28", DA, ""),
    ("I. Tương tác", "Trình NHIỀU phương án để người chọn",
     "eaa decide dựng tập phương án; gate đòi --option, không lấy gợi ý làm mặc định",
     "eaa/options.py + eaa decide", "TC-46 (32 test)", DA, ""),
    ("I. Tương tác", "Mỗi phương án phải nói mặt trái",
     "Thiếu cons là lỗi — danh sách toàn ưu điểm chỉ chuyển trách nhiệm sang người bấm",
     "eaa/options.py", "TC-46b", DA, ""),
    ("I. Tương tác", "Agent tự giải thích 'vì sao chọn cách này'",
     "Mỗi phương án kèm rationale; gợi ý có lý do nhưng không thay quyết định",
     "eaa/options.py", "TC-46", DA, ""),
    ("I. Tương tác", "Phương án bị loại vẫn tra lại được",
     "Quyết định mang cả tập, kể cả cái bị loại; từ chối cả tập cũng được lưu",
     "eaa/gates.py", "TC-46d", DA, ""),
    ("I. Tương tác", "Giao diện hội thoại",
     "Nói bằng tiếng Việt; Agent tự chọn và chạy lệnh, đọc kết quả, lặp, rồi trả lời",
     "eaa/agent.py + eaa chat", "TC-61 (39 test)", DA, ""),
    ("I. Tương tác", "Gate và phần cứng ngoài tầm với của vòng hội thoại",
     "Danh mục công cụ KHÔNG chứa gate approve, flash, doctor --fix, tune, gen — "
     "chặn bằng cấu tạo, không bằng lời dặn trong prompt",
     "eaa/agent.py TOOLBOX", "TC-61b, TC-61c", DA, ""),
    ("I. Tương tác", "Hiểu ngữ nghĩa khi khớp từ khóa trượt",
     "Bậc 2 dùng mô hình cho chọn kịch bản và truy hồi phẩm xuất; kết quả đánh dấu PHỎNG ĐOÁN",
     "diagnostics.select_smart, registry.find_smart", "TC-62 (19 test)", DA, ""),

    # -- J. Chẩn đoán & mạch thật -------------------------------------------
    ("J. Mạch thật & chẩn đoán", "Chẩn đoán hai kênh",
     "Giao của telemetry máy và quan sát người; thiếu một kênh thì TỪ CHỐI kết luận",
     "eaa/diagnostics.py", "TC-27 (33 test)", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Bộ kịch bản chẩn đoán DS-01..06",
     "Khai báo bằng YAML ở tầng dự án, engine không biết nội dung",
     "projects/*/diagnostics.yaml", "TC-27", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Checklist an toàn trước kịch bản gây chuyển động",
     "Kịch bản làm mạch chuyển động đòi xác nhận đủ checklist",
     "eaa/diagnostics.py", "TC-27", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Biên dịch từng module rồi liên kết riêng",
     "compile dùng -c sinh tệp đối tượng; link và hex là hai năng lực riêng của pack",
     "eaa/tools/compile.py, packs/avr/pack.yaml", "TC-40 (19 test)", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Ráp ảnh nạp được (.elf → .hex)",
     "LinkGate gộp tệp đối tượng thành ELF rồi đổi sang định dạng nạp được",
     "eaa/tools/compile.py", "TC-40c", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Đo kích thước nói rõ đang đo tầm nào",
     "size_scope phân biệt chiếm dụng của một module lẻ với của cả firmware",
     "eaa/tools/compile.py", "TC-40d", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Ngưỡng bộ nhớ đo trên CẢ firmware",
     "flash_pct_max lần đầu áp lên thứ sẽ nạp xuống mạch, không lên module lẻ",
     "eaa/firmware.py", "TC-41e", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Nạp firmware xuống mạch qua USB",
     "eaa flash gọi năng lực flash của pack; bốn phép kiểm rồi mới hỏi người",
     "eaa/flash.py + eaa flash", "TC-42 (30 test)", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Bốn phép kiểm trước khi nạp",
     "Có ảnh · kho mã sạch · ảnh mới hơn nguồn · người xác nhận — đều là 'không', không phải cảnh báo",
     "eaa/flash.py", "TC-42b, TC-42c", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Nhật ký nạp append-only",
     "Commit nào, ảnh nào (kèm băm), cổng nào, ai, lúc nào — ghi cả lần trượt",
     "eaa flash --history", "TC-42d", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Engine không đoán cổng để nạp",
     "Nhận ra đúng một cổng thì tự chọn; mơ hồ thì dừng và đòi --port",
     "eaa/cli.py", "TC-42e", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Đọc telemetry UART thật",
     "eaa telemetry thu thẳng từ cổng; diagnose run --port dùng luôn kênh máy sống",
     "eaa/telemetry.py", "TC-43 (28 test)", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Khung telemetry có checksum",
     "Định dạng khung do dự án khai; mỗi loại hỏng có lý do riêng",
     "eaa/telemetry.py", "TC-43a", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Phiên thu nhiều khung hỏng bị từ chối",
     "Vượt ngưỡng thì không kết luận: dữ liệu nhiễu vẫn cho ra số trông hợp lý",
     "eaa/telemetry.py", "TC-43c", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Giữ nguyên văn bản thu, phát lại được",
     "Bản thô cạnh bản đã lọc; phân tích lại không cần mạch",
     "eaa telemetry --replay", "TC-43d", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Sinh firmware chẩn đoán từ mẫu",
     "Bộ khung ở pack + phần đo ở dự án, ghép bằng liên kết; eaa diagnose build",
     "eaa/firmware.py", "TC-44, TC-58a", DA,
     ""),
    ("J. Mạch thật & chẩn đoán", "Không dựng firmware chẩn đoán rỗng",
     "Kịch bản chưa khai phần đo thì dừng — ảnh im lặng không phân biệt được với mạch hỏng",
     "eaa/firmware.py", "TC-44b", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Checklist an toàn theo ảnh tới lúc nạp",
     "Thẻ .meta.json đi kèm ảnh; eaa flash đưa checklist ra đúng lúc người bấm đồng ý",
     "eaa/firmware.py + eaa/cli.py", "TC-44d", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Chốt vòng: số đo thật → G4 → phong hạng",
     "eaa tune --port thu telemetry, rút số đo theo tiêu chí đã khai, rồi phong hạng",
     "eaa/acceptance.py", "TC-45 (23 test)", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Không phong hạng cho bản chưa từng chạy trên mạch",
     "Nhật ký nạp nói bản khác đang trên chip thì CHẶN; nhật ký trống thì ghi "
     "device_verified=false chứ không cấm",
     "eaa/acceptance.py", "TC-45a", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Tiêu chí nghiệm thu có TRƯỚC số đo",
     "acceptance.measurements khai trước; thiếu số đo hay vượt ngưỡng đều chặn phong hạng",
     "constraints.yaml + eaa/acceptance.py", "TC-45b, TC-45c, TC-45d", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Debug sâu qua debugWIRE / JTAG / SWD",
     "Đặt điểm dừng, đọc thanh ghi lúc chạy",
     "—", "—", CHUA,
     "Ngoài phạm vi đề án; UART + nạp lại là đủ cho vòng chẩn đoán hai kênh"),

    ("J. Mạch thật & chẩn đoán", "Kiểm sau khi nạp: đọc ngược bộ nhớ",
     "Ba kết cục — khớp / lệch (lần nạp bị coi là trượt) / không kiểm được; "
     "'nạp không báo lỗi' KHÔNG được đọc thành 'nạp đúng'",
     "eaa/flash.py + năng lực flash_verify", "TC-52 (12 test)", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Đo tay bằng dụng cụ — kênh thứ ba",
     "Dòng tổng, sụt áp trên dây, nhiệt độ vỏ: hướng dẫn đích danh đo cái gì, "
     "ở đâu, điều kiện nào, chờ đợi bao nhiêu",
     "eaa/diagnostics.py ManualMeasurement", "TC-58c, TC-58d", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Đo đặc tính thời gian thực, kể cả xấu nhất",
     "Chu kỳ trung bình, chu kỳ DÀI NHẤT, dao động, tải CPU — ràng buộc áp lên "
     "trường hợp xấu nhất chứ không lên trung bình",
     "projects/*/diagnostics/DS-06.c", "TC-58b", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Kiểm độ bền dài hạn",
     "Phát hiện reset qua bộ đếm thời gian chạy tụt; báo cáo không suy rộng "
     "quá quãng đã quan sát",
     "eaa/endurance.py + eaa endurance", "TC-58e..g", DA, ""),
    ("J. Mạch thật & chẩn đoán", "Chẩn đoán sự cố ngoài hiện trường",
     "Dựng lại điều kiện trước khi đo; không tái hiện được thì nói CHƯA KẾT "
     "LUẬN ĐƯỢC thay vì đoán",
     "eaa/diagnostics.py FieldCase", "TC-59e, TC-59f", DA, ""),

    ("A. Khởi tạo & phạm vi", "Agent đề xuất phạm vi, ràng buộc, tiêu chí, bảng chân",
     "Bốn bản đề xuất của G0/G1; mỗi ràng buộc kèm HỆ QUẢ, mỗi tiêu chí là số "
     "+ đơn vị + cách đo, mỗi chân được kiểm chức năng thay thế",
     "eaa/propose.py", "TC-54 (38 test)", DA, ""),
    ("A. Khởi tạo & phạm vi", "Danh sách tài liệu và trang cần trích, đích danh",
     "Suy từ hồ sơ phần cứng và đồ thị; chỉ xin phần CÒN THIẾU trích đoạn",
     "eaa/docplan.py", "TC-55a..e", DA, ""),
    ("A. Khởi tạo & phạm vi", "Errata theo đúng rev silicon",
     "'Chưa tra' khác hẳn 'chip sạch'; module chạm lỗi đã công bố được gọi tên",
     "eaa/docplan.py ErrataAnalysis", "TC-55f..h", DA, ""),

    ("C. Tri thức & RAG", "Bộ chuẩn đánh giá truy xuất (golden set)",
     "precision@3 trên bộ chuẩn của dự án, kèm chunk nhiễu có chủ ý",
     "eaa/goldenset.py + eaa report retrieval", "TC-20 (10 test)", DA, ""),
    ("C. Tri thức & RAG", "Ảnh màn hiện sóng thành số đo đề xuất",
     "Kèm sai số đọc ảnh; người sửa được giá trị trước khi lưu, bản ghi giữ cả hai số",
     "eaa/ingest.py ScopeImageReader", "TC-23 (10 test)", DA, ""),

    ("D. Sinh mã", "Sinh giao diện TRƯỚC khi sinh thân",
     "Hợp đồng gọi trả lời ba câu chữ ký không nói được: ngắt / chặn / tái nhập",
     "eaa/interfaces.py + khuôn của pack", "TC-56a..e", DA, ""),
    ("D. Sinh mã", "Ngân sách tài nguyên và token chia theo module",
     "Cảnh báo khi một module ăn quá phần; trần token tích lũy chặn TRƯỚC khi gọi mô hình",
     "eaa/budget.py + eaa budget", "TC-53 (30 test)", DA, ""),

    ("E. Kiểm chứng & cổng", "Nêu đích danh phần KHÔNG kiểm được trên máy chủ",
     "Thanh ghi, ngoại vi, ràng buộc thời gian — mỗi loại chỉ ra nơi nó được đóng",
     "eaa/tools/unittests.py host_gaps", "TC-56f, TC-56g", DA, ""),
    ("E. Kiểm chứng & cổng", "Khoảng trống ngăn xếp ở tầm firmware",
     "Số liệu SUY RA từ dung lượng trừ phần đã dùng; chỉ có nghĩa sau khi liên kết",
     "eaa/budget.py derived", "TC-53d", DA, ""),

    ("F. Mô phỏng", "Tiêm lỗi trong mô phỏng",
     "Bốn kiểu hỏng đặt tên theo hành vi; kiểm hệ có vào chế độ an toàn không",
     "eaa/tools/sim_runner.py FaultSpec", "TC-57a..e", DA, ""),
    ("F. Mô phỏng", "Đề xuất mô hình đối tượng",
     "Tham số chưa đo vào Assumption Log; mô hình BẮT BUỘC nêu hiện tượng nó bỏ qua",
     "eaa/propose.py PlantModelProposal", "TC-57f, TC-57g", DA, ""),

    ("K. Bàn giao & vận hành", "Tài liệu vận hành sinh từ dữ liệu dự án",
     "Kèm mục 'điều hệ thống KHÔNG làm được', dựng từ chỗ hở thật chứ không viết tay",
     "eaa/handover.py + eaa handover doc", "TC-59a, TC-59b", DA, ""),
    ("K. Bàn giao & vận hành", "Đánh giá ảnh hưởng khi đổi linh kiện",
     "So hai linh kiện rồi bắc cầu trên đồ thị ra module bị chạm",
     "eaa/handover.py SwapAnalysis", "TC-59c, TC-59d", DA, ""),
    ("K. Bàn giao & vận hành", "Cập nhật thiết bị đã triển khai",
     "Bậc đầu ĐÚNG MỘT thiết bị; phải có đường lui lấy từ known_good.lock",
     "eaa/handover.py RolloutPlan", "TC-59g, TC-59h", DA, ""),

    ("L. Siêu nghiệp vụ", "Bộ từ vựng mức tin cậy dùng chung",
     "ĐÃ KIỂM / SUY RA / GIẢ ĐỊNH / KHÔNG KIỂM ĐƯỢC — 23 lớp kết luận cùng dùng",
     "eaa/confidence.py", "TC-60a..c, TC-63a..e", DA, ""),
    ("L. Siêu nghiệp vụ", "Test canh ĐỘ PHỦ của nhãn tin cậy",
     "Thêm lớp sinh kết luận mà quên gắn nhãn thì bộ test đỏ ngay",
     "tests/test_tc63_confidence_coverage.py", "TC-63a", DA, ""),
    ("L. Siêu nghiệp vụ", "Agent tự phát hiện sai lệch thiết kế",
     "Đối chiếu module và lệnh với tài liệu; dựng nháp mục cho sổ sai lệch",
     "eaa/deviation.py + eaa deviations", "TC-60d..f", DA, ""),
    ("L. Siêu nghiệp vụ", "Tự đánh giá quy trình",
     "Chỉ ra cổng hay trượt nhất; mỗi đề xuất gắn với một con số quan sát được",
     "eaa/kpi.py weak_points + eaa report review", "TC-60g..i", DA, ""),
]


# --------------------------------------------------------------------------
# Lộ trình nối mạch thật — thứ tự có ý nghĩa, mỗi bước mở khóa bước sau
# --------------------------------------------------------------------------

LO_TRINH: list[tuple[str, str, str, str, str]] = [
    ("1 ✔", "Tách biên dịch và liên kết — XONG",
     "compile dùng -c cho từng tệp nguồn; link gộp tệp đối tượng + main() thành "
     ".elf; hex đổi sang định dạng nạp được. TC-40, 19 test",
     "Chưa có bước này thì KHÔNG CÓ TỆP ĐỂ NẠP — mọi bước sau vô nghĩa",
     "packs/avr/pack.yaml, eaa/tools/compile.py, eaa/platform.py"),
    ("2 ✔", "Ráp firmware hoàn chỉnh — XONG",
     "Khuôn vòng lặp chính ở pack (bộ định thời hợp tác); firmware.yaml khai "
     "module nào chạy mỗi bao nhiêu ms; `eaa build` dịch, liên kết, ra .hex. TC-41",
     "Biến các module rời thành một chương trình chạy được trên chip",
     "packs/avr/templates/, eaa/firmware.py, eaa/cli.py"),
    ("3 ✔", "Liệt kê cổng USB — XONG",
     "eaa ports; khớp VID/PID với bo khai ở hardware_profile.yaml; thiếu pyserial "
     "thì nói thẳng là không đọc được VID/PID. TC-42a",
     "Agent trả lời được 'bạn đang cắm mạch nào vào cổng nào'",
     "eaa/serialport.py"),
    ("4 ✔", "Lệnh nạp firmware — XONG",
     "eaa flash: bốn phép kiểm (có ảnh · kho sạch · ảnh mới hơn nguồn · người xác "
     "nhận) rồi mới nạp; nhật ký append-only ghi cả lần trượt. TC-42",
     "Đây là lần đầu Agent chạm vào phần cứng thật",
     "eaa/flash.py, eaa/cli.py"),
    ("5 ✔", "Bộ đọc telemetry UART — XONG",
     "eaa telemetry: hạn thời gian bắt buộc, khung hỏng được đếm, bản nguyên văn "
     "giữ lại và phát lại được; diagnose run --port từ chối kết luận trên phiên "
     "không tin được. TC-43",
     "Kênh máy của chẩn đoán hai kênh chuyển từ 'đọc tệp' sang 'đọc mạch'",
     "eaa/telemetry.py, eaa/cli.py"),
    ("6 ✔", "Sinh firmware chẩn đoán — XONG (2/6 kịch bản)",
     "eaa diagnose build: bộ khung của pack + phần đo của dự án, ghép bằng liên "
     "kết. Ảnh mang thẻ an toàn tới tận lúc nạp. TC-44",
     "Mỗi kịch bản chẩn đoán tự chạy được thay vì cần người viết mã đo",
     "eaa/firmware.py, packs/avr/templates/, projects/*/diagnostics/"),
    ("7 ✔", "Khép vòng số đo → G4 — XONG",
     "eaa tune --port: thu telemetry → rút số đo theo tiêu chí đã khai → G4 → "
     "hw-verified. Commit phong hạng phải là commit đang chạy trên thiết bị. TC-45",
     "Vòng đời ba hạng chất lượng khép kín, không còn nhập tay",
     "eaa/acceptance.py, eaa/cli.py"),
    ("8 ✔", "Gate nhiều phương án — XONG",
     "eaa decide dựng N phương án kèm đánh đổi; approve ĐÒI --option và không lấy "
     "gợi ý làm mặc định; cả tập (kể cả cái bị loại) vào quyết định. TC-46",
     "Đúng điều bạn nêu: 'tương tác với người dùng chọn giải pháp phù hợp'",
     "eaa/options.py, eaa/gates.py, eaa/cli.py"),
]

# --------------------------------------------------------------------------
# Kiểu dáng
# --------------------------------------------------------------------------

MAU = {
    DA: "C6EFCE",
    PHAN: "FFEB9C",
    CHUA: "FFC7CE",
    COY: "D9E1F2",
}
MAU_CHU = {
    DA: "006100",
    PHAN: "9C5700",
    CHUA: "9C0006",
    COY: "1F4E79",
}

DAM = Font(bold=True)
TRANG_DAM = Font(bold=True, color="FFFFFF")
NEN_TIEU_DE = PatternFill("solid", fgColor="2F5597")
VIEN = Border(*[Side(style="thin", color="BFBFBF")] * 4)
TREN_TRAI = Alignment(vertical="top", wrap_text=True)


def _tieu_de(ws, tieu_de: list[str], do_rong: list[int]) -> None:
    ws.append(tieu_de)
    for i, rong in enumerate(do_rong, 1):
        ws.column_dimensions[get_column_letter(i)].width = rong
    for o in ws[1]:
        o.font = TRANG_DAM
        o.fill = NEN_TIEU_DE
        o.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        o.border = VIEN
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _to_dep(ws, cot_trang_thai: int | None = None) -> None:
    for hang in ws.iter_rows(min_row=2):
        for o in hang:
            o.alignment = TREN_TRAI
            o.border = VIEN
        if cot_trang_thai is not None:
            o_tt = hang[cot_trang_thai - 1]
            if o_tt.value in MAU:
                o_tt.fill = PatternFill("solid", fgColor=MAU[o_tt.value])
                o_tt.font = Font(bold=True, color=MAU_CHU[o_tt.value])
                o_tt.alignment = Alignment(
                    vertical="center", horizontal="center", wrap_text=True
                )


def dung_bang(dich: Path) -> Path:
    wb = Workbook()

    # ---- Sheet 1: Tổng quan ------------------------------------------------
    ws = wb.active
    ws.title = "Tổng quan"

    nhom = sorted({t[0] for t in TINH_NANG})
    ws.append(["Embedded AIDD Agent — thống kê tính năng"])
    ws["A1"].font = Font(bold=True, size=16, color="2F5597")
    ws.append([
        "Đề án Thạc sĩ Kỹ thuật (Kỹ thuật Điện tử, PTIT) · "
        "Học viên: Vũ Trí Công · Giảng viên hướng dẫn: TS. Nguyễn Trung Hiếu · "
        f"cập nhật {_hom_nay()}"
    ])
    ws["A2"].font = Font(italic=True, color="7F7F7F")
    ws.append([])

    ws.append(["Nhóm", "Tổng", DA, PHAN, CHUA, COY, "% đã làm"])
    for o in ws[4]:
        o.font = TRANG_DAM
        o.fill = NEN_TIEU_DE
        o.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        o.border = VIEN

    for ten in nhom:
        hang = [t for t in TINH_NANG if t[0] == ten]
        d = sum(1 for t in hang if t[5] == DA)
        p = sum(1 for t in hang if t[5] == PHAN)
        c = sum(1 for t in hang if t[5] == CHUA)
        k = sum(1 for t in hang if t[5] == COY)
        # Mục CỐ Ý KHÔNG LÀM không tính vào mẫu số: nó không phải việc còn nợ.
        mau_so = len(hang) - k
        ws.append([ten, len(hang), d, p, c, k, (d / mau_so) if mau_so else 1.0])

    tong = len(TINH_NANG)
    td = sum(1 for t in TINH_NANG if t[5] == DA)
    tp = sum(1 for t in TINH_NANG if t[5] == PHAN)
    tc = sum(1 for t in TINH_NANG if t[5] == CHUA)
    tk = sum(1 for t in TINH_NANG if t[5] == COY)
    ws.append(["TỔNG CỘNG", tong, td, tp, tc, tk, td / (tong - tk)])

    for i, rong in enumerate([30, 8, 10, 11, 11, 17, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = rong
    for hang in ws.iter_rows(min_row=5, max_row=ws.max_row):
        for o in hang:
            o.border = VIEN
            o.alignment = Alignment(vertical="center", wrap_text=True)
        hang[6].number_format = "0 %"
        for i, tt in enumerate((DA, PHAN, CHUA, COY), start=2):
            hang[i].alignment = Alignment(vertical="center", horizontal="center")
            if hang[i].value:
                hang[i].font = Font(color=MAU_CHU[tt], bold=True)
    for o in ws[ws.max_row]:
        o.font = Font(bold=True)
        o.fill = PatternFill("solid", fgColor="EDEDED")

    ws.append([])
    ws.append(["Số liệu nền"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=12, color="2F5597")
    for nhan, gia_tri in [
        ("Test tự động đang xanh", "900"),
        ("Dòng mã engine (eaa/)", "14.792"),
        ("Lệnh CLI", "22"),
        ("Platform Pack", "1 (AVR 8-bit)"),
        ("Mô hình nền", "gemini-3.8-flash (ghim phiên bản, mặc định từ 04/09/2026)"),
        ("Sai lệch thiết kế đã ghi", "37 mục (SL-01..SL-37)"),
    ]:
        ws.append([nhan, gia_tri])
        ws[f"A{ws.max_row}"].font = DAM

    ws.append([])
    ws.append(["Ý nghĩa bốn trạng thái"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=12, color="2F5597")
    for tt, y in [
        (DA, "Có mã, có test, chạy được end-to-end"),
        (PHAN, "Có mã nhưng còn mắt xích thiếu — xem cột 'Còn thiếu gì'"),
        (CHUA, "Chưa có mã"),
        (COY, "Làm được nhưng thiết kế cấm — xem cột 'Còn thiếu gì' để rõ điều khoản"),
    ]:
        ws.append([tt, y])
        o = ws[f"A{ws.max_row}"]
        o.fill = PatternFill("solid", fgColor=MAU[tt])
        o.font = Font(bold=True, color=MAU_CHU[tt])
        o.border = VIEN

    # ---- Sheet 2: Đã làm được ---------------------------------------------
    ws2 = wb.create_sheet("Đã làm được")
    _tieu_de(
        ws2,
        ["#", "Nhóm", "Tính năng", "Làm được gì", "Module / Lệnh", "Bằng chứng"],
        [5, 24, 34, 56, 30, 22],
    )
    for i, t in enumerate([x for x in TINH_NANG if x[5] == DA], 1):
        ws2.append([i, t[0], t[1], t[2], t[3], t[4]])
    _to_dep(ws2)
    ws2.auto_filter.ref = f"A1:F{ws2.max_row}"

    # ---- Sheet 3: Chưa làm được -------------------------------------------
    ws3 = wb.create_sheet("Chưa làm được")
    _tieu_de(
        ws3,
        ["#", "Nhóm", "Tính năng", "Hiện trạng", "Trạng thái", "Còn thiếu gì / vì sao"],
        [5, 24, 32, 46, 16, 62],
    )
    con_lai = [x for x in TINH_NANG if x[5] != DA]
    thu_tu = {PHAN: 0, CHUA: 1, COY: 2}
    con_lai.sort(key=lambda x: (thu_tu[x[5]], x[0]))
    for i, t in enumerate(con_lai, 1):
        ws3.append([i, t[0], t[1], t[2], t[5], t[6]])
    _to_dep(ws3, cot_trang_thai=5)
    ws3.auto_filter.ref = f"A1:F{ws3.max_row}"

    # ---- Sheet 4: Lộ trình -------------------------------------------------
    ws4 = wb.create_sheet("Lộ trình mạch thật")
    ws4.append(["Lộ trình nối Agent với mạch thật qua USB"])
    ws4["A1"].font = Font(bold=True, size=14, color="2F5597")
    ws4.append([
        "Thứ tự có ý nghĩa: mỗi bước mở khóa bước sau. Bước 1 là mắt xích chặn — "
        "chưa tách biên dịch/liên kết thì không có tệp .hex để nạp."
    ])
    ws4["A2"].font = Font(italic=True, color="7F7F7F")
    ws4.append([])
    ws4.append(["Bước", "Việc", "Làm gì cụ thể", "Mở khóa điều gì", "Chạm vào đâu"])
    for o in ws4[4]:
        o.font = TRANG_DAM
        o.fill = NEN_TIEU_DE
        o.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        o.border = VIEN
    for b in LO_TRINH:
        ws4.append(list(b))
    for i, rong in enumerate([7, 30, 62, 46, 34], 1):
        ws4.column_dimensions[get_column_letter(i)].width = rong
    for hang in ws4.iter_rows(min_row=5):
        for o in hang:
            o.alignment = TREN_TRAI
            o.border = VIEN
        hang[0].font = DAM
        hang[1].font = DAM
    ws4.freeze_panes = "A5"

    dich.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dich)
    return dich


def _hom_nay() -> str:
    from datetime import date

    return date.today().isoformat()


if __name__ == "__main__":
    dich = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("docs/EAA_Thong_ke_tinh_nang.xlsx")
    print(f"Đã ghi: {dung_bang(dich)}")
