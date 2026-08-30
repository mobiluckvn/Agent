#!/usr/bin/env python3
"""Dựng bảng năng lực Agent ra Excel — nền chung ĐỐI CHIẾU với nhúng riêng.

    python scripts/lam_bang_nang_luc.py

Sinh ``docs/EAA_Bang_nang_luc.xlsx``.

Vì sao có tệp này bên cạnh ``eaa capabilities``
-----------------------------------------------

``eaa capabilities`` trả lời "máy này đang có gì chạy được" — nó đọc registry
và kiểm sự có mặt, nên nó chỉ thấy được thứ đã được cài vào hệ. Nó không thấy
được thứ CHƯA AI VIẾT. Bảng này trả lời câu khác: **so với một khung năng lực
agent tổng quát, chỗ nào của Agent này còn trống?**

Cột "Bằng chứng" là chỗ giữ cho bảng khỏi thành lời tự khen: mọi dòng ĐỦ phải
chỉ được ra module hoặc lệnh cụ thể. Dòng nào không chỉ được thì không phải ĐỦ.

Khung năng lực nền (C1–C9) lấy theo bản phân tích người dùng cung cấp
30/08/2026, giữ nguyên thứ tự và cách chia nhóm để đối chiếu được. C10 là phần
thêm, dựng từ chính ghi chú cuối của bản ấy: quy trình không nên đóng cứng.

Học viên: Vũ Trí Công — GVHD: TS. Nguyễn Trung Hiếu (PTIT).
"""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

GOC = Path(__file__).resolve().parent.parent
RA = GOC / "docs" / "EAA_Bang_nang_luc.xlsx"

DU = "ĐỦ"
PHAN = "MỘT PHẦN"
CHUA = "CHƯA"
COY = "CỐ Ý KHÔNG"

#: Mức ưu tiên bổ sung — do người rà soát chốt lại, đây chỉ là đề xuất.
CAO, VUA, THAP, KHONG = "Cao", "Vừa", "Thấp", "—"

# (mã, nhóm, năng lực, trạng thái, bằng chứng trong mã, khoảng trống / ghi chú, ưu tiên)
NEN: list[tuple[str, str, str, str, str, str, str]] = [
    # ---------------------------------------------------------------- C1 ---
    ("C1.1", "1. Năng lực nền", "Hiểu và phân rã nhiệm vụ thành bước con, xác định mỗi bước cần loại công cụ gì", PHAN,
     "eaa/agent.py AgentLoop chọn công cụ theo từng lượt; eaa/decompose.py + `eaa plan propose` phân rã bài toán thành module",
     "Phân rã vẫn trong danh mục công cụ, NHƯNG từ SL-77 Agent tự VIẾT được công cụ mới khi việc cần làm không có lệnh nào sẵn (`tool propose`). Còn thiếu: nó chưa tự NHẬN RA lúc nào nên viết", VUA),
    ("C1.2", "1. Năng lực nền", "Lập kế hoạch & theo dõi tiến độ; dừng hỏi khi thiếu thông tin quyết định", DU,
     "eaa/ledger.py sổ module có trạng thái; eaa/state.py ghi nguyên tử (TC-03); `eaa plan list`; 5 gate + eaa/brief.py tự hỏi",
     "", KHONG),
    ("C1.3", "1. Năng lực nền", "Suy luận có kiểm chứng — phân biệt 'tôi tin là' với 'tôi đã xác minh'", DU,
     "eaa/confidence.py 4 mức ĐÃ KIỂM / SUY RA / GIẢ ĐỊNH / KHÔNG KIỂM ĐƯỢC (N-903); TC-63 canh phủ mọi đầu ra",
     "Mạnh hơn khung yêu cầu: 4 mức thay vì 2, và có test chặn đầu ra nào quên gắn nhãn", KHONG),
    ("C1.4", "1. Năng lực nền", "Quản lý ngữ cảnh — tóm tắt log dài, bỏ output thừa, không tràn context", DU,
     "eaa/composer.py nén K1–K7; ngân sách 8.000 token đếm TRƯỚC khi gọi LLM (TC-16); agent.py MAX_OUTPUT_CHARS=3200 cắt đầu ra lệnh",
     "", KHONG),
    ("C1.5", "1. Năng lực nền", "An toàn & xin phép — phân loại hành động, hỏi trước việc không hoàn tác được", DU,
     "eaa/policy.py + eaa/gates.py 5 gate không lệnh nào vượt được; agent.py NGOÀI_DANH_MỤC chặn 12 lệnh ở tầng danh mục; Tool.writes phân loại đọc/ghi",
     "Mạnh hơn khung: ranh giới nằm ở DANH MỤC công cụ, không ở lời dặn trong prompt — Agent không có đường gọi chứ không phải được dặn đừng gọi", KHONG),

    # ---------------------------------------------------------------- C2 ---
    ("C2.1", "2. Kiểm tra công cụ hiện có", "Liệt kê tool đã đăng ký (tên, mô tả, phiên bản, trạng thái)", DU,
     "tools.yaml + packs/<pack>/tools.yaml: name, check, min_version, level, gates, install, smoke, approved_by/at; `eaa doctor`, `eaa capabilities`",
     "", KHONG),
    ("C2.2", "2. Kiểm tra công cụ hiện có", "Dò môi trường: hệ điều hành", DU,
     "eaa/doctor.py _os_key() → macos / linux / windows, dùng để chọn lệnh cài; `eaa ports` liệt kê cổng nối tiếp",
     "", KHONG),
    ("C2.3", "2. Kiểm tra công cụ hiện có", "Dò môi trường: kiến trúc CPU, quyền root, mạng ra ngoài, domain bị chặn", DU,
     "eaa/environ.py probe(): kiến trúc CPU, quyền quản trị, biến proxy (che thông tin đăng nhập), và MẠNG thử nối thật · lệnh `eaa environ` · TC-67",
     "", KHONG),
    ("C2.4", "2. Kiểm tra công cụ hiện có", "Dò runtime: python --version, pip list, npm ls, which, docker ps", PHAN,
     "doctor chạy argv `check` từng công cụ; `eaa environ` liệt kê trình quản lý gói có mặt",
     "Vẫn chưa liệt kê được gói đã cài (pip list / npm ls)", THAP),
    ("C2.5", "2. Kiểm tra công cụ hiện có", "Dò tài nguyên: dung lượng đĩa, RAM, GPU", DU,
     "eaa/environ.py: RAM, đĩa trống, số nhân CPU — kèm cảnh báo khi đĩa trống dưới 2 GB",
     "", KHONG),
    ("C2.6", "2. Kiểm tra công cụ hiện có", "Kiểm tool THỰC SỰ chạy được (smoke test), không chỉ tin registry", DU,
     "smoke + smoke_expect trong mỗi Tool Card; chạy sau khi cài, không đạt thì công cụ không được đánh dấu sẵn sàng",
     "Đúng đúng ý khung: registry nói 'có' không được tính là bằng chứng", KHONG),
    ("C2.7", "2. Kiểm tra công cụ hiện có", "So khớp năng lực ↔ nhu cầu, phát hiện khoảng trống năng lực", DU,
     "eaa/toolsearch.py derive_requirements() SUY nhu cầu từ pack.yaml (phần tử đầu mỗi command) rồi so với manifest; `eaa capabilities` bày 4 tầng",
     "Nhu cầu suy ra từ pack chứ không chép tay — nên không lệch khi pack đổi lệnh", KHONG),
    ("C2.8", "2. Kiểm tra công cụ hiện có", "Kiểm phiên bản / tương thích", DU,
     "min_version + version_regex trong Tool Card; `eaa doctor --lock` khoá phiên bản vào env.lock (N-023)",
     "", KHONG),
    ("C2.9", "2. Kiểm tra công cụ hiện có", "Phát hiện xung đột phụ thuộc", CHUA, "—",
     "Chưa gặp vấn đề thật vì công cụ nhúng là binary độc lập, ít phụ thuộc chéo", THAP),

    # ---------------------------------------------------------------- C3 ---
    ("C3.1", "3. Tìm thông tin để xác định công cụ cần cài", "Chuyển khoảng trống năng lực thành truy vấn tìm kiếm", DU,
     "eaa/websearch.py: JsonEndpointSearch (endpoint tự cấu hình) + GeminiGroundedSearch + ChainSearch; `eaa research` · TC-66",
     "ĐO ĐƯỢC: model chỉ thật sự đi tìm khi được RA LỆNH tìm — câu lệnh ấy nằm trong adapter", KHONG),
    ("C3.2", "3. Tìm thông tin để xác định công cụ cần cài", "Đọc tài liệu chính thống (docs, README, PyPI) bằng web_fetch", DU,
     "eaa/web.py WebFetcher: hai hạng tin cậy, chặn SSRF, trần byte/thời gian, kiểm lại hạng từng chặng chuyển hướng, bộ đệm để tái lập; `eaa read` · TC-65",
     "", KHONG),
    ("C3.3", "3. Tìm thông tin để xác định công cụ cần cài", "Đánh giá & chọn công cụ: còn bảo trì, license, độ phổ biến, phụ thuộc hệ thống", PHAN,
     "toolsearch đọc TRANG CÀI ĐẶT THẬT rồi mới đề xuất; ToolProposal.evidence ghi URL đã tải, và bản in nói rõ khi đề xuất chỉ dựa vào trí nhớ mô hình (SL-80)",
     "Chưa tự đọc chỉ số bảo trì / license / độ phổ biến từ kho gói", VUA),
    ("C3.4", "3. Tìm thông tin để xác định công cụ cần cài", "Kiểm bảo mật nguồn cài", DU,
     "eaa/toolsearch.py PACKAGE_MANAGERS là danh sách trắng; validate_proposal chặn URL ngoài miền cho phép và ĐÒI checksum khi tải trực tiếp (AIS §9.4); eaa/ingest.check_web_source lọc miền",
     "Chưa chống typosquatting theo TÊN GÓI (avr-gcc vs avrgcc). Bù lại: mọi đề xuất đều phải qua người duyệt tại gate", VUA),
    ("C3.5", "3. Tìm thông tin để xác định công cụ cần cài", "Xác định lệnh cài chính xác theo OS, kèm cờ đặc thù", DU,
     "install: {linux, macos, windows} trong mỗi Tool Card, lưu dạng argv chứ không dạng chuỗi shell (không có chỗ chèn lệnh)",
     "", KHONG),
    ("C3.6", "3. Tìm thông tin để xác định công cụ cần cài", "Nhận biết giới hạn mạng, tìm phương án thay thế", DU,
     "eaa/environ.py NetworkCheck thử nối thật + đọc biến proxy; bản báo cáo nói thẳng 'mất mạng → mọi năng lực tra cứu sẽ hỏng'",
     "", KHONG),

    # ---------------------------------------------------------------- C4 ---
    ("C4.1", "4. Cài đặt công cụ", "Chọn cơ chế cài phù hợp", PHAN,
     "Trình quản lý gói (apt / brew / choco / winget) và tải trực tiếp kèm checksum",
     "Không build từ nguồn, không chạy container. Đủ cho toolchain nhúng phổ biến, thiếu cho công cụ lạ", THAP),
    ("C4.2", "4. Cài đặt công cụ", "Cô lập môi trường (venv riêng cho tool mới)", PHAN,
     "Cổng chạy thử của xưởng công cụ chạy trong thư mục tạm riêng, tiến trình riêng, không mạng, không khóa API",
     "Chưa có venv riêng cho công cụ NGOÀI (avr-gcc, cppcheck…)", THAP),
    ("C4.3", "4. Cài đặt công cụ", "Cài theo thứ tự phụ thuộc (system lib trước, runtime package sau)", CHUA, "—",
     "Hiện mỗi Tool Card độc lập, không khai phụ thuộc lẫn nhau", THAP),
    ("C4.4", "4. Cài đặt công cụ", "Xác minh sau cài: chạy --version, chạy test nhỏ với input thật", DU,
     "Sau khi cài: đọc lại phiên bản qua check + chạy smoke; cả hai đạt mới ghi approved_at vào Tool Card",
     "", KHONG),
    ("C4.5", "4. Cài đặt công cụ", "Ghi lại quá trình để tái tạo và rollback", DU,
     "Tool Card ghi approved_by / approved_at; `eaa doctor --lock` → env.lock ghim phiên bản toàn môi trường (N-023)",
     "", KHONG),
    ("C4.6", "4. Cài đặt công cụ", "Agent TỰ chạy lệnh cài", COY,
     "`eaa doctor --fix` in lệnh và hỏi từng cái; `doctor` nằm trong NGOÀI_DANH_MỤC của Agent",
     "Cố ý: cài phần mềm là thay đổi máy của người dùng (N-022 ở mức tự chủ T2). Đây KHÔNG phải khoảng trống cần lấp", KHONG),

    # ---------------------------------------------------------------- C5 ---
    ("C5.1", "5. Xử lý lỗi", "Đọc và phân loại lỗi (mạng / quyền / phụ thuộc / build / runtime)", DU,
     "eaa/installerr.py classify(): 6 loại (mạng/quyền/phụ thuộc/build/không tìm thấy/khác), thứ tự mẫu có chủ ý · TC-69",
     "", KHONG),
    ("C5.2", "5. Xử lý lỗi", "Thử lại có kiểm soát (retry + backoff) cho lỗi mạng", DU,
     "installerr.retry_delays() 2s/4s/8s + WebFetcher.max_retries — và CHỈ lỗi mạng mới retryable, có test canh điều đó",
     "", KHONG),
    ("C5.3", "5. Xử lý lỗi", "Đổi tham số: ghim phiên bản cũ hơn, cờ khác, mirror khác", DU,
     "installerr.remedies(): bậc đổi kho/mirror, bậc ghim phiên bản cũ hơn, bậc cài ngoại tuyến",
     "", KHONG),
    ("C5.4", "5. Xử lý lỗi", "Tìm thông báo lỗi trên web (GitHub issues, StackOverflow)", DU,
     "`eaa research` tra nguyên văn dòng lỗi; thang gỡ của loại KHÁC gọi thẳng lệnh ấy kèm dòng lỗi",
     "", KHONG),
    ("C5.5", "5. Xử lý lỗi", "Đổi sang công cụ thay thế tương đương", DU,
     "installerr.remedies(alternatives=…) — bậc ấy KHÔNG cho Agent tự làm: đổi công cụ là đổi cả cổng kiểm chứng",
     "", KHONG),
    ("C5.6", "5. Xử lý lỗi", "Tự viết giải pháp tối thiểu thay thế", PHAN,
     "eaa/toolforge.py viết được một công cụ thay thế tối thiểu",
     "Chưa nối tự động vào thang gỡ lỗi cài đặt", THAP),
    ("C5.7", "5. Xử lý lỗi", "Báo cáo người dùng: lỗi cụ thể + đã thử gì + gợi ý", DU,
     "ToolReport ghi kết quả từng cổng kèm đầu ra thật; `eaa report review`; vòng tự sửa ghi lại từng lần thử",
     "Đủ cho vòng sinh mã. Với lỗi cài đặt thì mỏng hơn vì C5.1–C5.5 còn trống", KHONG),
    ("C5.8", "5. Xử lý lỗi", "Giới hạn vòng lặp, đặt ngân sách số lần thử", DU,
     "N=3 lần tự sửa, dạng patch chứ không gửi lại cả tệp; quá N thì dừng và bàn giao người, thoát mã 3 (TC-06, TC-19)",
     "Mạnh hơn khung: giới hạn là bất biến có test chặn, không phải một hằng số ai cũng sửa được", KHONG),
    ("C5.9", "5. Xử lý lỗi", "Rollback về trạng thái trước", DU,
     "installerr.rollback_command() suy lệnh gỡ từ chính lệnh cài; không suy được thì trả RỖNG chứ không đoán",
     "", KHONG),
    ("C5.10", "5. Xử lý lỗi", "Phân biệt lỗi của tool với lỗi do input", DU,
     "Tầng cổng: env error vs code error. Tầng cài đặt: 6 loại của installerr",
     "", KHONG),

    # ---------------------------------------------------------------- C6 ---
    ("C6.1", "6. Tự viết code tạo công cụ mới", "Nhận diện khi nào nên tự viết thay vì đi tìm", PHAN,
     "`eaa tool propose` có sẵn và nằm trong danh mục Agent tự gọi",
     "Agent chưa tự NHẬN RA lúc nào nên viết — vẫn cần người gợi ý", VUA),
    ("C6.2", "6. Tự viết code tạo công cụ mới", "Sinh tool đúng chuẩn: chữ ký rõ, type hints, docstring", DU,
     "eaa/toolforge.py ToolForge.design(): sinh mã có MO_TA, SCHEMA, run(), test_ · TC-70",
     "", KHONG),
    ("C6.3", "6. Tự viết code tạo công cụ mới", "Sinh JSON schema tham số để đăng ký vào registry", DU,
     "ForgedTool.schema + check_arguments() kiểm kiểu/bắt buộc/tên lạ TRƯỚC khi nạp mã",
     "", KHONG),
    ("C6.4", "6. Tự viết code tạo công cụ mới", "Input validation, ngoại lệ có nghĩa, timeout, không side-effect ẩn", DU,
     "check_arguments(): kiểu, bắt buộc, tên lạ — và bool không trôi vào chỗ đợi số. Chạy thử có hạn giờ",
     "", KHONG),
    ("C6.5", "6. Tự viết code tạo công cụ mới", "Viết test tối thiểu, chạy trước khi đăng ký", DU,
     "Cổng cấu tạo ĐÒI ít nhất một hàm test_; cổng chạy thử chạy chúng trong tiến trình riêng",
     "", KHONG),
    ("C6.6", "6. Tự viết code tạo công cụ mới", "Chạy trong sandbox, giới hạn quyền file/mạng", DU,
     "run_tests(): tiến trình riêng, thư mục tạm riêng, hạn giờ 30s, EAA_NO_NET=1, xóa khóa API khỏi môi trường con",
     "", KHONG),
    ("C6.7", "6. Tự viết code tạo công cụ mới", "Kiểm bảo mật mã tự sinh: không hard-code secret, không eval input", DU,
     "check_safety() quét theo CÂY CÚ PHÁP: 28 cấu trúc cấm + chuỗi trông giống khóa. Quét chuỗi con sẽ chặn cả re.compile",
     "", KHONG),
    ("C6.8", "6. Tự viết code tạo công cụ mới", "Tài liệu hoá tool mới: README ngắn, ví dụ, giới hạn đã biết", PHAN,
     "MO_TA + SCHEMA + purpose hiện trong `eaa tool list` và trong prompt của Agent",
     "Chưa sinh README riêng cho từng công cụ", THAP),

    # ---------------------------------------------------------------- C7 ---
    ("C7.1", "7. Đăng ký & mở rộng năng lực", "Nạp động tool vào registry (hot-reload, không khởi động lại)", DU,
     "Công cụ đã duyệt được bơm vào prompt lúc chạy (_mo_ta_cong_cu_tu_sinh) — KHÔNG cần khởi động lại. Quyền vẫn tĩnh trong Git: một mục `tool run`",
     "", KHONG),
    ("C7.2", "7. Đăng ký & mở rộng năng lực", "Cập nhật bản đồ năng lực để lần sau không tìm lại", DU,
     "`eaa capabilities` đọc thẳng từ parser, TOOLBOX, pack.yaml và Tool Manifest — không có danh sách chép tay nào để lệch",
     "", KHONG),
    ("C7.3", "7. Đăng ký & mở rộng năng lực", "Versioning tool, giữ bản cũ nếu bản mới hỏng", PHAN,
     "env.lock ghim phiên bản; `eaa report versions` đối chiếu",
     "Không giữ song song hai bản để quay lui", THAP),
    ("C7.4", "7. Đăng ký & mở rộng năng lực", "Đánh giá chất lượng sau khi dùng thật (hay hỏng? chậm?)", PHAN,
     "kpi_log.csv đo vòng sinh mã; sổ tay lỗi đo tỉ lệ trúng từng CÁCH SỬA",
     "Chưa đo độ tin cậy từng công cụ tự sinh sau khi dùng thật", VUA),
    ("C7.5", "7. Đăng ký & mở rộng năng lực", "Gộp chuỗi tool hay dùng thành tool cấp cao hơn (skill/workflow)", CHUA, "—",
     "Hiện chuỗi ấy được đóng cứng trong quy trình G0→G10 thay vì được rút ra thành skill — xem C10.1", CAO),

    # ---------------------------------------------------------------- C8 ---
    ("C8.1", "8. Ghi nhớ & học liên phiên", "Bộ nhớ dài hạn TRONG một dự án", DU,
     "state.json ghi nguyên tử sống sót qua crash (TC-03, N-907); KB append-only + supersede không ghi đè vật lý; tools.yaml; env.lock",
     "", KHONG),
    ("C8.2", "8. Ghi nhớ & học liên phiên", "Bộ nhớ LIÊN dự án", DU,
     "eaa/memory.py MemoryStore ở gốc kho: append-only + supersede, phạm vi toàn cục / mcu:<họ> / dự án:<tên> · TC-68",
     "", KHONG),
    ("C8.3", "8. Ghi nhớ & học liên phiên", "Sổ tay lỗi (playbook): lỗi X → cách sửa Y, tra trước khi tìm web", DU,
     "eaa/playbook.py: vân tay lỗi chuẩn hoá, hai bộ đếm trúng/trượt, xếp theo tỉ lệ Laplace. Agent được dặn tra đây TRƯỚC khi ra web · TC-68",
     "", KHONG),
    ("C8.4", "8. Ghi nhớ & học liên phiên", "Cache kết quả tìm kiếm / đánh giá tool", DU,
     "WebCache ở memory/web_cache — lưu kèm băm và mốc thời gian: để tái lập trước, để nhanh sau",
     "", KHONG),
    ("C8.5", "8. Ghi nhớ & học liên phiên", "Tự đánh giá cuối nhiệm vụ, đề xuất tạo tool/skill mới", PHAN,
     "`eaa report kpi` + sổ tay lỗi tích luỹ",
     "Chưa tự đề xuất 'nên tạo công cụ mới cho việc này'", VUA),

    # ---------------------------------------------------------------- C9 ---
    ("C9.1", "9. Giao tiếp với người dùng", "Báo cáo ngắn theo mốc, không dội log thô", DU,
     "`eaa chat` thuật lại từng bước; đầu ra lệnh cắt ở 3.200 ký tự trước khi vào ngữ cảnh",
     "", KHONG),
    ("C9.2", "9. Giao tiếp với người dùng", "Hỏi đúng lúc: thiếu quyền, cần quyết định giữa các phương án", DU,
     "5 gate; `eaa brief` tự dò rồi tự hỏi; bậc 2 của eaa/gapsearch.py hỏi ĐÍCH DANH thứ còn thiếu thay vì 'thiếu thông tin, bạn bổ sung đi'",
     "", KHONG),
    ("C9.3", "9. Giao tiếp với người dùng", "Giải thích rủi ro trước khi cài thứ ảnh hưởng hệ thống", DU,
     "`eaa doctor --fix` in nguyên lệnh sẽ chạy, nêu phạm vi, hỏi từng cái một",
     "", KHONG),
    ("C9.4", "9. Giao tiếp với người dùng", "Tổng kết cuối: có gì mới, dùng ra sao, chú ý gì", DU,
     "`eaa report`; `eaa handover doc` sinh sổ tay vận hành với giới hạn lấy từ SỐ ĐO THẬT chứ không từ mô tả",
     "", KHONG),

    # --------------------------------------------------------------- C10 ---
    ("C10.1", "10. Không đóng cứng quy trình", "Dùng được cho việc NHỎ mà không bắt đi hết vòng đời", PHAN,
     "`eaa scratch` sinh sẵn constraints + hardware_profile; 4 lệnh mới (environ/research/read + memory/playbook/tool) chạy KHÔNG cần dự án nào",
     "Vẫn phải `eaa init` trước khi vào `eaa chat`", VUA),
    ("C10.2", "10. Không đóng cứng quy trình", "Vào giữa quy trình, bỏ qua giai đoạn không cần", PHAN,
     "`eaa resume` khôi phục sau gián đoạn (N-907)",
     "Khôi phục ≠ bắt đầu giữa chừng. Không có đường vào thẳng G5 mà bỏ G0–G4", CAO),
    ("C10.3", "10. Không đóng cứng quy trình", "Chế độ không cần toolchain / không cần phần cứng", PHAN,
     "Mô phỏng + MockLLM chạy không cần bo; `eaa scratch` nói rõ ràng buộc là GIẢ ĐỊNH",
     "Cổng compile/static vẫn đòi toolchain — đúng thiết kế cho bản bàn giao", THAP),
    ("C10.4", "10. Không đóng cứng quy trình", "Người dùng tự chọn cổng nào bật, cổng nào tắt", CHUA,
     "Bộ cổng do pack.yaml khai, cố định cho cả pack",
     "Cẩn trọng: cổng tắt được thì bất biến 'merge chỉ khi TOÀN BỘ passed' mất nghĩa. Nếu làm thì phải là 'chế độ nháp, không bao giờ merge được'", VUA),
]


def _nap_nghiep_vu():
    """Đọc bảng nghiệp vụ nhúng từ chính script đã có, không chép lại."""
    spec = importlib.util.spec_from_file_location("nv", GOC / "scripts" / "lam_bang_nghiep_vu.py")
    m = importlib.util.module_from_spec(spec)
    sys.modules["nv"] = m
    spec.loader.exec_module(m)
    return m


def main() -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    nv = _nap_nghiep_vu()

    DAM = Font(bold=True, color="FFFFFF")
    NEN_TIEU_DE = PatternFill("solid", fgColor="1F3864")
    NEN_NHOM = PatternFill("solid", fgColor="D9E2F3")
    MAU = {
        DU: PatternFill("solid", fgColor="C6EFCE"),
        PHAN: PatternFill("solid", fgColor="FFEB9C"),
        CHUA: PatternFill("solid", fgColor="FFC7CE"),
        COY: PatternFill("solid", fgColor="E7E6E6"),
    }
    MAU_UT = {CAO: PatternFill("solid", fgColor="FF9999"), VUA: PatternFill("solid", fgColor="FFE699")}
    VIEN = Border(*[Side("thin", color="BFBFBF")] * 4)
    TREN = Alignment(vertical="top", wrap_text=True)

    wb = Workbook()

    def dat_tieu_de(ws, cot: list[tuple[str, int]]) -> None:
        for i, (ten, rong) in enumerate(cot, start=1):
            o = ws.cell(row=1, column=i, value=ten)
            o.font, o.fill = DAM, NEN_TIEU_DE
            o.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(i)].width = rong
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

    # ═══════════════════════════════════════════════════ 1. Đọc trước ═══
    ws = wb.active
    ws.title = "Đọc trước"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 104
    dong = [
        ("BẢNG NĂNG LỰC AGENT — EAA", ""),
        ("", ""),
        ("Đề án", "Agent lập trình nhúng tổng quát (Embedded AIDD Agent)"),
        ("Học viên", "Vũ Trí Công"),
        ("GVHD", "TS. Nguyễn Trung Hiếu"),
        ("Dựng ngày", "30/08/2026 — sinh lại bằng: python scripts/lam_bang_nang_luc.py"),
        ("", ""),
        ("BẢNG NÀY SO CÁI GÌ", ""),
        ("Sheet 'Năng lực nền'",
         "Khung năng lực agent tổng quát do người dùng cung cấp (9 nhóm), đối chiếu với mã hiện có. "
         "Nhóm C10 là phần thêm, dựng từ ghi chú cuối của bản ấy: quy trình không nên đóng cứng."),
        ("Sheet 'Năng lực nhúng'",
         "74 nghiệp vụ riêng của lập trình nhúng, xếp theo 11 giai đoạn vòng đời + nhóm xuyên suốt. "
         "Đây là phần KHÔNG có trong khung chung — giá trị riêng của đề án nằm ở đây."),
        ("Sheet 'Khoảng trống'", "Lọc mọi dòng CHƯA / MỘT PHẦN của cả hai sheet, xếp theo mức ưu tiên đề xuất."),
        ("Sheet 'Bản đồ lệnh'", "38 lệnh CLI: lệnh nào Agent tự gọi được, lệnh nào đòi hồ sơ dự án."),
        ("", ""),
        ("CÁCH ĐỌC CỘT TRẠNG THÁI", ""),
        (DU, "Có mã chạy được VÀ chỉ ra được module/lệnh cụ thể ở cột Bằng chứng. Dòng không chỉ được thì không phải ĐỦ."),
        (PHAN, "Có phần lõi, thiếu một nhánh. Cột Khoảng trống nói rõ thiếu gì."),
        (CHUA, "Không có mã nào làm việc này."),
        (COY, "Quyết định thiết kế, không phải thiếu sót. Không nên lấp."),
        ("", ""),
        ("THANG MỨC TỰ CHỦ (sheet Năng lực nhúng)", ""),
    ]
    for ma, nghia, _ in nv.THANG_TU_CHU:
        dong.append((ma, nghia))
    dong += [
        ("", ""),
        ("ĐIỀU BẢNG NÀY KHÔNG LÀM", ""),
        ("", "Nó không chạy thử năng lực nào — nó đối chiếu khung với mã. Câu 'nó chạy đúng không' "
             "thuộc về bộ test (pytest -q, 1428 test) và scripts/kiem_on_dinh.py."),
        ("", "Nó cũng không nói năng lực nào ĐÁNG có. Cột 'Ưu tiên' chỉ là đề xuất để rà soát, "
             "người chốt lại là người đọc."),
    ]
    for r, (a, b) in enumerate(dong, start=1):
        ws.cell(row=r, column=1, value=a).alignment = TREN
        ws.cell(row=r, column=2, value=b).alignment = TREN
        if a and not b:
            ws.cell(row=r, column=1).font = Font(bold=True, size=12 if r > 2 else 14)
        if a in MAU:
            ws.cell(row=r, column=1).fill = MAU[a]

    # ══════════════════════════════════════════════ 2. Năng lực nền ═══
    ws = wb.create_sheet("Năng lực nền")
    dat_tieu_de(ws, [("Mã", 8), ("Nhóm", 26), ("Năng lực", 46), ("Trạng thái", 12),
                     ("Bằng chứng trong mã", 58), ("Khoảng trống / ghi chú", 58), ("Ưu tiên", 9)])
    r = 2
    nhom_truoc = ""
    for ma, nhom, ten, tt, bc, gc, ut in NEN:
        if nhom != nhom_truoc:
            nhom_truoc = nhom
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = NEN_NHOM
            o = ws.cell(row=r, column=1, value=nhom)
            o.font = Font(bold=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            r += 1
        for c, v in enumerate([ma, nhom, ten, tt, bc, gc, ut], start=1):
            o = ws.cell(row=r, column=c, value=v)
            o.alignment = TREN
            o.border = VIEN
        ws.cell(row=r, column=4).fill = MAU.get(tt, PatternFill())
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center", vertical="top")
        if ut in MAU_UT:
            ws.cell(row=r, column=7).fill = MAU_UT[ut]
        ws.cell(row=r, column=7).alignment = Alignment(horizontal="center", vertical="top")
        r += 1
    ws.auto_filter.ref = f"A1:G{r-1}"

    # ═══════════════════════════════════════════ 3. Năng lực nhúng ═══
    ws = wb.create_sheet("Năng lực nhúng")
    dat_tieu_de(ws, [("Mã", 9), ("Giai đoạn", 26), ("Nghiệp vụ", 48), ("Trạng thái", 12),
                     ("Tự chủ", 8), ("Bằng chứng", 56), ("Vì sao dễ sai nếu thiếu", 56)])
    # Nhãn trạng thái của bảng nghiệp vụ khác nhãn của bảng này; quy về một bộ
    # để lọc và tô màu dùng chung được cho cả hai sheet.
    NHAN_TT = {nv.DU: DU, nv.PHAN: PHAN, nv.CHUA: CHUA, nv.COY: COY}
    dc = {d[0]: d for d in nv.DOI_CHIEU}
    r = 2
    gd_truoc = ""
    for muc in nv.NGHIEP_VU:
        ma, gd, ten, tu_chu, vi_sao = muc[0], muc[1], muc[2], muc[6], muc[8]
        d = dc.get(ma)
        tt = NHAN_TT.get(d[1], d[1]) if d else ""
        # Cột 2 của DOI_CHIEU là bằng chứng (module · mã TC); cột 3 là chỗ còn
        # thiếu — chỉ có nội dung ở dòng chưa đủ.
        bang_chung = (d[2] if d else "") or ""
        if d and d[3]:
            bang_chung = f"{bang_chung}\n⚠ {d[3]}" if bang_chung not in ("", "—") else d[3]
        if gd != gd_truoc:
            gd_truoc = gd
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = NEN_NHOM
            ws.cell(row=r, column=1, value=gd).font = Font(bold=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            r += 1
        for c, v in enumerate([ma, gd, ten, tt, tu_chu, bang_chung, vi_sao], start=1):
            o = ws.cell(row=r, column=c, value=v)
            o.alignment = TREN
            o.border = VIEN
        ws.cell(row=r, column=4).fill = MAU.get(tt, PatternFill())
        for c in (4, 5):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="top")
        r += 1
    ws.auto_filter.ref = f"A1:G{r-1}"

    # ══════════════════════════════════════════════ 4. Khoảng trống ═══
    ws = wb.create_sheet("Khoảng trống")
    dat_tieu_de(ws, [("Ưu tiên", 9), ("Mã", 9), ("Thuộc", 14), ("Năng lực", 48),
                     ("Trạng thái", 12), ("Còn thiếu gì", 74)])
    thieu = [(ut, ma, "nền", ten, tt, gc or "—") for ma, _n, ten, tt, _bc, gc, ut in NEN
             if tt in (CHUA, PHAN)]
    for muc in nv.NGHIEP_VU:
        d = dc.get(muc[0])
        if not d:
            continue
        tt = NHAN_TT.get(d[1], d[1])
        if tt in (CHUA, PHAN):
            thieu.append((VUA, muc[0], "nhúng", muc[2], tt, d[3] or "—"))
    thu_tu = {CAO: 0, VUA: 1, THAP: 2, KHONG: 3}
    thieu.sort(key=lambda x: (thu_tu.get(x[0], 9), x[1]))
    for r, hang in enumerate(thieu, start=2):
        for c, v in enumerate(hang, start=1):
            o = ws.cell(row=r, column=c, value=v)
            o.alignment = TREN
            o.border = VIEN
        if hang[0] in MAU_UT:
            ws.cell(row=r, column=1).fill = MAU_UT[hang[0]]
        ws.cell(row=r, column=5).fill = MAU.get(hang[4], PatternFill())
        for c in (1, 5):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="top")
    ws.auto_filter.ref = f"A1:F{len(thieu)+1}"

    # ═════════════════════════════════════════════ 5. Bản đồ lệnh ═══
    ws = wb.create_sheet("Bản đồ lệnh")
    dat_tieu_de(ws, [("Lệnh", 30), ("Agent tự gọi?", 14), ("Ghi ra tệp?", 12),
                     ("Cần hồ sơ dự án?", 16), ("Vì sao Agent không tự gọi", 80)])
    import ast

    cay = ast.parse((GOC / "eaa" / "cli.py").read_text(encoding="utf-8"))
    can_du_an: set[str] = set()
    for n in cay.body:
        if isinstance(n, ast.FunctionDef) and n.name.startswith("cmd_"):
            goi = {x.func.id for x in ast.walk(n) if isinstance(x, ast.Call) and isinstance(x.func, ast.Name)}
            if "resolve_project" in goi:
                can_du_an.add(n.name[4:].replace("_", "-"))

    sys.path.insert(0, str(GOC))
    from eaa.agent import NGOAI_DANH_MUC, TOOLBOX
    from eaa.cli import build_parser

    lenh: set[str] = set()
    p = build_parser()
    if getattr(p, "_subparsers", None):
        for hd in p._subparsers._group_actions:
            lenh |= set(getattr(hd, "choices", {}) or {})

    # Một lệnh CLI có thể sinh nhiều công cụ cho Agent (`plan list`, `plan add`),
    # nên hai con số 38 và 39 không so trực tiếp với nhau được. Sheet này liệt
    # kê THEO CÔNG CỤ, rồi thêm những lệnh CLI không sinh công cụ nào.
    hang: list[tuple[str, str, str, str, str]] = []
    goc_co_cong_cu = {t.argv[0] for t in TOOLBOX}
    for t in sorted(TOOLBOX, key=lambda x: x.argv):
        hang.append((" ".join(t.argv), "có", "có" if t.writes else "—",
                     "có" if t.argv[0] in can_du_an else "không", ""))
    for ten in sorted(lenh - goc_co_cong_cu):
        hang.append((ten, "KHÔNG", "", "có" if ten in can_du_an else "không",
                     NGOAI_DANH_MUC.get(ten, "")))
    for r, h in enumerate(hang, start=2):
        for c, v in enumerate(h, start=1):
            o = ws.cell(row=r, column=c, value=v)
            o.alignment = TREN
            o.border = VIEN
        ws.cell(row=r, column=2).fill = MAU[DU] if h[1] == "có" else MAU[COY]
        for c in (2, 3, 4):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="top")
    ws.auto_filter.ref = f"A1:E{len(hang)+1}"
    tu_goi = TOOLBOX

    RA.parent.mkdir(parents=True, exist_ok=True)
    wb.save(RA)

    tong = len(NEN)
    dem = {k: sum(1 for x in NEN if x[3] == k) for k in (DU, PHAN, CHUA, COY)}
    print(f"Đã ghi {RA.relative_to(GOC)}")
    print(f"  Năng lực nền : {tong} mục — ĐỦ {dem[DU]} · MỘT PHẦN {dem[PHAN]} · "
          f"CHƯA {dem[CHUA]} · CỐ Ý {dem[COY]}")
    print(f"  Năng lực nhúng: {len(nv.NGHIEP_VU)} nghiệp vụ")
    print(f"  Khoảng trống  : {len(thieu)} dòng")
    print(f"  Bản đồ lệnh   : {len(lenh)} lệnh CLI → {len(TOOLBOX)} công cụ Agent tự gọi "
          f"(một lệnh có thể sinh nhiều công cụ), {len(lenh - goc_co_cong_cu)} lệnh chỉ người gọi được, "
          f"{len(can_du_an)}/{len(lenh)} lệnh đòi hồ sơ dự án")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
