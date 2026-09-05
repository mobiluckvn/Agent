#!/usr/bin/env python3
"""Sinh `docs/EAA_Backlog_Tien_hoa.xlsx` — bảng quản trị quá trình tiến hoá.

Thay cho `scripts/lam_bang_viec_phai_lam.py` (bảng 14 việc, 05/09/2026). Mọi mã
V1..V14 của bảng cũ được mang sang nguyên vẹn trong cột **Mã cũ**, nên không
mất dấu vết. Bảng cũ bị rút đi chứ không để song song: hai danh sách việc phải
làm là hai danh sách sẽ lệch nhau, và lúc lệch thì không ai biết cái nào đúng.

Khác bảng cũ ở ba chỗ
----------------------

1. **Chia MẢNG.** 32 việc xếp vào sáu mảng, mỗi mảng có một câu hỏi riêng nó
   trả lời. Một danh sách phẳng dài ra thì không quản trị được.
2. **MÔ PHỎNG và IDE thành mảng đầy đủ**, mỗi mảng một sheet đi sâu. Bảng cũ
   cho IDE đúng một dòng và không có dòng nào cho mô phỏng — trong khi kho đã
   có cổng SIL chạy firmware thật từ Sprint 3.
3. **Quản trị được**: cột `Chặn việc nào`, sheet `Phụ thuộc`, sheet
   `Nhật ký tiến hoá`. Một backlog không nói được "việc này đang chặn việc kia"
   thì nó là một danh sách ước muốn.

Chạy: python3 scripts/lam_bang_tien_hoa.py
"""

from __future__ import annotations

import pathlib
import re
import sys

RA = pathlib.Path("docs/EAA_Backlog_Tien_hoa.xlsx")

# ── hạng sở cứ ───────────────────────────────────────────────────────────────

DO = "ĐO"
SO = "SỔ SAI LỆCH"
VL = "VĂN LIỆU"
KHAI = "KHAI (đối thủ)"
SR = "SUY RA"

# ── mảng ─────────────────────────────────────────────────────────────────────

MANG: list[tuple[str, str, str]] = [
    ("A · Thước đo và bằng chứng",
     "Câu nào của đề án đang là LỜI KHAI mà chưa có số?",
     "Mảng đứng đầu vì không có số thì mọi mảng sau đều không chứng minh được gì"),
    ("B · Tri thức vào prompt",
     "Mô hình biết gì lúc nó viết dòng mã ấy?",
     "Chỗ chất lượng sinh mã được quyết định, và chỗ dễ khai quá lời nhất"),
    ("C · Kiểm chứng trên phần cứng",
     "Cái gì chỉ bo thật mới nói được?",
     "Vùng mù vật lý — nơi mọi cổng phần mềm đều mù"),
    ("D · Mô phỏng",
     "Cái gì kiểm được TRƯỚC khi chạm vào bo?",
     "Đã có MiL và SIL chạy firmware thật; thiếu tiêm lỗi, chu kỳ máy, "
     "và phép đo khoảng cách sim↔bo"),
    ("E · IDE và trải nghiệm",
     "Kỹ sư ngồi trong đâu cả ngày, và họ thấy gì ở đó?",
     "Bị chặn ở một chỗ rất cụ thể: CLI chưa có đầu ra máy đọc được"),
    ("F · Truy vết và bàn giao",
     "Người sau đọc lại được không?",
     "Mảng gần triết lý các bộ công cụ chứng nhận nhất"),
]

# ── sổ sở cứ ─────────────────────────────────────────────────────────────────

SO_CU: list[tuple[str, str, str, str]] = [
    ("SC-01", DO,
     "3 lần từ chối G3 là mã TỰ CHỈNH CHO VỪA ĐỒ ĐO của chính nó, và cả ba qua "
     "sạch bốn cổng máy",
     "docs/TIEP_TUC_TU_DAY.md · docs/V3_KET_QUA.md"),
    ("SC-02", DO,
     "Thông báo lỗi nói được việc phải làm: 150/182 = 82%; mốc trước 25/182 = 14%",
     "pytest tests/test_tc145_thong_bao_loi_noi_duoc_viec_phai_lam.py"),
    ("SC-03", DO,
     "Bốn trục đo mới của đề án đều đang báo CHƯA ĐO ĐƯỢC — thước đã có, chưa đo gì",
     "eaa report bench · eaa/bench.py · SL-177"),
    ("SC-04", DO,
     "V3 đã chạy: bắt được 3/13 · trúng một phần 1 · kêu trật 7 · BỎ SÓT 0. "
     "`contract` kêu sớm hơn người 54 phút ở ca `imu_start_read`",
     "docs/V3_KET_QUA.md · python3 scripts/do_nguoc_lich_su.py"),
    ("SC-05", VL,
     "arXiv 2603.19583 — tri thức do NGƯỜI nén theo từng ngoại vi nâng kết quả "
     "lên gần trần; mốc DUY NHẤT trong khảo sát có ablation đo được",
     "docs/KHAO_SAT_AGENT_NHUNG.md"),
    ("SC-06", SO,
     "SL-125 — hồ sơ phần cứng gõ tay lệch với mạch thật",
     "docs/SAI_LECH_THIET_KE.md mục SL-125"),
    ("SC-07", DO,
     "`eaa observe`: 9/9 module chưa khai dấu hiệu quan sát được, 0 kênh quan "
     "sát trong hồ sơ phần cứng",
     "eaa observe · SL-175 · TC-135"),
    ("SC-08", KHAI,
     "Embedder cược vào vòng kín trên silicon thật — quan sát phần cứng là kênh "
     "phản hồi CHÍNH",
     "docs/KHAO_SAT_AGENT_NHUNG.md"),
    ("SC-09", DO,
     "Năm lỗi chỉ phần cứng mới chỉ ra được; bootloader đo trên bo là 57600 chứ "
     "không phải 115200 như tài liệu",
     "git log (abae855, 89c5177)"),
    ("SC-10", DO,
     "Cảnh báo TRÔI hồ sơ phần cứng vẫn bật; G1 chưa được duyệt lại",
     "eaa status · docs/TIEP_TUC_TU_DAY.md"),
    ("SC-11", SO,
     "SL-174 — `eaa/dimension.py` cố ý chưa nối vào bản đồ thanh ghi",
     "docs/SAI_LECH_THIET_KE.md mục SL-174"),
    ("SC-12", KHAI,
     "Parasoft · LDRA · QA Systems bán bằng chứng chứng nhận: truy vết hai "
     "chiều yêu cầu ↔ ca kiểm ↔ mã",
     "docs/EAA_Benchmark_San_pham.md Bảng 1"),
    ("SC-13", DO,
     "ISR bước 50 kHz trên AVR 16 MHz — chưa ai đo nó ăn bao nhiêu phần trăm CPU",
     "docs/TIEP_TUC_TU_DAY.md"),
    ("SC-14", DO,
     "Bảng năng lực sau rà soát vòng hai: 137 ĐỦ · 3 MỘT PHẦN · 4 CHƯA / 146 dòng",
     "docs/RA_SOAT_NANG_LUC_04_09.md §7.3"),
    ("SC-15", SR,
     "Cổng nối tiếp là tài nguyên độc chiếm; chưa ai chạy hai phiên nên CHƯA đo được",
     "lập luận, chưa có số — nhãn SUY RA là cố ý"),
    ("SC-16", DO,
     "CLI có 103 lệnh, 0 lệnh thiếu trợ giúp — và 0 lệnh nào có đầu ra máy đọc "
     "được. Một extension biên tập không đọc được văn xuôi tiếng Việt",
     "eaa --help · grep '--json' eaa/cli.py cho 0 kết quả"),
    ("SC-17", DO,
     "Lớp truy cập mạng phân hai hạng nguồn theo URL CUỐI (SL-71..80); danh sách "
     "miền chính chủ hiện nghiêng về pack thứ nhất",
     "eaa/web.py · pytest tests/test_tc65*.py"),
    ("SC-18", DO,
     "`ToolError` đã mang sẵn `file`, `line`, `rule_id`, `severity` — đủ cho một "
     "chẩn đoán trong biên tập. Chỉ thiếu chỗ xuất ra",
     "eaa/tools/base.py:40-59"),
    ("SC-19", DO,
     "Từ chối G3 #12: *'bus chết thì robot giữ nguyên lệnh động cơ cuối cùng và "
     "không bao giờ phát hiện mình đã ngã'* — cả bốn cổng đều xanh, NGƯỜI bắt",
     "projects/robot_balance/gates/decisions.jsonl, quyết định thứ 12"),
    ("SC-20", DO,
     "Từ chối G3 #11: vòng bơm cảm biến trượt thì *'im lặng dùng góc cũ (bài "
     "kiểm không thấy vì nó dùng driver giả)'*",
     "gates/decisions.jsonl, quyết định thứ 11"),
    ("SC-21", DO,
     "Cổng SIL đã chạy FIRMWARE C THẬT qua lớp HAL giả lập, không chạy bản "
     "Python song song; `sim/controller.py` khai rõ là bản tham chiếu MiL",
     "packs/avr/pack.yaml §sim · projects/robot_balance/sim/controller.py"),
    ("SC-22", DO,
     "Mô hình vật lý được kiểm chứng bằng nghiệm giải tích trước khi làm cổng chặn",
     "tests/test_sim_verification.py · eaa/tools/sim.py docstring"),
    ("SC-23", DO,
     "Cổng mô phỏng chỉ trả ĐẠT/TRƯỢT theo kịch bản. Một bản sinh lại làm thời "
     "gian xác lập xấu đi vẫn ĐẠT — không mốc nào so được",
     "eaa/tools/sim.py `SimGate.run`"),
    ("SC-24", DO,
     "12 lượt sinh lại module trong lịch sử; không lượt nào có số liệu mô phỏng "
     "được lưu để so với lượt trước",
     "docs/V3_KET_QUA.md · git -C projects/robot_balance/firmware log"),
    ("SC-25", DO,
     "Máy trạng thái ứng dụng có 5 trạng thái; từ chối #12 cho thấy nhánh NGÃ "
     "không được kịch bản nào đi tới",
     "gates/decisions.jsonl quyết định 11-12 · sim/scenarios.yaml"),
    ("SC-26", DO,
     "`board_facts.jsonl` và `measurements.jsonl` đã có số đo trên bo; chưa ai "
     "so chúng với số mô phỏng dự đoán",
     "projects/robot_balance/measurements.jsonl · eaa/measured.py"),
    ("SC-27", SR,
     "Dung sai linh kiện và nhiễu cảm biến chưa được quét; chưa có số về độ "
     "nhạy của bộ tham số với chúng",
     "lập luận, chưa có số — nhãn SUY RA là cố ý"),
    ("SC-28", DO,
     "`SimGate` chỉ chạy trong chuỗi cổng; không lệnh nào cho người chạy thử "
     "một kịch bản rồi xem ngay",
     "eaa/tools/sim.py · eaa --help (không có `sim`)"),
    ("SC-29", DO,
     "13 lý do từ chối G3 dài trung bình hơn 800 ký tự, nhắc đích danh hàm và "
     "hằng số; người đọc phải tự dò về chỗ trong mã",
     "gates/decisions.jsonl"),
    ("SC-30", DO,
     "FR-RAG-02 bắt mã cấu hình thanh ghi phải mang `// ref:`; muốn đọc trích "
     "đoạn ấy hiện phải rời biên tập sang một lệnh khác",
     "eaa/kb.py `Chunk.citation` · TC-17"),
]

# ── việc ─────────────────────────────────────────────────────────────────────
# (mã, mã cũ, mảng, việc, hạng, sở cứ, làm thế nào, tệp, nghiệm thu, bài canh,
#  phụ thuộc, chặn việc nào, rủi ro, công, ưu tiên, trạng thái)

BS = "Biến lời khai thành SỐ"
CL = "Chặn một hạng lỗi"
MR = "Mở rộng mặt tiếp xúc"
NG = "Việc của NGƯỜI"
NM = "Nền móng cho việc khác"

A, B, C, D, E, F = (m[0] for m in MANG)

VIEC: list[tuple[str, ...]] = [
    # ── A · thước đo ────────────────────────────────────────────────────────
    ("A1", "V2", A,
     "Dựng BỘ NHIỆM VỤ cho thước đo, chạy trên ≥ 2 Platform Pack",
     BS, "SC-03",
     "1) N nhiệm vụ sinh module, mỗi nhiệm vụ có tiêu chí máy chấm được. "
     "2) k lượt mỗi nhiệm vụ để tính pass@k đúng công thức không chệch. "
     "3) Hạng kết cục lấy từ BÁO CÁO CHUỖI CỔNG, không từ trường tự khai. "
     "4) Bo thật và máy chủ đứng RIÊNG. 5) `bench.py` chỉ GOM số từ bộ dò đã có.",
     "eaa/bench.py · packs/avr · packs/stm32 · projects/*",
     "`eaa report bench` ra số thật cho cả bốn trục, trên ≥2 pack",
     "TC-138 canh thước; cần thêm bài canh bộ nhiệm vụ",
     "D9 · ngân sách gọi mô hình", "—",
     "Trộn hai hạng bằng chứng rồi báo một con số. Và cám dỗ chọn nhiệm vụ "
     "mà ta biết mình làm tốt",
     "3–5 ngày + token", "1", "CHỜ NGÂN SÁCH"),
    ("A2", "V3", A,
     "Đo NGƯỢC lịch sử dự án bằng bốn bộ dò",
     BS, "SC-04, SC-01",
     "Chốt chân lý nền TRƯỚC khi chạy · dựng cặp từ `llm_calls.jsonl` · chạy "
     "bốn bộ dò · tách BẮT ĐƯỢC / BỎ SÓT / KHÔNG CHẠY ĐƯỢC · công bố cả phần "
     "dự đoán sai.",
     "scripts/do_nguoc_lich_su.py · docs/CHAN_LY_NEN_V3.md",
     "ĐÃ ĐO: bắt 3/13 · bỏ sót 0 · kêu trật 7 · sớm hơn người 54 phút",
     "TC-146 (8 bài) — canh cả việc chân lý nền không bị sửa sau khi chốt",
     "—", "—",
     "Tự chấm chính mình. Chân lý nền phải chốt trước và commit riêng",
     "2 ngày", "1", "XONG (SL-179)"),
    ("A3", "—", A,
     "Chạy lại phép đo ngược trên Platform Pack thứ hai",
     BS, "SC-04",
     "1) Lịch sử pack thứ hai ngắn hơn — nói rõ cỡ mẫu. 2) Tính sạch KHÔNG "
     "còn: bộ dò nay đã tồn tại trước dữ liệu mới, nên đây là phép đo khác "
     "hạng với A2 và không được gộp số. 3) Báo riêng hai bảng.",
     "scripts/do_nguoc_lich_su.py",
     "Bảng thứ hai, kèm câu khai rằng nó KHÔNG phải hồi cứu sạch",
     "TC-146 mở rộng",
     "A1 · A2", "—",
     "Gộp số của hai phép đo khác hạng là hạng lỗi bản benchmark đã cảnh báo",
     "1 ngày", "3", "CHƯA"),

    # ── B · tri thức ────────────────────────────────────────────────────────
    ("B1", "V4", B,
     "Thủ tục theo NGOẠI VI, nạp vào prompt và duyệt qua G2",
     CL, "SC-05",
     "1) Thủ tục KHÁC trích đoạn: trích đoạn nói bit, thủ tục nói THỨ TỰ và "
     "bẫy. 2) Ngân sách MƯỢN của lớp trích đoạn, không cộng thêm — nếu không "
     "thì ablation đo 'nhiều ngữ cảnh hơn' chứ không đo thủ tục. 3) Vào kho "
     "qua G2, không cửa sau. 4) Bẫy phải có xuất xứ và mức tin cậy.",
     "eaa/procedure.py · eaa/composer.py · packs/avr/procedures/ · "
     "projects/*/procedures/",
     "CƠ CHẾ XONG. Con số pass@k trước/sau CHƯA ĐO ĐƯỢC — cần A1",
     "TC-139 (21 bài, 5 đột biến đều bị bắt)",
     "A1", "—",
     "Con số mượn 800 token là CANH BẠC chứ không phải phép đo. Ablation thua "
     "thì con số phải giảm, không phải phép đo bị giải thích lại",
     "3–4 ngày", "2", "CƠ CHẾ XONG — số chờ A1"),
    ("B2", "V6", B,
     "Nối bộ soi THỨ NGUYÊN vào bản đồ thanh ghi",
     CL, "SC-11",
     "1) `dimension.py` đang tra sổ số đo để lấy đơn vị. 2) Nối nguồn thứ hai: "
     "bề rộng và mặt nạ trường bit từ `regmap.py`. 3) Bắt hạng 'gán vượt bề "
     "rộng trường mà chú thích vẫn khai đúng'. 4) Giữ BA TRẠNG THÁI, không ép "
     "về đúng/sai.",
     "eaa/dimension.py · eaa/regmap.py",
     "Số chú thích sai thứ nguyên bắt thêm được trên 13 ca của A2",
     "TC-134 mở rộng",
     "A2", "—",
     "Bộ soi hay báo nhầm sớm muộn cũng bị tắt đi — A2 đã đo 7/13 kêu trật",
     "1–2 ngày", "4", "CHƯA"),
    ("B3", "V11", B,
     "Mở rộng danh sách miền CHÍNH CHỦ sang hãng chip của pack thứ hai",
     MR, "SC-17",
     "1) Thêm miền của hãng. 2) KHÔNG nới luật: hạng vẫn tính theo URL CUỐI "
     "sau chuyển hướng, và trang chính chủ vẫn là SUY RA. 3) Thêm miền là "
     "thêm DỮ LIỆU, không sửa mã cưỡng chế.",
     "eaa/web.py · packs/stm32/",
     "Số datasheet chính chủ nạp được cho pack thứ hai",
     "TC-65, TC-66 đã canh luật",
     "—", "—",
     "Nới một miền vì tiện là cách danh sách trắng thành trang trí",
     "0,5 ngày", "5", "CHƯA"),
    ("B4", "V8", B,
     "Đóng nốt hai dòng CHƯA và hai dòng MỘT PHẦN của bảng năng lực",
     BS, "SC-14",
     "1) Kiểm lại nhánh chưa chạy được của N-036/N-100. 2) N-908/N-909 còn "
     "thiếu phần đòi biết bài toán. 3) Với hai dòng ấy: lấp nốt HOẶC chốt "
     "rằng nhánh ấy thuộc về người và ghi lý do — không để treo.",
     "eaa/lifecycle.py · eaa/sensitivity.py · eaa/instrument.py",
     "Sheet Khoảng trống còn bao nhiêu dòng, và mỗi dòng có một quyết định",
     "TC-129, TC-131, TC-132 đã có",
     "A2", "—",
     "Lấp cho đủ số tệ hơn để trống có lý do",
     "2 ngày", "6", "CHƯA"),

    # ── C · phần cứng ───────────────────────────────────────────────────────
    ("C1", "V5", C,
     "Đọc netlist và ĐỐI CHIẾU với hồ sơ phần cứng tại G1",
     CL, "SC-06",
     "1) Đọc netlist ra mô hình trung lập. 2) KHÔNG thay hồ sơ — chỉ nêu chỗ "
     "lệch. 3) Chỗ lệch trình ở G1 cho người phân xử. 4) Netlist là SUY RA: "
     "nó là ý định thiết kế, không phải mạch đã hàn.",
     "eaa/netlist.py (mới) · eaa/gates.py",
     "Số chỗ lệch tìm được trên dự án thật",
     "TC-140 (đã đặt chỗ, chưa viết)",
     "—", "—",
     "Tự động 'sửa' hồ sơ theo netlist là làm mất chính thứ G1 sinh ra để hỏi",
     "3 ngày", "5", "CHƯA"),
    ("C2", "V7", C,
     "Trọng tài phần cứng: khoá cổng nối tiếp giữa các phiên",
     CL, "SC-15",
     "1) Khoá theo tệp, có tên phiên và dấu thời gian. 2) Phiên thứ hai bị "
     "chặn kèm thông báo nói RÕ ai đang giữ. 3) Khoá chết phải tự hết hạn.",
     "eaa/serialport.py · eaa/telemetry.py",
     "Hai phiên đồng thời: phiên hai bị chặn với thông báo rõ",
     "cần TC mới",
     "—", "—",
     "Sở cứ hạng SUY RA — chưa ai chạy hai phiên. Vào bảng kèm nhãn ấy",
     "1 ngày", "7", "CHƯA"),
    ("C3", "V12", C,
     "Đo tải CPU của ISR bước 50 kHz trên bo thật",
     BS, "SC-13",
     "1) Lật chân GPIO đầu/cuối ISR, đọc bằng máy hiện sóng hoặc bộ đếm. "
     "2) Số đo vào `board_facts.jsonl` qua G4, rồi chảy ngược vào prompt ở "
     "lớp K8. 3) Vượt ngưỡng thì đó là ràng buộc thiết kế, không phải lỗi mã.",
     "projects/robot_balance/board_facts.jsonl",
     "Một con số phần trăm CPU, có xuất xứ và đơn vị, đã duyệt qua G4",
     "TC-133 canh đường số đo vào prompt",
     "bo và máy đo", "—",
     "Số đo trên bo đã cãi lại tài liệu một lần. Đo bằng suy luận thay vì "
     "bằng que đo là lặp lại chính lỗi ấy",
     "0,5 ngày + thiết bị", "3", "CHƯA"),
    ("C4", "V10", C,
     "Ba việc chờ NGƯỜI bấm — không được uỷ quyền cho hệ",
     NG, "SC-10, SC-07",
     "1) Duyệt lại G1 cho hồ sơ đang bật cảnh báo TRÔI. 2) Khai `observable:` "
     "trong `hardware_profile.yaml`. 3) Ghi số đo gia tốc mốc vào sổ số đo "
     "nếu đó là số đã duyệt.",
     "projects/robot_balance/hardware_profile.yaml",
     "Cảnh báo TRÔI tắt; `eaa observe` không còn báo 0 kênh quan sát",
     "TC-135 canh phần hệ",
     "—", "—",
     "Hệ tự bấm ba việc này là phá đúng bất biến 5 cổng người duyệt",
     "30 phút của người", "2", "CHỜ NGƯỜI"),
    ("C5", "V14", C,
     "Điều khiển máy đo và phiên gỡ lỗi tự động",
     MR, "SC-08, SC-09",
     "1) Khoảng cách phần cứng lớn nhất so với mốc Embedder. 2) Tốn THIẾT BỊ "
     "chứ không chỉ tốn mã. 3) Chỉ có nghĩa khi đã có thước (A1) để chứng "
     "minh nó cải thiện được gì.",
     "eaa/instrumentctl.py (mới)",
     "Số kịch bản chẩn đoán chạy được không cần người cầm que đo",
     "cần TC mới",
     "A1 · thiết bị", "—",
     "Làm sớm là đúng cái bẫy bảng này sinh ra để tránh: thêm năng lực không "
     "ai có, rồi không có số nào chứng minh nó tốt hơn",
     "1–2 tuần + thiết bị", "9", "CHƯA"),

    # ── D · mô phỏng ────────────────────────────────────────────────────────
    ("D1", "—", D,
     "TIÊM LỖI vào kịch bản mô phỏng",
     CL, "SC-19, SC-20",
     "1) Thêm khối `faults:` vào `scenarios.yaml`: bus chết ở giây thứ N, cảm "
     "biến kẹt giá trị, mẫu không về, sụt áp. Lớp HAL giả lập thực thi. "
     "2) Mỗi kịch bản lỗi phải khai HÀNH VI MONG ĐỢI (chuyển trạng thái an "
     "toàn), không chỉ 'không sập' — một kịch bản chỉ đòi không sập sẽ xanh "
     "với mã đứng im. 3) Chạy NGƯỢC trên 13 ca của A2 để đo nó bắt được mấy ca.",
     "projects/*/sim/scenarios.yaml · packs/*/hostmock/ · eaa/tools/sim.py",
     "Số lần từ chối G3 hạng 'im lặng dùng số cũ' mà một kịch bản tiêm lỗi bắt "
     "được, đo ngược trên 13 ca. KHÔNG tốn token API",
     "cần TC mới",
     "A2 · D6", "—",
     "Kịch bản lỗi viết lỏng thì nó xanh với mã đứng im — đó là bài kiểm rỗng "
     "ở tầng mô phỏng, đúng hạng `sensitivity.py` sinh ra để bắt",
     "3 ngày", "1", "CHƯA"),
    ("D2", "—", D,
     "Mô phỏng CHU KỲ MÁY (PIL) cho ngân sách thời gian",
     BS, "SC-13, SC-21",
     "1) Cổng SIL biên dịch cho MÁY CHỦ nên không có mô hình chu kỳ — nó "
     "không trả lời được câu 'ISR ăn bao nhiêu CPU'. 2) Nối một bộ mô phỏng "
     "tập lệnh của họ chip làm cổng CẢNH BÁO trước, chặn sau. 3) Khai rõ mức "
     "tin cậy: kết quả PIL là SUY RA, không phải ĐÃ KIỂM — bo thật vẫn là C3.",
     "packs/avr/tools.yaml · eaa/tools/pil.py (mới)",
     "Số chu kỳ mỗi vòng điều khiển, và phần trăm ngân sách đã dùng",
     "cần TC mới",
     "—", "—",
     "Thêm một NGUỒN SỰ THẬT thứ ba. Nếu PIL và bo thật lệch nhau thì phải có "
     "luật ai thắng — và luật ấy phải viết TRƯỚC khi có số, không phải sau",
     "4–5 ngày", "4", "CHƯA"),
    ("D3", "—", D,
     "Đo KHOẢNG CÁCH giữa mô phỏng và bo thật",
     BS, "SC-26, SC-09",
     "1) Với mỗi đại lượng mô phỏng DỰ ĐOÁN và bo ĐO ĐƯỢC, ghi cặp (dự đoán, "
     "đo) và sai số tương đối. 2) Sai số vượt ngưỡng → mô hình phải được hiệu "
     "chỉnh, HOẶC mức tin cậy của mọi kết luận từ mô phỏng bị hạ xuống. "
     "3) Con số này vào báo cáo Chương 3 như một đại lượng riêng.",
     "eaa/simgap.py (mới) · measurements.jsonl · board_facts.jsonl",
     "Sai số tương đối theo từng đại lượng, có xuất xứ hai đầu",
     "cần TC mới",
     "C3 · C4", "—",
     "Đây là ĐÓNG GÓP: không sản phẩm nào trong khảo sát công bố con số này. "
     "Cũng vì thế không có mốc để đối chiếu, nên phải nói rõ nó chưa được "
     "ai khác kiểm",
     "2–3 ngày", "2", "CHƯA"),
    ("D4", "—", D,
     "MỐC HỒI QUY cho mô phỏng qua các lượt sinh lại",
     CL, "SC-23, SC-24",
     "1) Cổng hiện chỉ ĐẠT/TRƯỢT: một bản sinh lại làm thời gian xác lập xấu "
     "đi vẫn ĐẠT. 2) Lưu số liệu mô phỏng theo từng bản ĐÃ MERGE. 3) Cổng "
     "cảnh báo khi xấu đi quá ngưỡng, và ngưỡng là dữ liệu của dự án chứ "
     "không đóng cứng trong engine.",
     "eaa/tools/sim.py · projects/*/sim_baseline.jsonl (mới)",
     "Số lượt sinh lại bị bắt vì xấu đi, đo ngược trên 12 lượt của A2",
     "cần TC mới",
     "A2 · D1 · D6", "—",
     "Ngưỡng đặt lỏng thì cổng vô dụng; đặt chặt thì nó kêu mỗi lượt và bị "
     "tắt đi. Ngưỡng phải rút TỪ dữ liệu 12 lượt cũ, không gõ tay",
     "2 ngày", "3", "CHƯA"),
    ("D5", "—", D,
     "Độ phủ kịch bản theo TRẠNG THÁI và nhánh",
     BS, "SC-25",
     "1) Đo trạng thái/nhánh nào kịch bản chạm tới. 2) NÊU RA chỗ chưa chạm — "
     "từ chối #12 cho thấy nhánh NGÃ không kịch bản nào đi tới, và lỗi nằm "
     "đúng ở đó. 3) Không tự sinh kịch bản: chỗ chưa phủ là câu hỏi cho người.",
     "eaa/simcov.py (mới) · packs/*/hostmock/",
     "Danh sách trạng thái chưa kịch bản nào chạm tới",
     "cần TC mới",
     "D1", "—",
     "Độ phủ cao KHÔNG có nghĩa là đúng. Báo độ phủ mà không nói câu ấy là "
     "mời người đọc kết luận sai",
     "2 ngày", "5", "CHƯA"),
    ("D6", "—", D,
     "Tất định: hạt giống và chạy lại ra cùng số",
     NM, "SC-22",
     "1) Mọi lượt chạy mô phỏng ghi hạt giống vào báo cáo. 2) Cùng hạt giống "
     "cho cùng số liệu, tới đúng chữ số. 3) Cùng luật TC-15 đã đặt cho lượt "
     "gọi mô hình.",
     "eaa/tools/sim.py · eaa/tools/sim_runner.py",
     "Chạy hai lần cùng hạt giống cho cùng số liệu",
     "cần TC mới (đối xứng với TC-15)",
     "—", "—",
     "Không tất định thì D4 báo hồi quy giả, và người sẽ thôi tin cả cổng",
     "1 ngày", "2", "CHƯA"),
    ("D7", "—", D,
     "Quét Monte-Carlo theo dung sai linh kiện và nhiễu cảm biến",
     BS, "SC-27",
     "1) Khai phân bố dung sai trong hồ sơ phần cứng. 2) Quét N lượt, báo "
     "phần trăm lượt đạt chứ không báo một lượt. 3) Kết quả là XÁC SUẤT, và "
     "phải được đọc như xác suất — không được viết thành 'robot đứng được'.",
     "projects/*/hardware_profile.yaml · eaa/tools/sim.py `sweep`",
     "Phần trăm lượt đạt trên N lượt, kèm phân bố đã dùng",
     "cần TC mới",
     "D6", "—",
     "Sở cứ hạng SUY RA. Và một con số xác suất rất dễ bị trích dẫn lại thành "
     "một lời khẳng định chắc chắn",
     "3 ngày", "7", "CHƯA"),
    ("D8", "—", D,
     "`eaa sim` thành lệnh của NGƯỜI, không chỉ là cổng của máy",
     MR, "SC-28",
     "1) `eaa sim run <kịch bản>` · `eaa sim sweep` · `eaa sim report`. "
     "2) Gọi ĐÚNG `SimGate` mà chuỗi cổng gọi — không dựng đường chạy thứ "
     "hai. 3) Chạy tay KHÔNG ghi kết quả vào Project State: xem thử không "
     "phải là kiểm chứng.",
     "eaa/cli.py · eaa/tools/sim.py",
     "Số bước từ 'muốn thử một tham số' tới 'thấy đồ thị'",
     "cần TC mới",
     "—", "—",
     "Một đường chạy thứ hai tới kết quả mô phỏng là chỗ để lách: chạy tay "
     "cho đẹp rồi khai là cổng đã đạt",
     "1–2 ngày", "3", "CHƯA"),
    ("D9", "—", D,
     "Mô phỏng cho Platform Pack thứ hai",
     BS, "SC-21",
     "1) A1 đòi ≥2 pack; pack thứ hai hiện chưa có mô hình vật lý nào. "
     "2) Mô hình mới phải qua đúng phép kiểm chứng bằng nghiệm giải tích như "
     "mô hình thứ nhất TRƯỚC khi được làm cổng chặn.",
     "packs/stm32/ · projects/<dự án stm32>/sim/",
     "Cổng SIL chạy được trên pack thứ hai, mô hình đã kiểm chứng",
     "đối xứng với tests/test_sim_verification.py",
     "—", "—",
     "Một cổng dựa trên mô hình CHƯA kiểm chứng còn tệ hơn không có cổng — nó "
     "phát ra phán quyết có vẻ khách quan về thứ nó không mô tả đúng",
     "4–5 ngày", "6", "CHƯA"),

    # ── E · IDE ─────────────────────────────────────────────────────────────
    ("E1", "V13", E,
     "`--json` cho MỌI lệnh CHỈ ĐỌC",
     NM, "SC-16",
     "1) Đây là chỗ chặn thật của cả mảng E: không có đầu ra máy đọc được thì "
     "không extension nào làm được gì. 2) CHỈ lệnh chỉ đọc. Lệnh đổi trạng "
     "thái vẫn đi đúng đường cũ — không có đường thứ hai tới merge (TC-01, "
     "TC-02). 3) Lược đồ đầu ra là HỢP ĐỒNG: đổi nó phải qua một mục sai lệch.",
     "eaa/cli.py · eaa/jsonout.py (mới)",
     "Số lệnh chỉ đọc có `--json`, trên tổng số lệnh chỉ đọc",
     "cần TC mới — canh cả việc lệnh GHI không có `--json`",
     "—", "—",
     "Cám dỗ thêm `--json` cho lệnh ghi 'cho tiện tự động hoá'. Đó chính là "
     "đường thứ hai tới merge mà bất biến số một cấm",
     "3 ngày", "1", "CHƯA"),
    ("E2", "—", E,
     "Lỗi cổng thành CHẨN ĐOÁN CÓ VỊ TRÍ trong biên tập",
     MR, "SC-18",
     "1) `ToolError` đã mang sẵn `file`, `line`, `rule_id`, `severity` — việc "
     "này gần như chỉ là xuất ra. 2) `eaa verify --json` trả danh sách chẩn "
     "đoán. 3) Lỗi KHÔNG có vị trí vẫn phải xuất, gắn vào tệp dự án — nuốt "
     "chúng đi là giấu đúng những lỗi khó nhất.",
     "eaa/cli.py · eaa/tools/base.py",
     "Phần trăm lỗi cổng có `file`+`line` dùng được để vẽ gạch đỏ",
     "cần TC mới",
     "E1", "—",
     "Chỉ hiện lỗi có vị trí thì lỗi thiết kế — hạng chiếm 8/13 lần từ chối — "
     "biến mất khỏi màn hình",
     "1–2 ngày", "2", "CHƯA"),
    ("E3", "—", E,
     "Bảng TRẠNG THÁI GATE ngay trong biên tập",
     MR, "SC-16",
     "1) Hiện 5 gate, gate nào đang chờ, lý do từ chối gần nhất. 2) Bấm duyệt "
     "trong IDE vẫn gọi ĐÚNG `eaa gate approve` với cùng dấu vân tay nội "
     "dung — không có API riêng, không có đường vòng. 3) IDE không được tự "
     "bấm bất cứ gate nào.",
     "tích hợp riêng · eaa/cli.py (chỉ đọc)",
     "Số bước từ mở dự án tới thấy 'đang mắc ở gate nào'",
     "TC-01, TC-02 đã canh bất biến; cần TC cho lớp mỏng",
     "E1", "—",
     "Một nút 'Duyệt' đặt cạnh mã rất dễ bị bấm theo phản xạ. Hồ sơ gate phải "
     "hiện ĐỦ trước khi nút bấm được",
     "3 ngày", "3", "CHƯA"),
    ("E4", "—", E,
     "Nhảy từ LÝ DO TỪ CHỐI tới đúng dòng mã",
     MR, "SC-29",
     "1) 13 lý do từ chối dài trung bình hơn 800 ký tự, nhắc đích danh hàm và "
     "hằng số. 2) Rút tên hàm/hằng số trong lý do thành liên kết tới vị trí. "
     "3) Rút KHÔNG chắc thì không tạo liên kết — một liên kết trỏ sai chỗ tệ "
     "hơn không có liên kết.",
     "eaa/gatelink.py (mới) · gates/decisions.jsonl",
     "Phần trăm lý do từ chối có ít nhất một liên kết đúng chỗ",
     "cần TC mới",
     "E1 · E2", "—",
     "Cùng hạng lỗi TC-145 đã chặn ở gợi ý CLI: mũi tên chỉ vào tường",
     "2 ngày", "5", "CHƯA"),
    ("E5", "—", E,
     "Di chuột lên `// ref:` hiện ngay TRÍCH ĐOẠN",
     MR, "SC-30",
     "1) FR-RAG-02 bắt mã cấu hình thanh ghi phải mang `// ref:`; muốn đọc "
     "trích đoạn ấy hiện phải rời biên tập. 2) Hiện kèm TRẠNG THÁI chunk: đã "
     "duyệt G2 hay chưa, có bị thay thế chưa. 3) Trích dẫn trỏ vào chunk "
     "không có thật phải hiện thành CẢNH BÁO, không hiện thành rỗng.",
     "tích hợp riêng · eaa/kb.py (chỉ đọc)",
     "Số lần phải rời biên tập để tra một trích dẫn",
     "TC-17 canh trích dẫn bắt buộc",
     "E1", "—",
     "Hiện rỗng khi trích dẫn sai là biến một lỗi thành một khoảng lặng",
     "2 ngày", "4", "CHƯA"),
    ("E6", "—", E,
     "So bản ỨNG VIÊN với bản ĐÃ MERGE, kèm cảnh báo bộ dò",
     MR, "SC-04",
     "1) A2 cho thấy 12 lượt sinh lại và 67 lượt bộ dò kêu. 2) Hiện diff kèm "
     "cảnh báo của `contract`/`instrument`/`sensitivity` đúng chỗ dòng. "
     "3) Cảnh báo phải kèm HẠNG: A2 đo được 7/13 là kêu trật, nên hiện chúng "
     "ngang hàng với cảnh báo đúng là dạy người bỏ qua cả hai.",
     "tích hợp riêng · eaa/contract.py · instrument.py · sensitivity.py",
     "Thời gian từ 'có bản mới' tới 'biết nó đổi gì đáng ngờ'",
     "cần TC mới",
     "E1 · A2", "—",
     "Bộ dò kêu trật 7/13. Đưa cả 67 cảnh báo lên màn hình là cách nhanh "
     "nhất để người tắt hết chúng đi",
     "3 ngày", "6", "CHƯA"),
    ("E7", "—", E,
     "Bảng TELEMETRY trực tiếp trong biên tập",
     MR, "SC-08",
     "1) `eaa ports --watch` đã có đường đọc. 2) Vẽ đồ thị đại lượng theo "
     "thời gian, cạnh mã sinh ra nó. 3) Số hiện trên bảng là SỐ ĐO — muốn nó "
     "vào kho tri thức vẫn phải qua G4.",
     "tích hợp riêng · eaa/telemetry.py (chỉ đọc)",
     "Số bước từ 'nạp xong' tới 'thấy đồ thị góc nghiêng'",
     "cần TC mới",
     "E1 · C2", "—",
     "Một số đẹp trên đồ thị rất dễ bị chép thẳng vào tài liệu mà không qua "
     "G4 — lúc ấy nó là số ĐÃ KIỂM giả",
     "3 ngày", "7", "CHƯA"),
    ("E8", "—", E,
     "Một lệnh dựng lại môi trường cho máy mới, gọi từ IDE",
     MR, "SC-16",
     "1) `eaa doctor` đã có. 2) IDE gọi nó lúc mở dự án và hiện kết quả dưới "
     "dạng danh sách việc phải làm. 3) Cài công cụ VẪN cần người xác nhận — "
     "IDE không được tự cài (TC-28, TC-34).",
     "eaa/doctor.py (chỉ đọc) · tích hợp riêng",
     "Số bước từ clone kho tới lượt sinh đầu tiên trên máy mới",
     "TC-28, TC-34 đã canh cổng cài",
     "E1", "—",
     "Một nút 'Sửa hết đi' là cách bỏ qua cổng cài mà không ai nhận ra",
     "1 ngày", "8", "CHƯA"),

    # ── F · truy vết ────────────────────────────────────────────────────────
    ("F1", "V9", F,
     "Truy vết HAI CHIỀU: yêu cầu ↔ ca kiểm ↔ mã ↔ quyết định gate",
     MR, "SC-12",
     "1) Đã có một nửa: mỗi SL có TC canh, mỗi lượt sinh có prompt hash và "
     "chunk id. 2) Còn thiếu chiều NGƯỢC: từ một yêu cầu, ra mọi TC và mọi "
     "dòng mã phục vụ nó. 3) Dựng chỉ mục từ dữ liệu ĐÃ CÓ trong Git, không "
     "bắt người khai thêm.",
     "eaa/trace.py (mới) · eaa/vcs.py",
     "Từ một mã yêu cầu bất kỳ, ra được danh sách TC và commit — kiểm tay 10 mẫu",
     "cần TC mới",
     "—", "—",
     "Mốc gần ta nhất về triết lý, nên cũng là chỗ dễ khai quá lời nhất. Ta "
     "KHÔNG làm chứng nhận; ta làm truy vết. Phải nói rõ khác biệt",
     "3 ngày", "4", "CHƯA"),
    ("F2", "V1", F,
     "Thông báo lỗi nói được VIỆC PHẢI LÀM ≥ 80%",
     MR, "SC-02",
     "Đo lại mốc bằng phép đếm ngoặc cân · gom gợi ý vào MỘT bảng · gắn ở "
     "`main()` nơi biết người dùng vừa gõ lệnh nào · thông báo đã tự nêu lệnh "
     "thì không gắn thêm.",
     "eaa/cli.py",
     "ĐÃ ĐO: 150/182 = 82%, mốc cũ 25/182 = 14%",
     "TC-145 (11 bài, 4 đột biến đều bị bắt)",
     "—", "—",
     "Cám dỗ nới định nghĩa phép đo cho dễ đạt",
     "1 ngày", "1", "XONG (SL-178)"),
    ("F3", "—", F,
     "Bản bàn giao dựng LẠI ĐƯỢC từ trạng thái kho, không gõ tay",
     BS, "SC-14",
     "1) `docs/TIEP_TUC_TU_DAY.md` hiện gõ tay, nên nó cũ đi mà không ai "
     "biết — cùng hạng lỗi đã sửa cho con số '1.428 test'. 2) Sinh phần SỐ từ "
     "phép đếm tại chỗ. 3) Phần PHÁN ĐOÁN ('nên làm gì tiếp') vẫn do người "
     "viết — máy không suy được ý định.",
     "scripts/lam_ban_ban_giao.py (mới) · docs/TIEP_TUC_TU_DAY.md",
     "Số con số gõ tay còn lại trong bản bàn giao",
     "cần TC mới",
     "—", "—",
     "Sinh cả phần phán đoán là để máy nói hộ người một câu chỉ người biết",
     "1 ngày", "6", "CHƯA"),
]

def _suy_ra_chan(viec: list[tuple[str, ...]]) -> list[tuple[str, ...]]:
    """Cột `Chặn việc nào` SUY RA từ cột `Phụ thuộc`, không khai tay.

    Bản đầu khai cả hai chiều bằng tay, và phép kiểm toàn vẹn tìm ra **15 cạnh
    không đối xứng**: A khai chặn B trong khi B không khai chờ A. Một đồ thị
    phụ thuộc tự mâu thuẫn còn tệ hơn không có đồ thị — người đọc tin nó.

    Cùng hình dạng lỗi mà V3 tìm ra trong `contract.py`: hai danh sách cho
    cùng một mục đích là hai danh sách sẽ lệch nhau. Khai một chiều, suy ra
    chiều kia.
    """
    chan: dict[str, list[str]] = {}
    for v in viec:
        for x in v[10].split(" · "):
            x = x.strip()
            if re.fullmatch(r"[A-F]\d+", x):
                chan.setdefault(x, []).append(v[0])
    return [
        v[:11] + (" · ".join(sorted(chan.get(v[0], []))) or "—",) + v[12:]
        for v in viec
    ]


VIEC = _suy_ra_chan(VIEC)

COT = [
    ("Mã", 6), ("Mã cũ", 7), ("Mảng", 24), ("Việc — làm gì", 42),
    ("Hạng việc", 17), ("Sở cứ", 13), ("Làm NHƯ THẾ NÀO", 66),
    ("Đụng tệp nào", 32), ("Nghiệm thu — xong thì ĐO bằng gì", 36),
    ("Bài kiểm canh", 26), ("Phụ thuộc", 15), ("Chặn việc nào", 15),
    ("Rủi ro — cái dễ làm hỏng", 42), ("Công", 14), ("Ưu tiên", 8),
    ("Trạng thái", 18),
]

# ── nhật ký tiến hoá ─────────────────────────────────────────────────────────

NHAT_KY: list[tuple[str, str, str, str]] = [
    ("05/09/2026", "V1 → F2", "XONG — 82%, SL-178, TC-145",
     "Đính chính mốc đã công bố: 36% là SAI, mốc thật 14%"),
    ("05/09/2026", "bảng 14 việc", "Lập lần đầu",
     "Sinh từ bản benchmark sản phẩm và bản rà soát năng lực"),
    ("05/09/2026", "V3 → A2", "XONG — SL-179, TC-146",
     "Bắt 3/13 · bỏ sót 0 · kêu trật 7. Dự đoán chốt trước là 4–5/13; "
     "ba chỗ dự đoán sai được công bố nguyên vẹn"),
    ("05/09/2026", "V4 → B1", "CƠ CHẾ XONG — SL-180, TC-139",
     "Con số ablation vẫn chờ A1. Bản nháp ghi đè mất eaa/skills.py, đã khôi "
     "phục và đổi tên thành eaa/procedure.py"),
    ("05/09/2026", "V13 → E1..E8", "MỘT DÒNG nở thành TÁM việc",
     "Bảng cũ cho IDE đúng một dòng 'lớp mỏng tích hợp IDE'. Rà soát lại thì "
     "chỗ chặn nằm ở E1: CLI chưa có đầu ra máy đọc được. Mã V13 giữ ở E1"),
    ("05/09/2026", "cột 'Chặn việc nào'", "Đổi từ KHAI TAY sang SUY RA",
     "Phép kiểm toàn vẹn tìm ra 15 cạnh không đối xứng và 3 cặp vòng tròn "
     "(C1↔C4, C3↔D2, D2↔D3) do khai cả hai chiều bằng tay. Nay khai một "
     "chiều, suy ra chiều kia — cùng hình dạng lỗi V3 tìm ra trong contract.py"),
    ("05/09/2026", "toàn bảng", "THIẾT KẾ LẠI thành backlog tiến hoá",
     "14 việc → 32 việc, chia sáu mảng. Mảng D (mô phỏng) và E (IDE) dựng "
     "mới sau khi rà soát: kho đã có MiL + SIL chạy firmware thật, còn IDE "
     "bị chặn ở chỗ CLI chưa có đầu ra máy đọc được"),
]

# ── bảng này KHÔNG nói gì ────────────────────────────────────────────────────

KHONG_NOI: list[str] = [
    "KHÔNG cho điểm tổng. Sở cứ thuộc năm hạng khác nhau; cộng chúng lại "
    "thành một con số là làm ra một con số sai theo hướng có lợi cho người "
    "viết — tức hướng khó tự phát hiện nhất.",
    "Cột Công là ƯỚC LƯỢNG. Ba việc đã xong (F2, A2, B1) chưa được đối chiếu "
    "ước lượng với thời gian thật, nên cột này vẫn chưa có độ tin cậy nào.",
    "Bốn việc dựa một phần vào sở cứ hạng SUY RA: C2, D7, và phần lập luận "
    "của C5. Chúng nằm trong bảng kèm nhãn, không lẫn vào việc dựa trên lỗi "
    "đã xảy ra thật.",
    "Mảng E (IDE) mô tả một mặt tiếp xúc CHƯA TỒN TẠI. Mọi con số nghiệm thu "
    "của mảng ấy là con số sẽ đo, không phải con số đã đo.",
    "Mảng D (mô phỏng) xây trên nền đã có — cổng SIL chạy firmware C thật qua "
    "lớp HAL giả lập, mô hình vật lý đã kiểm chứng bằng nghiệm giải tích. "
    "Chín việc của mảng là phần THIẾU, không phải phần dựng lại từ đầu.",
    "Thứ tự ưu tiên xếp theo 'nuôi được chiều sâu bao nhiêu', không theo 'dễ "
    "làm bao nhiêu'. A1, D1, E1 đứng đầu vì mỗi cái chặn nhiều việc khác.",
    "Một việc XONG nghĩa là có bài kiểm canh nó, không phải đã chạy thử một lần. "
    "B1 ghi 'XONG CƠ CHẾ' chứ không ghi XONG, vì con số chứng minh nó có tác "
    "dụng thì chưa có.",
]


def main() -> int:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Cần openpyxl: pip install openpyxl", file=sys.stderr)
        return 1

    DAM = Font(bold=True, color="FFFFFF")
    NEN = PatternFill("solid", fgColor="1F3864")
    TREN = Alignment(vertical="top", wrap_text=True)
    GIUA = Alignment(horizontal="center", vertical="top", wrap_text=True)
    VIEN = Border(*(Side(style="thin", color="BFBFBF"),) * 4)

    MAU_HANG = {
        DO: PatternFill("solid", fgColor="C6EFCE"),
        SO: PatternFill("solid", fgColor="DDEBF7"),
        VL: PatternFill("solid", fgColor="E4DFEC"),
        KHAI: PatternFill("solid", fgColor="FFF2CC"),
        SR: PatternFill("solid", fgColor="FFC7CE"),
    }
    MAU_TT = {
        "XONG": PatternFill("solid", fgColor="C6EFCE"),
        # Không đặt tên trạng thái này bắt đầu bằng "XONG": một việc mới xong
        # một nửa mà đọc thành đã xong là chỗ bảng tự nói dối. TC-147 bắt được.
        "CƠ CHẾ XONG": PatternFill("solid", fgColor="FFF2CC"),
        "ĐANG LÀM": PatternFill("solid", fgColor="FFF2CC"),
        "CHỜ NGÂN SÁCH": PatternFill("solid", fgColor="E7E6E6"),
        "CHỜ NGƯỜI": PatternFill("solid", fgColor="DDEBF7"),
        "CHƯA": PatternFill("solid", fgColor="FFC7CE"),
    }
    MAU_MANG = {
        A: "DDEBF7", B: "E2EFDA", C: "FCE4D6",
        D: "FFF2CC", E: "E4DFEC", F: "F2F2F2",
    }

    wb = Workbook()

    def dau(ws, cot):
        for i, (ten, rong) in enumerate(cot, 1):
            o = ws.cell(row=1, column=i, value=ten)
            o.font, o.fill, o.alignment, o.border = DAM, NEN, GIUA, VIEN
            ws.column_dimensions[get_column_letter(i)].width = rong
        ws.row_dimensions[1].height = 32
        ws.freeze_panes = "A2"

    # ── 1 · đọc trước ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Đọc trước"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 116
    doc = [
        ("EAA — backlog tiến hoá", ""),
        ("", ""),
        ("Bảng này để làm gì",
         "Quản trị quá trình tiến hoá của sản phẩm: việc gì · dựa vào đâu · "
         "làm thế nào · xong thì đo bằng gì · đang chặn việc nào."),
        ("Thay cho bảng nào",
         "Thay `docs/EAA_Viec_phai_lam.xlsx` (14 việc). Mọi mã V1..V14 mang "
         "sang trong cột 'Mã cũ'. Bảng cũ được rút đi chứ không để song song: "
         "hai danh sách việc phải làm là hai danh sách sẽ lệch nhau."),
        ("Sáu mảng",
         " · ".join(m[0] for m in MANG)),
        ("Hai mảng dựng mới",
         "D · Mô phỏng và E · IDE. Bảng cũ cho IDE đúng một dòng và không có "
         "dòng nào cho mô phỏng. Rà soát lại thì kho đã có MiL + cổng SIL "
         "chạy FIRMWARE C THẬT qua lớp HAL giả lập từ Sprint 3, còn IDE bị "
         "chặn ở một chỗ rất cụ thể: CLI chưa có đầu ra máy đọc được."),
        ("Cột quan trọng nhất",
         "'Sở cứ' và 'Chặn việc nào'. Cột đầu chặn bảng biến thành danh sách "
         "những thứ người viết thấy hay; cột sau là thứ làm nó quản trị được "
         "thay vì chỉ đọc được."),
        ("Năm hạng sở cứ",
         "ĐO — chạy lại được trên chính kho này.  SỔ SAI LỆCH — lỗi đã xảy ra "
         "thật, có bài kiểm canh.  VĂN LIỆU — kết quả trong bài báo có "
         "ablation.  KHAI (đối thủ) — nhà cung cấp công bố.  SUY RA — lập "
         "luận của ta, chưa có số."),
        ("Cập nhật thế nào",
         "Sửa `scripts/lam_bang_tien_hoa.py` rồi chạy lại. Mỗi lần đổi trạng "
         "thái một việc, thêm một dòng vào sheet 'Nhật ký tiến hoá' — bảng "
         "không có nhật ký thì không ai kiểm lại được nó đã đi qua đâu."),
        ("", ""),
        ("Sinh lại", "python3 scripts/lam_bang_tien_hoa.py"),
        ("Đọc cùng",
         "docs/EAA_Benchmark_San_pham.docx · docs/V3_KET_QUA.md · "
         "docs/KE_HOACH_VUOT_LEN.md · docs/SAI_LECH_THIET_KE.md"),
    ]
    for r, (a, b) in enumerate(doc, 1):
        ws.cell(row=r, column=1, value=a).font = Font(bold=True, size=14 if r == 1 else 11)
        ws.cell(row=r, column=1).alignment = TREN
        ws.cell(row=r, column=2, value=b).alignment = TREN

    # ── 2 · lộ trình theo mảng ───────────────────────────────────────────────
    ws = wb.create_sheet("Lộ trình")
    dau(ws, [("Mảng", 26), ("Câu hỏi mảng này trả lời", 44),
             ("Vì sao mảng này có mặt", 52), ("Việc", 8), ("Xong", 7),
             ("Đang chờ", 9), ("Chưa", 7), ("Việc chặn nhiều nhất", 24)])
    for r, (ten, hoi, vi_sao) in enumerate(MANG, 2):
        cua = [v for v in VIEC if v[2] == ten]
        xong = sum(1 for v in cua if v[15].startswith("XONG"))
        cho = sum(1 for v in cua if v[15].startswith(("CHỜ", "CƠ CHẾ")))
        chan = max(cua, key=lambda v: 0 if v[11] == "—" else len(v[11].split(" · ")))
        so_chan = 0 if chan[11] == "—" else len(chan[11].split(" · "))
        gia_tri = (ten, hoi, vi_sao, len(cua), xong, cho,
                   len(cua) - xong - cho,
                   f"{chan[0]} — chặn {so_chan}" if so_chan else "—")
        for c, v in enumerate(gia_tri, 1):
            o = ws.cell(row=r, column=c, value=v)
            o.alignment = GIUA if c >= 4 else TREN
            o.border = VIEN
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=MAU_MANG[ten])
        ws.row_dimensions[r].height = 56

    # ── 3 · việc ─────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Việc")
    dau(ws, COT)
    for r, hang in enumerate(VIEC, 2):
        for c, v in enumerate(hang, 1):
            o = ws.cell(row=r, column=c, value=v)
            o.alignment = GIUA if c in (1, 2, 11, 12, 14, 15) else TREN
            o.border = VIEN
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor=MAU_MANG[hang[2]])
        tt = next((k for k in MAU_TT if hang[15].startswith(k)), "")
        ws.cell(row=r, column=16).fill = MAU_TT.get(tt, PatternFill())
        ws.row_dimensions[r].height = 120
    ws.auto_filter.ref = f"A1:P{len(VIEC) + 1}"

    # ── 4 và 5 · hai mảng đi sâu ────────────────────────────────────────────
    for ten_sheet, ma_mang, dan in (
        ("Mô phỏng", D,
         "Nền đã có: bộ điều khiển tham chiếu MiL bằng Python để quét tham số, "
         "và cổng SIL chạy FIRMWARE C THẬT qua lớp HAL giả lập rồi nối vào mô "
         "hình vật lý — mô hình ấy đã được kiểm chứng bằng nghiệm giải tích "
         "trước khi được làm cổng chặn. Chín việc dưới đây là phần THIẾU. "
         "Thiếu lớn nhất: cổng chỉ trả ĐẠT/TRƯỢT, không kịch bản nào tiêm lỗi, "
         "và không ai so số mô phỏng với số đo trên bo."),
        ("IDE", E,
         "Mảng này mô tả một mặt tiếp xúc CHƯA TỒN TẠI, nên mọi con số nghiệm "
         "thu là con số sẽ đo. Chỗ chặn thật rất cụ thể và rất rẻ: CLI có 103 "
         "lệnh, 0 lệnh thiếu trợ giúp, và 0 lệnh nào có đầu ra máy đọc được. "
         "Đổi lại, `ToolError` đã mang sẵn file/line/rule_id/severity — đủ cho "
         "một chẩn đoán trong biên tập, chỉ thiếu chỗ xuất ra. Luật xuyên suốt "
         "cả mảng: IDE gọi ĐÚNG các lệnh đã có, không dựng đường thứ hai tới "
         "merge."),
    ):
        ws = wb.create_sheet(ten_sheet)
        ws.cell(row=1, column=1, value=dan).alignment = TREN
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws.row_dimensions[1].height = 96
        cot = [("Mã", 6), ("Việc", 40), ("Sở cứ", 12),
               ("Làm như thế nào", 70), ("Nghiệm thu", 36),
               ("Rủi ro — cái dễ làm hỏng", 46)]
        for i, (t, rong) in enumerate(cot, 1):
            o = ws.cell(row=2, column=i, value=t)
            o.font, o.fill, o.alignment, o.border = DAM, NEN, GIUA, VIEN
            ws.column_dimensions[get_column_letter(i)].width = rong
        ws.freeze_panes = "A3"
        r = 3
        for v in (x for x in VIEC if x[2] == ma_mang):
            for c, gia in enumerate((v[0], v[3], v[5], v[6], v[8], v[12]), 1):
                o = ws.cell(row=r, column=c, value=gia)
                o.alignment = GIUA if c in (1, 3) else TREN
                o.border = VIEN
            ws.cell(row=r, column=1).font = Font(bold=True)
            ws.row_dimensions[r].height = 130
            r += 1

    # ── 6 · sở cứ ────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Sở cứ")
    dau(ws, [("Mã", 8), ("Hạng", 16), ("Sở cứ nói gì", 82),
             ("Tra ở đâu / chạy lại bằng gì", 52), ("Việc dùng nó", 20)])
    dung: dict[str, list[str]] = {}
    for v in VIEC:
        for m in v[5].split(", "):
            dung.setdefault(m, []).append(v[0])
    for r, (ma, hang, noi, tra) in enumerate(SO_CU, 2):
        for c, gia in enumerate(
            (ma, hang, noi, tra, ", ".join(dung.get(ma, ["—"]))), 1
        ):
            o = ws.cell(row=r, column=c, value=gia)
            o.alignment = GIUA if c in (1, 2, 5) else TREN
            o.border = VIEN
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=2).fill = MAU_HANG.get(hang, PatternFill())
        ws.row_dimensions[r].height = 56
    ws.auto_filter.ref = f"A1:E{len(SO_CU) + 1}"

    # ── 7 · phụ thuộc ────────────────────────────────────────────────────────
    ws = wb.create_sheet("Phụ thuộc")
    dau(ws, [("Việc", 8), ("Tên việc", 44), ("Chờ việc nào", 22),
             ("Chặn việc nào", 26), ("Số việc nó chặn", 14), ("Trạng thái", 18)])
    canh = sorted(VIEC, key=lambda v: -(0 if v[11] == "—" else len(v[11].split(" · "))))
    for r, v in enumerate(canh, 2):
        so = 0 if v[11] == "—" else len(v[11].split(" · "))
        for c, gia in enumerate((v[0], v[3], v[10], v[11], so, v[15]), 1):
            o = ws.cell(row=r, column=c, value=gia)
            o.alignment = GIUA if c in (1, 5) else TREN
            o.border = VIEN
        ws.cell(row=r, column=1).font = Font(bold=True)
        if so >= 3:
            ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor="FFC7CE")
        ws.row_dimensions[r].height = 40

    # ── 8 · nhật ký ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Nhật ký tiến hoá")
    dau(ws, [("Ngày", 14), ("Việc", 22), ("Đổi thành", 40), ("Ghi chú", 88)])
    for r, hang in enumerate(NHAT_KY, 2):
        for c, gia in enumerate(hang, 1):
            o = ws.cell(row=r, column=c, value=gia)
            o.alignment = GIUA if c == 1 else TREN
            o.border = VIEN
        ws.row_dimensions[r].height = 46

    # ── 9 · không nói gì ─────────────────────────────────────────────────────
    ws = wb.create_sheet("KHÔNG nói gì")
    dau(ws, [("#", 6), ("Điều bảng này KHÔNG chứng minh", 124)])
    for r, d in enumerate(KHONG_NOI, 2):
        ws.cell(row=r, column=1, value=r - 1).alignment = GIUA
        o = ws.cell(row=r, column=2, value=d)
        o.alignment, o.border = TREN, VIEN
        ws.row_dimensions[r].height = 52

    RA.parent.mkdir(parents=True, exist_ok=True)
    wb.save(RA)

    print(f"Đã ghi {RA}  ({RA.stat().st_size:,} byte)")
    print(f"  việc  : {len(VIEC)} — " + " · ".join(
        f"{t}: {sum(1 for v in VIEC if v[15].startswith(t))}"
        for t in ("XONG", "CƠ CHẾ XONG", "ĐANG LÀM", "CHỜ NGÂN SÁCH",
                  "CHỜ NGƯỜI", "CHƯA")))
    for ten, _, _ in MANG:
        print(f"    {ten:<28} {sum(1 for v in VIEC if v[2] == ten)} việc")
    hang_dem: dict[str, int] = {}
    for _, h, _, _ in SO_CU:
        hang_dem[h] = hang_dem.get(h, 0) + 1
    print(f"  sở cứ : {len(SO_CU)} — " + " · ".join(
        f"{k}: {v}" for k, v in hang_dem.items()))
    print(f"  sheet : {len(wb.sheetnames)} — {' · '.join(wb.sheetnames)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
